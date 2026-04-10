import matplotlib.pyplot as plt
import pandas as pd
import numpy as np
import traceback
import os
from datetime import datetime, timedelta
from dateutil.relativedelta import relativedelta
import logging
from logging import FileHandler
import matplotlib
from pathlib import Path
import base64
from analysis.defect_analyzer import analyze_flatness, analyze_warp, analyze_growing, analyze_broken, analyze_nano, analyze_pit, analyze_scratch, analyze_chip, analyze_edge, analyze_HUMAN_ERR, analyze_VISUAL, analyze_NOSALE, analyze_OTHER, analyze_GR, analyze_sample,analyze_particle
from config.mappings import REJ_GROUP_TO_MID_MAPPING, NAME_TO_EQP, MID_TO_EQP
import tempfile
from inspect import signature
import re
from openpyxl import Workbook
from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from matplotlib.ticker import PercentFormatter
from decimal import Decimal
import textwrap


# =========================================================
# Excel Export Utilities (공통화: 이미지/표/시트 빌더)
# - PNG 파일 I/O 최소화: (가능하면) 메모리 이미지로 처리
# - DataFrame 표 생성 공통화
# - 커서 기반 SheetBuilder 제공
# =========================================================
from io import BytesIO

try:
    from PIL import Image as PILImage  # optional (권장)
except Exception:  # pragma: no cover
    PILImage = None


# ---- 공통 스타일 상수 (재사용) ----
_THIN = Side(style="thin")
_BORDER_THIN = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)

HEADER_FONT = Font(bold=True, size=10)
BODY_FONT = Font(size=9)
CENTER_WRAP = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)

HEADER_FILL = PatternFill("solid", fgColor="D3D3D3")

# 한글 폰트 설정
matplotlib.rcParams['font.family'] = 'Malgun Gothic'  # Windows
matplotlib.rcParams['font.size'] = 10
matplotlib.rcParams['axes.unicode_minus'] = False  # 마이너스 기호 깨짐 방지

# 결과 저장 폴더
REPORT_DIR = "./daily_reports_debug"
os.makedirs(REPORT_DIR, exist_ok=True)

# 기존 로거 설정 대체
logger = logging.getLogger(__name__)
logger.setLevel(logging.INFO)

# 기존 핸들러 제거
if logger.hasHandlers():
    logger.handlers.clear()

# UTF-8로 기록하는 FileHandler 추가
file_handler = FileHandler('daily_report.log', encoding='utf-8')
formatter = logging.Formatter('%(asctime)s - %(name)s - %(levelname)s - %(message)s')
file_handler.setFormatter(formatter)
logger.addHandler(file_handler)

class DailyReportGenerator:
    _ms6_mapping_cache = None

    def __init__(self, data, target_date=None):
        self.data = data
        self.target_date = target_date or (datetime.now() - timedelta(days=1)).strftime('%Y%m%d')
        self.target_date_obj = datetime.strptime(self.target_date, '%Y%m%d').date()

    def _calculate_total_loss_influence(self, df):
        """
        전체 불량 기준 제품 영향성 분석
        - REJ_GROUP != '분모' 인 모든 LOSS_QTY 합산
        - PRODUCT_TYPE별 전체 불량률 계산
        """

        # ──────────────────────────────────────────────────
        # [신규] PRODUCT_TYPE 그룹화: [P]SEC F3 & [P]SEC UB → '[P]SEC F3/UB'
        # ──────────────────────────────────────────────────
        def group_product_type(pt):
            if pd.isna(pt):
                return "Unknown"
            pt = str(pt).strip()
            if pt in ['[P]SEC F3', '[P]SEC UB']: #[P]SEC F3 과 [P]SEC UB
                return '[P]SEC F3/UB'
            return pt

        df['PRODUCT_TYPE'] = df['PRODUCT_TYPE'].apply(group_product_type)
        # ──────────────────────────────────────────────────

        if df is None or getattr(df, "empty", True):
            return pd.DataFrame(), 0

        df = df.copy()
        df = df[df['GRD_CD_NM_CS'] == 'Prime']

        # 전체 불량: '분모' 제외 모든 불량
        df_total = df[df['REJ_GROUP'] != '분모']

        for col in ["IN_QTY", "LOSS_QTY"]:
            if col in df.columns:
                df[col] = pd.to_numeric(df[col], errors="coerce")
            else:
                df[col] = 0

        # 전체 수량
        denominator = df[df['REJ_GROUP'] == '분모']
        total_qty = denominator['IN_QTY'].sum()

        df_total = df[df['REJ_GROUP']!= '분모']

        loss_summary = (
            df_total.groupby("PRODUCT_TYPE", dropna=False)["LOSS_QTY"]
            .sum()
            .reset_index()
        )
        loss_summary.rename(columns={"LOSS_QTY": "전체_불량개수"}, inplace=True)

        # Compile 수량
        compile_summary = (
            df[df['REJ_GROUP'] == '분모']
            .groupby("PRODUCT_TYPE", dropna=False)["IN_QTY"]
            .sum()
            .reset_index()
        )
        compile_summary.rename(columns={"IN_QTY": "Compile_수량"}, inplace=True)
 

        # 병합
        result = pd.merge(loss_summary, compile_summary, on="PRODUCT_TYPE", how="outer")
        result["전체_불량개수"] = result["전체_불량개수"].fillna(0.0)
        result["Compile_수량"] = result["Compile_수량"].fillna(0.0)

        # 물량비
        result["물량비(%)"] = (
            (result["Compile_수량"] / total_qty * 100) if total_qty > 0 else 0.0 #compile수량/전체수량 * 100 -> 이미 %로 변경함.
        ).round(2)


        # 제품별 전체 불량률
        result["전체_불량률(%)"] = np.where(
            result["Compile_수량"] > 0,
            (result["전체_불량개수"] / result["Compile_수량"] ) * 100, # 각 제품별 불량률, 제품 불량 개수/ 제품 총 in 수량
            0.0
        ).round(2)

        result = result[[
            "PRODUCT_TYPE", "전체_불량개수", "Compile_수량", "전체_불량률(%)", "물량비(%)"
        ]].sort_values("전체_불량률(%)", ascending=False).reset_index(drop=True)

        return result, total_qty

   
    # ──────────────────────────────────────────────────
    # [신규] MS6.csv 기반 제품 정보 병합 함수
    # ──────────────────────────────────────────────────
    @classmethod
    def _load_ms6_mapping(cls):
        """MS6.csv에서 (MS6 -> 제품1) 매핑을 1회만 로드하여 캐시"""
        if cls._ms6_mapping_cache is not None:
            return cls._ms6_mapping_cache

        project_root = Path(__file__).parent.parent
        ms6_path = project_root / "queries" / "MS6.csv"

        if not ms6_path.exists():
            cls._ms6_mapping_cache = {}
            return cls._ms6_mapping_cache

        try:
            df_ms6 = pd.read_csv(ms6_path, dtype=str, encoding="utf-8")
        except UnicodeDecodeError:
            df_ms6 = pd.read_csv(ms6_path, dtype=str, encoding="cp949")

        if "MS6" not in df_ms6.columns or "제품1" not in df_ms6.columns:
            cls._ms6_mapping_cache = {}
            return cls._ms6_mapping_cache

        df_ms6 = df_ms6.dropna(subset=["MS6", "제품1"]).copy()
        cls._ms6_mapping_cache = dict(
            zip(df_ms6["MS6"].astype(str).str.strip(), df_ms6["제품1"].astype(str).str.strip())
        )
        return cls._ms6_mapping_cache

    def _merge_product_type(self, df):
        """df에 PROD_ID 기준으로 MS6.csv의 '제품1' 컬럼을 병합하여 'PRODUCT_TYPE' 추가 (캐시 적용)"""
        if df is None or getattr(df, "empty", True):
            # empty or None
            try:
                df = df.copy()
            except Exception:
                df = pd.DataFrame()
            df["PRODUCT_TYPE"] = "Unknown"
            return df

        if "PROD_ID" not in df.columns:
            df = df.copy()
            df["PRODUCT_TYPE"] = "Unknown"
            return df

        mapping = self._load_ms6_mapping()
        df = df.copy()
        df["MS6"] = df["PROD_ID"].astype(str).str[:6]
        df["PRODUCT_TYPE"] = df["MS6"].map(mapping).fillna("Unknown")
        return df

    def _get_top3_rej_groups(self):
        """
        안전하게 상위 3개 REJ_GROUP 목록 가져오기
        """
        return self.data.get('DATA_3210_wafering_300', {}).get('top3_rej_groups', [])

    def _create_total_loss_ref(self):
        """6개월 전체 불량 기반 Ref 데이터 생성 (현재 날짜 기준 직전 6개월)"""
        PROJECT_ROOT = Path(__file__).parent.parent
        cache_dir = PROJECT_ROOT / "data_cache"
        pattern = "DATA_LOT_3210_wafering_300_*.parquet"
        parquet_files = list(cache_dir.glob(pattern))

        # 현재 날짜 기준으로 직전 6개월의 연월 생성 (YYYYMM 형식)
        current_date =  self.target_date_obj
        target_months = []
        for i in range(6, 0, -1):  # 최근 6개월: -1개월, -2개월, ..., -6개월
            month_date = current_date - relativedelta(months=i)
            target_months.append(month_date.strftime("%Y%m"))
        print(f"target_months:", target_months)
        df_list = []
        for file_path in parquet_files:
            try:
                stem = file_path.stem
                date_part = stem.split("_")[-1]
                if len(date_part) == 6 and date_part.isdigit() and date_part in target_months:
                    df_part = pd.read_parquet(file_path)
                    df_part = self._merge_product_type(df_part)
                    df_list.append(df_part)
            except Exception as e:
                continue  # 로그 추가를 원하면 logger.warning 사용

        if not df_list:
            return pd.DataFrame(), 0

        df_full = pd.concat(df_list, ignore_index=True)

        return self._calculate_total_loss_influence(df_full)








    # def _create_total_loss_ref(self):
    #     """6개월 전체 불량 기반 Ref 데이터 생성"""
    #     PROJECT_ROOT = Path(__file__).parent.parent
    #     cache_dir = PROJECT_ROOT / "data_cache"
    #     pattern = "DATA_LOT_3210_wafering_300_*.parquet"
    #     parquet_files = list(cache_dir.glob(pattern))

    #     target_months = [f"2025{str(m).zfill(2)}" for m in range(6, 13)]

    #     df_list = []
    #     for file_path in parquet_files:
    #         try:
    #             stem = file_path.stem
    #             date_part = stem.split("_")[-1]
    #             if len(date_part) == 6 and date_part.isdigit() and date_part in target_months:
    #                 df_part = pd.read_parquet(file_path)
    #                 df_part = self._merge_product_type(df_part)
    #                 df_list.append(df_part)
    #         except Exception as e:
    #             continue

    #     if not df_list:
    #         return pd.DataFrame(), 0

    #     df_full = pd.concat(df_list, ignore_index=True)

    #     return self._calculate_total_loss_influence(df_full)

    def _create_total_loss_daily(self):
        """금일 전체 불량 기반 Daily 데이터 생성"""
        key = "DATA_LOT_3210_wafering_300"
        if key not in self.data or self.data[key].empty:
            return pd.DataFrame(), 0

        df = self.data[key].copy()
        if "PRODUCT_TYPE" not in df.columns:
            df = self._merge_product_type(df)
            if "PRODUCT_TYPE" not in df.columns:
                return pd.DataFrame(), 0

        return self._calculate_total_loss_influence(df)


    def _analyze_total_loss_gap(self):
        """
        전체 불량률 기준 GAP 분석
        - Daily 상위 3개 제품 기준 → Ref 비교
        """
        if 'total_loss_ref' not in self.data or 'total_loss_daily' not in self.data:
            return pd.DataFrame()

        ref_df = self.data['total_loss_ref']
        daily_df = self.data['total_loss_daily']

        if ref_df.empty or daily_df.empty:
            return pd.DataFrame()

        # 전체 평균 불량률 
        total_ref_qty = self.data.get('total_loss_ref_total_qty', 1)
        total_daily_qty = self.data.get('total_loss_daily_total_qty', 1)

        total_ref_loss = ref_df['전체_불량개수'].sum()
        total_daily_loss = daily_df['전체_불량개수'].sum()

        overall_ref_loss_rate = (total_ref_loss / total_ref_qty * 100) if total_ref_qty > 0 else 0.0
        overall_daily_loss_rate = (total_daily_loss / total_daily_qty * 100) if total_daily_qty > 0 else 0.0

        gap_list = []
        all_products = daily_df['PRODUCT_TYPE'].unique().tolist()

        # 전체 제품에 대해 GAP 계산
        for pt in all_products:
            row = {'PRODUCT_TYPE': pt}

            # Daily 데이터
            daily_row = daily_df[daily_df['PRODUCT_TYPE'] == pt].iloc[0]
            row.update({
                'Daily_전체_불량률(%)': daily_row['전체_불량률(%)'],
                'Daily_Compile_수량': daily_row['Compile_수량'],
                'Daily_물량비(%)': daily_row['물량비(%)']
            })

            # Ref 데이터
            ref_row = ref_df[ref_df['PRODUCT_TYPE'] == pt]
            if not ref_row.empty:
                ref_val = ref_row.iloc[0]
                row['Ref_전체_불량률(%)'] = ref_val['전체_불량률(%)']
                row['Ref_Compile_수량'] = ref_val['Compile_수량']
                row['Ref_물량비(%)'] = ref_val['물량비(%)']
            else:
                row['Ref_전체_불량률(%)'] = 0.0
                row['Ref_Compile_수량'] = 0
                row['Ref_물량비(%)'] = 0.0

            # GAP 계산
            row['전체_불량률_GAP(%)'] = (row['Daily_전체_불량률(%)'] - row['Ref_전체_불량률(%)']).round(2)
            row['물량비_GAP(%)'] = (row['Daily_물량비(%)'] - row['Ref_물량비(%)']).round(2)

            #  제품 Mix비 변동 계산
            row['제품 Mix비 변동'] = (
                (row['Ref_전체_불량률(%)'] - overall_ref_loss_rate) * row['물량비_GAP(%)'] / 100.0
            ).round(4)

            gap_list.append(row)

        #전체 제품의 제품 Mix비 변동 합계 계산
        total_volume_defect_change = sum(item['제품 Mix비 변동'] for item in gap_list)   

        # 전체 결과 DataFrame 생성
        result = pd.DataFrame(gap_list)
        if result.empty or 'PRODUCT_TYPE' not in result.columns:
            return pd.DataFrame(), {}

        # 제품 Mix비 변동 기준 내림차순 정렬 → 상위 3개
        # result = result.sort_values('제품 Mix비 변동', key=abs, ascending=False).head(3) #절대값 기준으로 변경

        positive_top3 = result[result['제품 Mix비 변동'] > 0].sort_values('제품 Mix비 변동', ascending=False).head(3) # 양수 데이터: 0보다 큰 값 중에서 내림차순 정렬 후 상위 3개
        negative_top3 = result[result['제품 Mix비 변동'] < 0].sort_values('제품 Mix비 변동', ascending=True).head(3) # 음수 데이터: 0보다 작은 값 중에서 오름차순 정렬 (가장 작은 음수 = 가장 큰 감소) 후 상위 3개

        result = pd.concat([positive_top3, negative_top3], ignore_index=True)


        # 전체 평균 통계
        overall_stats = {
            'total_ref_qty': total_ref_qty,
            'total_daily_qty': total_daily_qty,
            'overall_ref_loss_rate': overall_ref_loss_rate,
            'overall_daily_loss_rate': overall_daily_loss_rate,
            'total_ref_loss': total_ref_loss,
            'total_daily_loss': total_daily_loss,
            'total_volume_defect_change' : total_volume_defect_change
        }

        return result.reset_index(drop=True), overall_stats

    def generate(self):
        """데일리 리포트 생성"""
        print(f"🔍 self.data 키 목록: {list(self.data.keys())}")
        if 'DATA_3010_epi_300' in self.data:
            print(f"📊 EPI 데이터 건수: {len(self.data['DATA_3010_epi_300'])}")
        else:
            print("🚨 DATA_3010_epi_300 키 없음 → 외부에서 추가 필요")

        try:
            logger.info("리포트 생성 시작")
            # ==================================================================
            # 모든 데이터에 PRODUCT_TYPE 일괄 병합 (가장 먼저 실행)
            # ===================================================================

            for key in ['DATA_LOT_3210_wafering_300', 'DATA_WAF_3210_wafering_300']:
                if key in self.data and not self.data[key].empty:
                    self.data[key] = self._merge_product_type(self.data[key])
                    if 'PRODUCT_TYPE' in self.data[key].columns:
                        sample = self.data[key].sample(1)[['PROD_ID', 'PRODUCT_TYPE']].to_dict('records')
                else:
                    print(f"{key} 없거나 빈 데이터")

            # 3010 보고서 생성 (WF)
            data_3010_details = self._create_3010_wafering_300()

            # 3010 보고서 생성 (EPI)
            # data_3010_epi_details = self._create_3010_epi_300()

            # 1. DATA_3210_wafering_300 생성 + 저장 (먼저!)
            data_3210_details = self._create_DATA_3210_wafering_300()
            self.data['DATA_3210_wafering_300'] = data_3210_details

            # 2. DATA_3210_wafering_300_3months 생성 + 저장 (핵심!)
            data_3210_3months = self._create_DATA_3210_wafering_300_3months()
            self.data['DATA_3210_wafering_300_3months'] = data_3210_3months  

            # STEP 1: 전체 불량 기반 분석 먼저
            total_loss_ref_df, total_loss_ref_qty = self._create_total_loss_ref()
            total_loss_daily_df, total_loss_daily_qty = self._create_total_loss_daily()

            self.data['total_loss_ref'] = total_loss_ref_df
            self.data['total_loss_daily'] = total_loss_daily_df
            self.data['total_loss_ref_total_qty'] = total_loss_ref_qty
            self.data['total_loss_daily_total_qty'] = total_loss_daily_qty

            total_loss_gap, overall_stats  = self._analyze_total_loss_gap()
            self.data['overall_stats'] = overall_stats

            data_lot_details = self._create_DATA_LOT_3210_wafering_300()
            data_waf_details = self._create_DATA_WAF_3210_wafering_300()

            # DATA_1511_SMAX_wafering_300 = self.data.get('DATA_1511_SMAX_wafering_300')
            self._create_DATA_1511_SMAX_wafering_300()
            self._calculate_loss_rate_by_process() # LOSS_RATE 계산(smax 1511 보고서 사용)

            self._plot_rej_group_top3_eqp_trend(output_dir="./daily_reports_debug")

            report = {
                'DATA_3010_wafering_300' : data_3010_details,
                'DATA_3210_wafering_300_details': data_3210_details,
                'DATA_3210_wafering_300_3months': data_3210_3months,
                'DATA_LOT_3210_wafering_300_details': data_lot_details,
                'DATA_WAF_3210_wafering_300_details': data_waf_details,
                'total_loss_gap' : total_loss_gap,
                'overall_stats': overall_stats,
                'raw_data': self.data
            }
            
            # Excel 생성 시 report 전체 전달
            try:
                excel_path = self._export_to_excel(report, output_dir="./daily_reports_debug")
                report['excel_report'] = str(excel_path)
                print(f"Excel 보고서도 생성됨: {excel_path}")
            except Exception as e:
                print(f"Excel 생성 실패: {e}")
                report['excel_report'] = None

            logger.info("리포트 생성 완료")
            return report
        except Exception as e:
            logger.error(f"리포트 생성 실패: {e}")
            raise
    

    def _create_3010_wafering_300(self):
        """3010 수율 데이터 분석 및 그래프 생성 (WF RTY만, 최신 일실적 기준)"""
        details = {}

        if 'DATA_3010_wafering_300' not in self.data or self.data['DATA_3010_wafering_300'].empty:
            print("DATA_3010_wafering_300 데이터 없음 또는 비어 있음")
            return details

        df = self.data['DATA_3010_wafering_300'].copy()

        # --- 전처리 ---
        df['rate'] = pd.to_numeric(df['rate'], errors='coerce')
        df['item_type'] = df['item_type'].astype(str).str.strip()
        df['dt_range_raw'] = df['dt_range'].astype(str).str.strip() # dt_range_raw: 문자열 정리

        # grade, yld_type 컬럼 처리
        if 'grade' not in df.columns:
            df['grade'] = 'Total'
        if 'yld_type' not in df.columns:
            df['yld_type'] = 'RTY'
        df['grade'] = df['grade'].astype(str).str.strip()
        df['yld_type'] = df['yld_type'].astype(str).str.strip()

        # item_type에 따라 파싱 전략 분기
        def parse_date(row):
            raw = row['dt_range_raw']
            item_type = row['item_type']
            
            # 빈 값 또는 이상한 값 필터링
            if pd.isna(raw) or raw in ['', 'None', 'nan', 'NaT']:
                return pd.NaT

            try:
                if item_type in ['월실적', '월사업계획']:
                    return pd.to_datetime(raw, format='%Y-%m', errors='coerce')
                else:
                    return pd.to_datetime(raw, errors='coerce')
            except:
                return pd.NaT

        df['dt_range'] = df.apply(parse_date, axis=1)

        # month_str 생성
        df['month_str'] = df['dt_range'].dt.strftime('%Y-%m')
        current_month = (datetime.now() - timedelta(days=1)).strftime('%Y-%m')
        # ──────────────────────────────────────────────────
        # 기준일: 어제
        # ──────────────────────────────────────────────────
        # target_date = (datetime.now().date() - timedelta(days=1))  # 2026-02-03
        target_date = self.target_date
        print(f"기준일: {target_date}")
        # ──────────────────────────────────────────────────
        # 3. 재사용 함수 정의
        # ──────────────────────────────────────────────────
        def get_latest_or_target(df, item_type, target_date):
            print(f"[DEBUG] 조회: item_type={item_type}, target_date={target_date}")
            
            # 동일 날짜 찾기
            same_day = df[
                (df['item_type'] == item_type) &
                (df['dt_range'].notna()) &
                (df['dt_range'].dt.date == target_date)
            ]
            if not same_day.empty:
                return same_day.iloc[0]

            # 없으면 최신 날짜 사용
            latest = df[
                (df['item_type'] == item_type) &
                (df['dt_range'].notna())
            ]
            if not latest.empty:
                latest_sorted = latest.sort_values('dt_range', ascending=False)
                return latest_sorted.iloc[0]
            
            return None

        # 월 데이터 추출 함수
        def get_monthly_value(df, item_type, current_month):
            monthly = df[
                (df['item_type'] == item_type) &
                (df['month_str'] == current_month)
            ]
            return float(monthly['rate'].iloc[0]) if not monthly.empty else 0.0

        # ──────────────────────────────────────────────────
        # 4개 조합 (Total/Prime × RTY/OAY) 데이터 추출
        # ──────────────────────────────────────────────────
        grades = ['Total', 'Prime']
        yld_types = ['RTY', 'OAY']
        results = {}
        daily_actual_date = "N/A"

        for grade in grades:
            for yld_type in yld_types:
                key = f"{grade}_{yld_type}"
                df_filtered = df[(df['grade'] == grade) & (df['yld_type'] == yld_type)].copy()
                
                # 월 목표/실적
                monthly_plan_val = get_monthly_value(df_filtered, '월사업계획', current_month)
                monthly_actual_val = get_monthly_value(df_filtered, '월실적', current_month)
                
                # 일 목표/실적
                daily_plan_row = get_latest_or_target(df_filtered, '일사업계획', target_date)
                daily_actual_row = get_latest_or_target(df_filtered, '일실적', target_date)
                
                daily_plan_val = float(daily_plan_row['rate']) if daily_plan_row is not None else 0.0
                if daily_actual_row is not None:
                    daily_actual_val = float(daily_actual_row['rate'])
                    daily_actual_date = daily_actual_row['dt_range'].strftime('%Y-%m-%d')
                else:
                    daily_actual_val = 0.0
                
                results[key] = {
                    'grade': grade,
                    'yld_type': yld_type,
                    'monthly_plan': float(monthly_plan_val),
                    'monthly_actual': float(monthly_actual_val),
                    'daily_plan': float(daily_plan_val),
                    'daily_actual': float(daily_actual_val),
                    'monthly_gap': float(monthly_actual_val - monthly_plan_val),
                    'daily_gap': float(daily_actual_val - daily_plan_val)
                }

        # ──────────────────────────────────────────────────
        # 4. 그래프 생성
        # ──────────────────────────────────────────────────
        # PROJECT_ROOT 및 날짜 폴더
        PROJECT_ROOT = Path(__file__).parent.parent
        # base_date = (datetime.now().date() - timedelta(days=1))
        # date_folder_name = base_date.strftime("%Y%m%d")
        # debug_dir = PROJECT_ROOT / "daily_reports_debug" / date_folder_name
        debug_dir = PROJECT_ROOT / "daily_reports_debug" / self.target_date
        debug_dir.mkdir(exist_ok=True, parents=True)

        chart_path = debug_dir / "3010_yield_WF_chart.png"

        if chart_path.exists():
            chart_path.unlink() #파일 삭제
            print(f"기존 그래프 파일 삭제됨 : {chart_path}")

        fig, axes = plt.subplots(1, 4, figsize=(16, 4), dpi=300)
        axes = axes.flatten()

        plot_order = ['Total_RTY', 'Prime_RTY', 'Total_OAY', 'Prime_OAY']
        titles = ['Total RTY', 'Prime RTY', 'Total OAY', 'Prime OAY']

        for idx, key in enumerate(plot_order):
            ax = axes[idx]
            data = results[key]

            # X축: 월, 일
            x_labels = ['월', '일']
            x = np.arange(len(x_labels))
            bar_width = 0.9
            month_color = '#0000ff'  # 파랑
            day_color = '#ff0000'    # 빨강

            # 월/일 각각 하나의 막대 (실적만), 목표는 점선
            bar_month = ax.bar(x[0], data['monthly_actual'], bar_width, color=month_color)
            bar_day = ax.bar(x[1], data['daily_actual'], bar_width, color=day_color)

            # 목표선 (공통 목표 = 일사업계획)
            ax.axhline(y=data['daily_plan'], color='black', linestyle='--', linewidth=1.2, label='목표')

            # X축 레이블 설정
            ax.set_xticks(x)
            ax.set_xticklabels(x_labels, fontsize=15, fontweight='bold')
            ax.set_xlabel('기간', fontsize=15)

            # Y축 범위
            all_vals = [data['monthly_actual'], data['daily_actual'], data['daily_plan']]
            min_ylim = min(all_vals)
            max_ylim = max(all_vals)
            ax.set_ylim(max(90, min_ylim - 1.0), max_ylim + 3.0)
            ylim_bottom = ax.get_ylim()[0]

            # 제목
            ax.set_title(titles[idx], fontsize=20, fontweight='bold', pad=14)
            ax.set_ylabel('수율 (%)', fontsize=15)

            # 막대 내부에 텍스트
            def add_internal_label(bar, value):
                for rect in bar:
                    height = rect.get_height()
                    pos = (ylim_bottom + height) / 2.0
                    ax.text(rect.get_x() + rect.get_width()/2., pos,
                            f'{value:.2f}%', ha='center', va='center',
                            fontsize=15, fontweight='bold', color='white')

            add_internal_label(bar_month, data['monthly_actual'])
            add_internal_label(bar_day, data['daily_actual'])

            # Gap 라벨 (막대 위, 색상 구분)
            def add_gap_label(x_pos, gap, base_height):
                color = 'blue' if gap >= 0 else 'red'
                sign = '+' if gap >= 0 else ''
                ax.text(x_pos, base_height + 0.3,
                        f'{sign}{gap:.2f}%', ha='center', va='bottom',
                        fontsize=15, fontweight='bold', color=color)

            add_gap_label(x[0], data['monthly_gap'], data['monthly_actual'])
            add_gap_label(x[1], data['daily_gap'], data['daily_actual'])

            # 범례 (상단 왼쪽)
            ax.legend(loc='upper center', ncol=1, frameon=False, fontsize=15, bbox_to_anchor=(0.5, 0.98))

            # 그리드
            ax.grid(axis='y', linestyle='--', alpha=0.7, zorder=0)
            ax.set_axisbelow(True)
            ax.spines['top'].set_visible(False)
            ax.spines['right'].set_visible(False)

        plt.tight_layout(pad=1.0)
        plt.savefig(chart_path, dpi=300, bbox_inches='tight', pad_inches=0.5)
        plt.close()

        # Base64 인코딩
        with open(chart_path, "rb") as img_file:
            img_base64 = base64.b64encode(img_file.read()).decode()
        # ──────────────────────────────────────────────────
        # 5. 표 생성 (계층형 DataFrame)
        # ──────────────────────────────────────────────────
        table_rows = []
        for yld_type in ['RTY', 'OAY']:
            for grade in ['Total', 'Prime']:
                key = f"{grade}_{yld_type}"
                data = results[key]
                table_rows.append({
                    'yld_type': yld_type,
                    'grade': grade,
                    'monthly_plan': data['monthly_plan'],
                    'monthly_actual': data['monthly_actual'],
                    'monthly_gap': data['monthly_gap'],
                    'daily_plan': data['daily_plan'],
                    'daily_actual': data['daily_actual'],
                    'daily_gap': data['daily_gap']
                })

        table_df = pd.DataFrame(table_rows)

        details.update({
            'chart_path': str(chart_path),
            'img_base64': img_base64,
            'table_df': table_df,
            'summary': table_df,
            'daily_actual_date': daily_actual_date,
            'results': results
        })

        return details


    # def _create_3010_epi_300(self):
    #     """3010 EPI 수율 데이터 분석 및 그래프 생성 (WF RTY만, 최신 일실적 기준)"""
    #     details = {}

    #     if 'DATA_3010_epi_300' not in self.data or self.data['DATA_3010_epi_300'].empty:
    #         print("DATA_3010_epi_300 데이터 없음 또는 비어 있음")
    #         return details

    #     df = self.data['DATA_3010_epi_300'].copy()

    #     # --- 전처리 ---
    #     df['rate'] = pd.to_numeric(df['rate'], errors='coerce')
    #     df['item_type'] = df['item_type'].astype(str).str.strip()
    #     df['dt_range_raw'] = df['dt_range'].astype(str).str.strip() # dt_range_raw: 문자열 정리

    #     # grade, yld_type 컬럼 처리
    #     if 'grade' not in df.columns:
    #         df['grade'] = 'Total'
    #     if 'yld_type' not in df.columns:
    #         df['yld_type'] = 'RTY'
    #     df['grade'] = df['grade'].astype(str).str.strip()
    #     df['yld_type'] = df['yld_type'].astype(str).str.strip()

    #     # item_type에 따라 파싱 전략 분기
    #     def parse_date(row):
    #         raw = row['dt_range_raw']
    #         item_type = row['item_type']
            
    #         # 빈 값 또는 이상한 값 필터링
    #         if pd.isna(raw) or raw in ['', 'None', 'nan', 'NaT']:
    #             return pd.NaT

    #         try:
    #             if item_type in ['월실적', '월사업계획']:
    #                 return pd.to_datetime(raw, format='%Y-%m', errors='coerce')
    #             else:
    #                 return pd.to_datetime(raw, errors='coerce')
    #         except:
    #             return pd.NaT

    #     df['dt_range'] = df.apply(parse_date, axis=1)

    #     # month_str 생성
    #     df['month_str'] = df['dt_range'].dt.strftime('%Y-%m')
    #     current_month = (datetime.now() - timedelta(days=1)).strftime('%Y-%m')
    #     # ──────────────────────────────────────────────────
    #     # 기준일: 어제
    #     # ──────────────────────────────────────────────────
    #     target_date = (datetime.now().date() - timedelta(days=1))  # 2026-02-03
    #     print(f"기준일: {target_date}")
    #     # ──────────────────────────────────────────────────
    #     # 3. 재사용 함수 정의
    #     # ──────────────────────────────────────────────────
    #     def get_latest_or_target(df, item_type, target_date):
    #         print(f"[DEBUG] 조회: item_type={item_type}, target_date={target_date}")
            
    #         # 동일 날짜 찾기
    #         same_day = df[
    #             (df['item_type'] == item_type) &
    #             (df['dt_range'].notna()) &
    #             (df['dt_range'].dt.date == target_date)
    #         ]
    #         if not same_day.empty:
    #             return same_day.iloc[0]

    #         # 없으면 최신 날짜 사용
    #         latest = df[
    #             (df['item_type'] == item_type) &
    #             (df['dt_range'].notna())
    #         ]
    #         if not latest.empty:
    #             latest_sorted = latest.sort_values('dt_range', ascending=False)
    #             return latest_sorted.iloc[0]
            
    #         return None

    #     # 월 데이터 추출 함수
    #     def get_monthly_value(df, item_type, current_month):
    #         monthly = df[
    #             (df['item_type'] == item_type) &
    #             (df['month_str'] == current_month)
    #         ]
    #         return float(monthly['rate'].iloc[0]) if not monthly.empty else 0.0

    #     # ──────────────────────────────────────────────────
    #     # 4개 조합 (Total/Prime × RTY/OAY) 데이터 추출
    #     # ──────────────────────────────────────────────────
    #     grades = ['Total', 'Prime']
    #     yld_types = ['RTY']
    #     results = {}
    #     daily_actual_date = "N/A"

    #     for grade in grades:
    #         for yld_type in yld_types:
    #             key = f"{grade}_{yld_type}"
    #             df_filtered = df[(df['grade'] == grade) & (df['yld_type'] == yld_type)].copy()
                
    #             # 월 목표/실적
    #             monthly_plan_val = get_monthly_value(df_filtered, '월사업계획', current_month)
    #             monthly_actual_val = get_monthly_value(df_filtered, '월실적', current_month)
                
    #             # 일 목표/실적
    #             daily_plan_row = get_latest_or_target(df_filtered, '일사업계획', target_date)
    #             daily_actual_row = get_latest_or_target(df_filtered, '일실적', target_date)
                
    #             daily_plan_val = float(daily_plan_row['rate']) if daily_plan_row is not None else 0.0
    #             if daily_actual_row is not None:
    #                 daily_actual_val = float(daily_actual_row['rate'])
    #                 daily_actual_date = daily_actual_row['dt_range'].strftime('%Y-%m-%d')
    #             else:
    #                 daily_actual_val = 0.0
                
    #             results[key] = {
    #                 'grade': grade,
    #                 'yld_type': yld_type,
    #                 'monthly_plan': float(monthly_plan_val),
    #                 'monthly_actual': float(monthly_actual_val),
    #                 'daily_plan': float(daily_plan_val),
    #                 'daily_actual': float(daily_actual_val),
    #                 'monthly_gap': float(monthly_actual_val - monthly_plan_val),
    #                 'daily_gap': float(daily_actual_val - daily_plan_val)
    #             }

    #     # ──────────────────────────────────────────────────
    #     # 4. 그래프 생성
    #     # ──────────────────────────────────────────────────
    #     # PROJECT_ROOT 및 날짜 폴더
    #     PROJECT_ROOT = Path(__file__).parent.parent
    #     base_date = (datetime.now().date() - timedelta(days=1))
    #     date_folder_name = base_date.strftime("%Y%m%d")
    #     debug_dir = PROJECT_ROOT / "daily_reports_debug" / date_folder_name
    #     debug_dir.mkdir(exist_ok=True, parents=True)

    #     chart_path = debug_dir / "3010_yield_chart_epi.png"

    #     if chart_path.exists():
    #         chart_path.unlink() #파일 삭제
    #         print(f"기존 EPI 그래프 파일 삭제됨 : {chart_path}")

    #     fig, axes = plt.subplots(1, 4, figsize=(16, 4), dpi=300)
    #     axes = axes.flatten()

    #     plot_order = ['Total_RTY', 'Prime_RTY']
    #     titles = ['Total RTY', 'Prime RTY']

    #     for idx, key in enumerate(plot_order):
    #         ax = axes[idx]
    #         data = results[key]

    #         # X축: 월, 일
    #         x_labels = ['월', '일']
    #         x = np.arange(len(x_labels))
    #         bar_width = 0.9
    #         month_color = '#0000ff'  # 파랑
    #         day_color = '#ff0000'    # 빨강

    #         # 월/일 각각 하나의 막대 (실적만), 목표는 점선
    #         bar_month = ax.bar(x[0], data['monthly_actual'], bar_width, color=month_color)
    #         bar_day = ax.bar(x[1], data['daily_actual'], bar_width, color=day_color)

    #         # 목표선 (공통 목표 = 일사업계획)
    #         ax.axhline(y=data['daily_plan'], color='black', linestyle='--', linewidth=1.2, label='목표')

    #         # X축 레이블 설정
    #         ax.set_xticks(x)
    #         ax.set_xticklabels(x_labels, fontsize=15, fontweight='bold')
    #         ax.set_xlabel('기간', fontsize=15)

    #         # Y축 범위
    #         all_vals = [data['monthly_actual'], data['daily_actual'], data['daily_plan']]
    #         min_ylim = min(all_vals)
    #         max_ylim = max(all_vals)
    #         ax.set_ylim(max(90, min_ylim - 1.0), max_ylim + 3.0)
    #         ylim_bottom = ax.get_ylim()[0]

    #         # 제목
    #         ax.set_title(titles[idx], fontsize=20, fontweight='bold', pad=14)
    #         ax.set_ylabel('수율 (%)', fontsize=15)

    #         # 막대 내부에 텍스트
    #         def add_internal_label(bar, value):
    #             for rect in bar:
    #                 height = rect.get_height()
    #                 pos = (ylim_bottom + height) / 2.0
    #                 ax.text(rect.get_x() + rect.get_width()/2., pos,
    #                         f'{value:.2f}%', ha='center', va='center',
    #                         fontsize=15, fontweight='bold', color='white')

    #         add_internal_label(bar_month, data['monthly_actual'])
    #         add_internal_label(bar_day, data['daily_actual'])

    #         # Gap 라벨 (막대 위, 색상 구분)
    #         def add_gap_label(x_pos, gap, base_height):
    #             color = 'blue' if gap >= 0 else 'red'
    #             sign = '+' if gap >= 0 else ''
    #             ax.text(x_pos, base_height + 0.3,
    #                     f'{sign}{gap:.2f}%', ha='center', va='bottom',
    #                     fontsize=15, fontweight='bold', color=color)

    #         add_gap_label(x[0], data['monthly_gap'], data['monthly_actual'])
    #         add_gap_label(x[1], data['daily_gap'], data['daily_actual'])

    #         # 범례 (상단 왼쪽)
    #         ax.legend(loc='upper center', ncol=1, frameon=False, fontsize=15, bbox_to_anchor=(0.5, 0.98))

    #         # 그리드
    #         ax.grid(axis='y', linestyle='--', alpha=0.7, zorder=0)
    #         ax.set_axisbelow(True)
    #         ax.spines['top'].set_visible(False)
    #         ax.spines['right'].set_visible(False)

    #     plt.tight_layout(pad=1.0)
    #     plt.savefig(chart_path, dpi=300, bbox_inches='tight', pad_inches=0.5)
    #     plt.close()

    #     # Base64 인코딩
    #     with open(chart_path, "rb") as img_file:
    #         img_base64 = base64.b64encode(img_file.read()).decode()
    #     # ──────────────────────────────────────────────────
    #     # 5. 표 생성 (계층형 DataFrame)
    #     # ──────────────────────────────────────────────────
    #     table_rows = []
    #     for yld_type in ['RTY']:
    #         for grade in ['Total', 'Prime']:
    #             key = f"{grade}_{yld_type}"
    #             data = results[key]
    #             table_rows.append({
    #                 'yld_type': yld_type,
    #                 'grade': grade,
    #                 'monthly_plan': data['monthly_plan'],
    #                 'monthly_actual': data['monthly_actual'],
    #                 'monthly_gap': data['monthly_gap'],
    #                 'daily_plan': data['daily_plan'],
    #                 'daily_actual': data['daily_actual'],
    #                 'daily_gap': data['daily_gap']
    #             })

    #     table_df = pd.DataFrame(table_rows)
    #     print(f"table_df:", table_df)

    #     details.update({
    #         'chart_path': str(chart_path),
    #         'img_base64': img_base64,
    #         'table_df': table_df,
    #         'summary': table_df,
    #         'daily_actual_date': daily_actual_date,
    #         'results': results
    #     })

    #     return details

    def _create_DATA_3210_wafering_300(self):
        """3210 불량률 상세 분석 """
        details = {}

        if 'DATA_3210_wafering_300' not in self.data or self.data['DATA_3210_wafering_300'].empty:
            print("DATA_3210_wafering_300 데이터 없음 또는 비어 있음")
            return details

        df = self.data['DATA_3210_wafering_300'].copy()

        # 컬럼 타입 변환
        numeric_cols = ['LOSS_RATIO', 'GOAL_RATIO', 'GOAL_RATIO_SUM', 'GAP_RATIO', 'LOSS_QTY', 'MGR_QTY']
        for col in numeric_cols:
            if col in df.columns:
                df[col] = pd.to_numeric(df[col], errors='coerce')

        # 그룹별 집계
        summary = df.groupby(['BASE_DT_NM', 'REJ_GROUP'], dropna=False).agg(
            AVG_LOSS_RATIO=('LOSS_RATIO', 'sum'),
            AVG_GOAL_RATIO=('GOAL_RATIO', 'mean'),
            TOTAL_MGR_QTY=('MGR_QTY', 'mean')
        ).reset_index()

        # 백분율 계산
        summary['LOSS_RATIO_PCT'] = (summary['AVG_LOSS_RATIO'] * 100).round(2)
        summary['GOAL_RATIO_PCT'] = (summary['AVG_GOAL_RATIO'] * 100).round(2)
        summary['GAP_PCT'] = (summary['LOSS_RATIO_PCT'] - summary['GOAL_RATIO_PCT']).round(2)

        # === Total 행 계산 ===
        total_row = {
            'BASE_DT_NM': summary['BASE_DT_NM'].iloc[0] if len(summary) > 0 else "Unknown",
            'REJ_GROUP': 'Total',
            'AVG_LOSS_RATIO': summary['AVG_LOSS_RATIO'].sum(),  # 전체 합 (원본이 sum이므로)
            'AVG_GOAL_RATIO': summary['AVG_GOAL_RATIO'].sum(), # 전체 합
            'TOTAL_MGR_QTY': summary['TOTAL_MGR_QTY'].mean(),   # 평균 유지
        }

        # 백분율 계산 (Total 기준)
        total_row['LOSS_RATIO_PCT'] = round(total_row['AVG_LOSS_RATIO'] * 100, 2)
        total_row['GOAL_RATIO_PCT'] = round(total_row['AVG_GOAL_RATIO'] * 100, 2)
        total_row['GAP_PCT'] = round(total_row['LOSS_RATIO_PCT'] - total_row['GOAL_RATIO_PCT'], 2)

        # Total 행을 데이터프레임으로 변환 후 summary에 추가
        total_df = pd.DataFrame([total_row])
        summary = pd.concat([summary, total_df], ignore_index=True)

        # 정렬: Total 제외하고 GAP_PCT 큰 순서대로, Total은 맨 아래
        summary_sorted = summary[summary['REJ_GROUP'] != 'Total'].sort_values('GAP_PCT', ascending=False)
        summary = pd.concat([summary_sorted, total_df], ignore_index=True)

        base_date = summary['BASE_DT_NM'].iloc[0] if len(summary) > 0 else "Unknown"
        print(f"분석 대상일: {base_date}")

        # 출력 디렉터리
        PROJECT_ROOT = Path(__file__).parent.parent
        # base_date = (datetime.now().date() - timedelta(days=1))
        # date_folder_name = base_date.strftime("%Y%m%d")
        # debug_dir = PROJECT_ROOT / "daily_reports_debug" / date_folder_name
        debug_dir = PROJECT_ROOT / "daily_reports_debug" / self.target_date
        debug_dir.mkdir(exist_ok=True, parents=True)

        # ──────────────────────────────────────────────────
        # 1. 그래프 저장 → Base64 인코딩
        # ──────────────────────────────────────────────────
        chart_path = debug_dir / "prime_gap_chart.png"

        if chart_path.exists():
            chart_path.unlink() #파일 삭제
            print(f"기존 그래프 파일 삭제됨 : {chart_path}")

        # 'Total' 제외한 데이터로 그래프 생성
        plot_data = summary[summary['REJ_GROUP'] != 'Total'].copy()
        if plot_data.empty:
            print("그래프를 그릴 데이터가 없습니다 (Total 제외 후).")
            return details  # 또는 기본 이미지 처리


        plt.figure(figsize=(10, 6))
        x = np.arange(len(plot_data))
        bars = plt.bar(x, plot_data['GAP_PCT'],
                    color=plot_data['GAP_PCT'].apply(lambda x: '#ff0000' if x > 0 else '#0000ff'), linewidth=1)

        plt.title(f"Gap 분석 - {base_date}", fontsize=14, fontweight='bold')
        plt.xlabel('REJ_GROUP', fontsize=14)
        plt.ylabel('GAP (%)', fontsize=14)
        plt.xticks(x, plot_data['REJ_GROUP'], rotation=90, ha='right')

        for i, bar in enumerate(bars):
            height = bar.get_height()
            # 라벨 위치 조정 (양수는 위, 음수는 아래)
            offset = 0.01 * (1 if height >= 0 else -1)
            va_pos = 'bottom' if height >= 0 else 'top'
            plt.text(bar.get_x() + bar.get_width() / 2, height + offset,
                    f"{height:.2f}%", ha='center', va=va_pos,
                    fontsize=14, fontweight='bold', color='black')

        plt.ylim(-0.5, 1.0) #-0.5 ~ 0.5로 변경 요청
        plt.grid(axis='y', linestyle='--', alpha=0.7, zorder=0)
        plt.tight_layout()
        plt.savefig(chart_path, dpi=300, bbox_inches='tight')
        plt.close()

        # Base64 인코딩
        with open(chart_path, "rb") as img_file:
            img_base64 = base64.b64encode(img_file.read()).decode()

        # ──────────────────────────────────────────────────
        # 2. 상위 3개 불량 상세분석
        # ──────────────────────────────────────────────────
        top3_rej_groups = summary[summary['REJ_GROUP'] != 'Total'].nlargest(3, 'GAP_PCT')['REJ_GROUP'].tolist()
        print(f"상위 3개 불량: {top3_rej_groups}")

        yesterday_mid_list = []
        for rej_group in top3_rej_groups:
            group_df = df[df['REJ_GROUP'] == rej_group].copy()

            # MID_GROUP 매핑 적용
            mid_mapping = REJ_GROUP_TO_MID_MAPPING.get(rej_group, {})
            group_df['MID_GROUP'] = group_df['AFT_BAD_RSN_CD'].map(mid_mapping)
            group_df['MID_GROUP'] = group_df['MID_GROUP'].fillna(group_df['AFT_BAD_RSN_CD'])

            # MID_GROUP별 평균 LOSS_RATIO 계산
            mid_agg = group_df.groupby('MID_GROUP', dropna=False).agg(
                YESTERDAY_LOSS_RATIO=('LOSS_RATIO', 'sum') #mean -> sum으로 변경
            ).reset_index()

            mid_agg['REJ_GROUP'] = rej_group
            mid_agg['YESTERDAY_LOSS_PCT'] = (mid_agg['YESTERDAY_LOSS_RATIO'] * 100).round(2)
            yesterday_mid_list.append(mid_agg[['REJ_GROUP', 'MID_GROUP', 'YESTERDAY_LOSS_RATIO', 'YESTERDAY_LOSS_PCT']])

        # 전체 yesterday MID_GROUP 실적
        yesterday_mid_summary = pd.concat(yesterday_mid_list, ignore_index=True) if yesterday_mid_list else pd.DataFrame()

        # ──────────────────────────────────────────────────
        #  5. details에 top3 + yesterday_mid_summary 저장
        # ──────────────────────────────────────────────────
        details.update({
            'summary': summary,
            'top3_rej_groups': top3_rej_groups,
            'yesterday_mid_summary': yesterday_mid_summary,  # 핵심: MID_GROUP 실적 저장
            'chart_path': str(chart_path),
            'img_base64': img_base64
            # 'detailed_analysis': detailed_analysis_dict #딕셔너리 저장
        })

        self.top3_rej_groups = top3_rej_groups

        return details


    def _create_DATA_3210_wafering_300_3months(self):
        """3210 불량률 상세 분석(3개월) """
        details = {}
        try:
            if 'DATA_3210_wafering_300_3months' not in self.data or self.data['DATA_3210_wafering_300_3months'].empty:
                print("DATA_3210_wafering_300_3months 데이터 없음 또는 비어 있음")
                return details

            df = self.data['DATA_3210_wafering_300_3months'].copy()

            # 컬럼 타입 변환
            numeric_cols = ['LOSS_QTY', 'MGR_QTY']
            for col in numeric_cols:
                if col in df.columns:
                    df[col] = pd.to_numeric(df[col], errors='coerce')
            
            # 2. 분모 계산: BASE_DT_NM 기준 MGR_QTY 중복 제거 후 전체 합계
            #    (동일 일자의 MGR_QTY 는 동일하므로, 일자별 고유 값의 합 = 기간 총 투입량)
            mgr_qty_daily = df[['BASE_DT_NM', 'MGR_QTY']].drop_duplicates(subset=['BASE_DT_NM', 'MGR_QTY'])
            total_mgr_qty = mgr_qty_daily['MGR_QTY'].sum()

            if total_mgr_qty == 0:
                print("기간 내 MGR_QTY 합계가 0 입니다.")
                return details

            # 3. 분자 계산: AFT_BAD_RSN_CD 별 LOSS_QTY 전체 합계
            summary_list = []
            for rej_group, group_df in df.groupby('REJ_GROUP', dropna=False):
                mid_mapping = REJ_GROUP_TO_MID_MAPPING.get(rej_group, {})
                group_df = group_df.copy()
                group_df['MID_GROUP'] = group_df['AFT_BAD_RSN_CD'].map(mid_mapping)
                group_df['MID_GROUP'] = group_df['MID_GROUP'].fillna(group_df['AFT_BAD_RSN_CD'])

                print(f"🔧 [DEBUG] {rej_group} → MID_GROUP 생성됨: {group_df['MID_GROUP'].nunique()} 종류")

                # 그룹 집계: REJ_GROUP + MID_GROUP + AFT_BAD_RSN_CD
                agg_df = group_df.groupby(['REJ_GROUP', 'MID_GROUP', 'AFT_BAD_RSN_CD'], dropna=False).agg(
                    TOTAL_LOSS_QTY=('LOSS_QTY', 'sum'),
                    COUNT_DAYS=('BASE_DT_NM', 'nunique')
                ).reset_index()

                # 4. LOSS_RATIO 계산: 분자 (TOTAL_LOSS_QTY) / 분모 (total_mgr_qty)
                agg_df['AVG_LOSS_RATIO'] = agg_df['TOTAL_LOSS_QTY'] / total_mgr_qty
                agg_df['LOSS_RATIO_PCT'] = (agg_df['AVG_LOSS_RATIO'] * 100).round(2)

                # 참조용: 분모 정보 저장 (검증용)
                agg_df['TOTAL_MGR_QTY'] = total_mgr_qty

                summary_list.append(agg_df)

            # 전체 요약 병합
            summary_3months = pd.concat(summary_list, ignore_index=True)
            if 'MID_GROUP' not in summary_3months.columns:
                print("❌ [ERROR] MID_GROUP 컬럼이 생성되지 않음!")
                return details

            summary_3months['LOSS_RATIO_PCT'] = (summary_3months['AVG_LOSS_RATIO'] * 100).round(2)

            # yesterday_mid_summary 가져오기
            yesterday_mid = self.data.get('DATA_3210_wafering_300', {}).get('yesterday_mid_summary', pd.DataFrame())

            if yesterday_mid.empty:
                details['summary'] = summary_3months
                return details

            # 상위 3개 REJ_GROUP 가져오기 (Gap 기준)
            top3_rej_groups = self.data.get('DATA_3210_wafering_300', {}).get('top3_rej_groups', [])

            # === AFT_BAD_RSN_CD 기준 Gap 계산 ===
            def get_code_gap_ref(rej_group):
                df_3months = self.data.get('DATA_3210_wafering_300_3months')  # raw 데이터 사용
                if not isinstance(df_3months, pd.DataFrame) or df_3months.empty:
                    return pd.Series()
                if 'REJ_GROUP' not in df_3months.columns:
                    return pd.Series()
                df_ref = df_3months[df_3months['REJ_GROUP'] == rej_group].copy()
                if df_ref.empty:
                    return pd.Series()
                if 'AFT_BAD_RSN_CD' not in df_ref.columns:
                    return pd.Series()
                loss_ratio_series = df_ref.groupby('AFT_BAD_RSN_CD')['LOSS_RATIO'].sum()
                return loss_ratio_series

            # 3개월 평균 (Ref) 준비
            ref_3months = summary_3months[summary_3months['REJ_GROUP'].isin(yesterday_mid['REJ_GROUP'])].copy()
            # REJ_GROUP + MID_GROUP 기준 집계 (AFT_BAD_RSN_CD 는 제거하고 MID_GROUP 수준으로 비교)
            ref_3months = ref_3months.groupby(['REJ_GROUP', 'MID_GROUP'], dropna=False).agg(
                REF_AVG_LOSS_RATIO=('AVG_LOSS_RATIO', 'sum'),  # 동일 MID_GROUP 내 불량코드 합계 비율
                REF_LOSS_QTY=('TOTAL_LOSS_QTY', 'sum')
            ).reset_index()


            # 병합 → Gap 계산 (전체 사용)
            merged = pd.merge(
                yesterday_mid,
                ref_3months,
                on=['REJ_GROUP', 'MID_GROUP'],
                how='inner'
            )
            merged['GAP'] = merged['YESTERDAY_LOSS_PCT'] - (merged['REF_AVG_LOSS_RATIO'] * 100)
            merged['Gap'] = merged['GAP'].round(2)
            merged['실적(%)'] = merged['YESTERDAY_LOSS_PCT']
            merged['Ref(3개월)'] = (merged['REF_AVG_LOSS_RATIO'] * 100).round(2)
            merged['범례'] = merged['MID_GROUP']

            # 모든 분석 통합 시작
            top_rej_mid_groups = []
            group_tables = {}
            analysis_text = "[ Prime 주요 열위 불량 세부코드 분석 Ref.(3개월) 比 일 실적 변동 (상위 3개) ]\n"
            detailed_analysis_dict = {}  # 세부분석 결과 저장

            # REJ_GROUP별 분석 함수 매핑 (내부 정의)
            REJ_GROUP_TO_ANALYZER = {
                'FLATNESS': analyze_flatness,
                'WARP&BOW': analyze_warp,
                'GROWING': analyze_growing,
                'BROKEN': analyze_broken,
                'NANO': analyze_nano,
                'PIT': analyze_pit,
                'SCRATCH': analyze_scratch,
                'CHIP': analyze_chip,
                'EDGE': analyze_edge,
                'HUMAN_ERR': analyze_HUMAN_ERR,
                'VISUAL': analyze_VISUAL,
                'NOSALE': analyze_NOSALE,
                'OTHER': analyze_OTHER,
                'GR_보증': analyze_GR,
                'SAMPLE': analyze_sample,
                'PARTICLE': analyze_particle
            }
            # 데이터 준비
            df_wafer = self.data.get('DATA_WAF_3210_wafering_300')
            df_lot = self.data.get('DATA_LOT_3210_wafering_300')

            # === AFT_BAD_RSN_CD 기준 Gap 계산 (실적) ===
            def get_code_gap_daily(rej_group):
                if not isinstance(df_wafer, pd.DataFrame) or df_wafer.empty:
                    return pd.Series()
                if 'REJ_GROUP' not in df_wafer.columns:
                    return pd.Series()
                df_daily = df_wafer[df_wafer['REJ_GROUP'] == rej_group].copy()
                if df_daily.empty:
                    return pd.Series()
                if 'AFT_BAD_RSN_CD' not in df_daily.columns or 'LOSS_RATIO' not in df_daily.columns:
                    return pd.Series()
                loss_ratio_series = df_daily.groupby('AFT_BAD_RSN_CD')['LOSS_RATIO'].sum()
                return loss_ratio_series

            # 전역 딕셔너리에 저장
            self.CODE_GAP_REF = {}
            self.CODE_GAP_DAILY = {}

            for rej in top3_rej_groups:
                ref_series = get_code_gap_ref(rej)
                daily_series = get_code_gap_daily(rej)  # ✅ 추가됨
                self.CODE_GAP_REF[rej] = ref_series
                self.CODE_GAP_DAILY[rej] = daily_series  # ✅ 수정됨
                # 1. top_rej_mid_groups 생성 (그래프 기준)
                group_df = merged[merged['REJ_GROUP'] == rej].copy()  # ✅ 추가됨
                if group_df.empty:
                    continue
                top3 = group_df.nlargest(3, 'Gap')
                top_rej_mid_groups.extend([(rej, mid) for mid in top3['MID_GROUP']])

                # 2. 표 생성
                table = top3[['MID_GROUP', 'YESTERDAY_LOSS_PCT', 'REF_AVG_LOSS_RATIO', 'Gap']].copy()
                table.rename(columns={
                    'YESTERDAY_LOSS_PCT': '실적(%)',
                    'REF_AVG_LOSS_RATIO': 'Ref(3개월)'
                }, inplace=True)
                table['Ref(3개월)'] *= 100
                table['Ref(3개월)'] = table['Ref(3개월)'].round(2)
                group_tables[rej] = table


                # 표 생성
                table = top3[['MID_GROUP', 'YESTERDAY_LOSS_PCT', 'REF_AVG_LOSS_RATIO', 'Gap']].copy()
                table.rename(columns={
                    'YESTERDAY_LOSS_PCT': '실적(%)',
                    'REF_AVG_LOSS_RATIO': 'Ref(3개월)'
                }, inplace=True)
                table['Ref(3개월)'] *= 100
                table['Ref(3개월)'] = table['Ref(3개월)'].round(2)

                # 3. 분석 텍스트
                top_row = top3.iloc[0]
                analysis_text += f"\n {rej} 최대 Gap: {top_row['MID_GROUP']} ({top_row['Gap']:.2f}%)"

                # 4. 세부분석 함수 실행
                if rej not in REJ_GROUP_TO_ANALYZER:
                    detailed_analysis_dict[rej] = [f"[{rej} 분석] 함수 없음"]
                    continue

                analyzer_func = REJ_GROUP_TO_ANALYZER[rej]
                sig = signature(analyzer_func)
                bound_args = {}
                missing = []


                for param_name, param in sig.parameters.items():
                    if 'wafer' in param_name.lower():
                        if df_wafer is not None:
                            df_target = df_wafer[df_wafer['REJ_GROUP'] == rej].copy()
                            bound_args[param_name] = df_target
                        else:
                            missing.append(param_name)
                    elif 'lot' in param_name.lower():
                        if df_lot is not None:
                            bound_args[param_name] = df_lot
                        else:
                            missing.append(param_name)
                    elif param.name == 'target_mids':
                        bound_args['target_mids'] = [mid for mid in top3['MID_GROUP']]
                        print(f"{rej}: target_mids → {bound_args['target_mids']}")
                    elif param.name == 'target_codes':
                        target_codes = top3['MID_GROUP'].tolist()
                        bound_args['target_codes'] = target_codes
                        print(f"{rej}: target_codes → {target_codes}")
                    elif param.default == param.empty:
                        missing.append(param_name)

                if missing:
                    detailed_analysis_dict[rej] = [f"[{rej} 분석] 누락: {missing}"]
                    continue

                try:
                    result = analyzer_func(**bound_args)
                    content_only = result[1:] if len(result) > 1 and result[0].startswith("[") else result
                    detailed_analysis_dict[rej] = content_only
                except Exception as e:
                    detailed_analysis_dict[rej] = [f"[{rej} 분석] 오류: {e}"]


            # 개별 플롯 생성
            plot_paths = self._create_top3_midgroup_plot_per_group(merged, top3_rej_groups)

            # 최종 details 업데이트
            details.update({
                'summary': summary_3months,
                'top_rej_mid_groups': top_rej_mid_groups,
                'top3_midgroup_analysis': {
                    'tables': group_tables,
                    'plot_paths': plot_paths,
                    'analysis': analysis_text.strip(),
                    'detailed_analysis': detailed_analysis_dict  #  세부분석 포함
                }
            })

            return details
        
        except Exception as e:
            print(f"\n❌ [ERROR] _create_DATA_3210_wafering_300_3months 중 오류 발생: {e}")
            import traceback
            traceback.print_exc()
            return details  # 빈 상태로 반환

    def _create_top3_midgroup_plot_per_group(self, merged_df, top3_rej_groups):
        """
        각 REJ_GROUP별로 Gap 상위 3개 MID_GROUP만 추출하여 개별 막대그래프 생성
        → 결과: {'GR_보증': 'path1.png', 'SAMPLE': 'path2.png', ...}
        """
        # PROJECT_ROOT 및 날짜 폴더
        PROJECT_ROOT = Path(__file__).parent.parent
        # base_date = (datetime.now().date() - timedelta(days=1))
        # date_folder_name = base_date.strftime("%Y%m%d")
        # debug_dir = PROJECT_ROOT / "daily_reports_debug" / date_folder_name
        debug_dir = PROJECT_ROOT / "daily_reports_debug" / self.target_date
        debug_dir.mkdir(exist_ok=True, parents=True)

        plot_paths = {}

        for rej_group in top3_rej_groups:
            try:
                # 해당 REJ_GROUP 데이터 필터링
                group_df = merged_df[merged_df['REJ_GROUP'] == rej_group].copy()
                if group_df.empty:
                    print(f"{rej_group}: 분석 데이터 없음")
                    continue

                # Gap 기준 상위 3개만 추출
                top3_mids = group_df.nlargest(3, 'Gap')

                # 파일명
                safe_rej = "".join(c if c.isalnum() else "_" for c in rej_group)
                plot_path = debug_dir  / f"prime_midgroup_top3_gap_{safe_rej}.png"

                # 기존 파일 삭제
                if plot_path.exists():
                    plot_path.unlink()
                # 서브플롯 생성 (가로 3개)
                fig, axes = plt.subplots(1, 3, figsize=(15, 4), dpi=300)
                # 그래프 수에 따라 안 보이는 서브플롯 숨기기
                n_plots = len(top3_mids)
                for i in range(n_plots, 3):
                    axes[i].set_visible(False)

                # 전체 Y축 범위 계산 (3개 MID 기준 통합)
                all_values = []
                for _, row in top3_mids.iterrows():
                    all_values.extend([float(row['Ref(3개월)']), float(row['실적(%)'])])
                
                global_min = min(all_values) if all_values else 0
                global_max = max(all_values) if all_values else 0

                # 여유 포함 (0 반드시 포함)
                y_min = min(0, float(global_min) * 0.95)
                y_max = max(float(global_max) * 1.15, 0.01)

                # 각 MID_GROUP 그래프 그리기 (동일 Y축)
                for i, (idx, row) in enumerate(top3_mids.iterrows()):
                    ax = axes[i]
                    mid_name = row['MID_GROUP']
                    ref_val = float(row['Ref(3개월)'])
                    actual_val = float(row['실적(%)'])
                    gap_val = actual_val - ref_val

                    # X 축 위치
                    x_center = 0
                    bar_width = 0.9
                    ref_x = x_center - bar_width/2
                    daily_x = x_center + bar_width/2

                    # 막대 그리기
                    bar_ref = ax.bar(ref_x, ref_val, bar_width, color='#0000ff')
                    bar_actual = ax.bar(daily_x, actual_val, bar_width, color='#ff0000')

                    #  Y 축 범위 (0 반드시 포함)
                    ax.set_ylim(y_min, y_max)
                    ax.set_title(f"[{mid_name}]", fontsize=14, fontweight='bold', pad=15)

                    # X 축 라벨
                    ax.set_xticks([ref_x, daily_x])
                    ax.set_xticklabels(['Ref.', '일'], fontsize=14, fontweight='bold')

                    #  내부 라벨 (흰색, 막대 중앙)
                    label_offset = 0.01
                    for bar, val in zip([bar_ref, bar_actual], [ref_val, actual_val]):
                        height = bar[0].get_height()
                        # 양수면 위, 음수면 아래
                        if height >= 0:
                            va = 'bottom'
                            label_y = height + label_offset
                        else:
                            va = 'top'
                            label_y = height - label_offset

                        ax.text(bar[0].get_x() + bar[0].get_width()/2., label_y,
                                f'{val:.2f}%', 
                                ha='center', va=va,                    # 가로 중앙, 세로 정렬 (bottom/top)
                                fontsize=15, fontweight='bold', 
                                color='black',                         
                                zorder=4)                              # 막대 위에 표시

                    #  Gap 라벨 (제목과 겹치지 않게)
                    max_val = max(ref_val, actual_val)
                    min_val = min(ref_val, actual_val)
                    
                    # 양수면 막대 위, 음수면 막대 아래
                    if max_val >= 0:
                        gap_y = max_val * 1.1  
                        va_align = 'bottom'
                    else:
                        gap_y = min_val * 1.1
                        va_align = 'top'
                    
                    gap_x = x_center
                    gap_color = '#ff0000' if gap_val >= 0 else '#0000ff'
                    ax.text(gap_x, gap_y, f'{gap_val:+.2f}%', ha='center', va=va_align,
                            fontsize=15, fontweight='bold', color=gap_color)

                    # 그리드
                    ax.grid(axis='y', linestyle='--', alpha=0.7, zorder=0)
                    ax.set_axisbelow(True)
                    ax.spines['top'].set_visible(False)
                    ax.spines['right'].set_visible(False)

                # 전체 제목
                fig.suptitle(f"[ {rej_group} 상위 3개 분석 ]", fontsize=14, fontweight='bold', y=1.02)

                plt.tight_layout()
                plt.savefig(str(plot_path), dpi=300, bbox_inches='tight')
                plt.close()

                if plot_path.exists():
                    plot_paths[rej_group] = str(plot_path)
                else:
                    raise RuntimeError(f"파일 생성 실패: {plot_path}")

            except Exception as e:
                print(f"{rej_group} 플롯 생성 실패: {e}")
                continue

        return plot_paths



    def _create_DATA_LOT_3210_wafering_300(self):
        """3210 LOT 상세 분석 - 캐시된 3개월 데이터 + self.data의 당일 데이터 모두 활용"""

        details = {}

        # ===================================================================
        # 1. [신규] data_cache에서 3개월 데이터 직접 로드 (장기 분석용)
        # ===================================================================
        PROJECT_ROOT = Path(__file__).parent.parent  

        # 어제 날짜 폴더 생성
        # base_date = (datetime.now().date() - timedelta(days=1))
        # date_folder_name = base_date.strftime("%Y%m%d")  # 예: 20260204
        # debug_dir = PROJECT_ROOT / "daily_reports_debug" / date_folder_name
        base_date = self.target_date_obj
        date_folder_name = base_date.strftime("%Y%m%d")
        debug_dir = PROJECT_ROOT / "daily_reports_debug" / date_folder_name 
        debug_dir.mkdir(exist_ok=True, parents=True)  # 폴더 생성

        target_months = []
        current = self.target_date_obj.replace(day=1)
        for _ in range(3):
            # 전월로 이동
            current = (current - timedelta(days=1)).replace(day=1)
            month_str = current.strftime("%Y%m")
            target_months.append(month_str)

        # 역순 정렬 (과거 → 최근)
        target_months = sorted(target_months)

        cache_dir = PROJECT_ROOT / "data_cache"
        pattern = "DATA_LOT_3210_wafering_300_*.parquet"
        parquet_files = list(cache_dir.glob(pattern))

        df_cached_3months = pd.DataFrame()

        if parquet_files:
            valid_files = []
            for file_path in parquet_files:
                try:
                    stem = file_path.stem  # 전체 이름 (확장자 제외)
                    date_part = stem.split('_')[-1]  # '202506'

                    if len(date_part) != 6 or not date_part.isdigit():
                        continue  # 형식 맞지 않으면 건너뜀

                    file_ym = date_part  # '202506' 형식
                except Exception as e:
                    print(f"[캐시] {file_path.name}에서 월 정보 추출 실패 → 건너뜀: {e}")
                    continue

                if file_ym in target_months:
                    valid_files.append(file_path)

            dfs = []
            for file_path in valid_files:
                try:
                    df_part = pd.read_parquet(file_path)
                    dfs.append(df_part)
                except Exception as e:
                    print(f"[캐시] {file_path.name} 읽기 실패: {e}")

            if dfs:
                df_cached_3months = pd.concat(dfs, ignore_index=True)
            else:
                print("[캐시] 모든 파일 로드 실패 → 3개월 데이터 없음")
        else:
            print("[캐시] data_cache에 DATA_LOT_3210_wafering_300_*.parquet 파일 없음")

        # ===================================================================
        # 2. [기존] self.data에서 당일 데이터 사용 (실시간 리포트용)
        # ===================================================================
        df_self_data = pd.DataFrame()
        if 'DATA_LOT_3210_wafering_300' in self.data and not self.data['DATA_LOT_3210_wafering_300'].empty:
            df_self_data = self.data['DATA_LOT_3210_wafering_300']
        else:
            print("[self.data] DATA_LOT_3210_wafering_300 없거나 빈 데이터")

        # ===================================================================
        # [핵심] MS6 기반 PRODUCT_TYPE 병합
        # ===================================================================
        if not df_cached_3months.empty:
            df_cached_3months = self._merge_product_type(df_cached_3months)

        if not df_self_data.empty:
            df_self_data = self._merge_product_type(df_self_data)

        # ===================================================================
        # 3. [핵심] 3개월 데이터 기반 Loss Rate 분석
        # ===================================================================
        if not df_cached_3months.empty:
            total_months = 3

            # ===================================================================
            # (1) 전체 (Total) 데이터 기준 분석
            # ===================================================================
            denominator_data = df_cached_3months[df_cached_3months['REJ_GROUP'] == '분모']
            total_in_qty = denominator_data['IN_QTY'].sum() 
            avg_in_qty = total_in_qty / total_months

            if avg_in_qty == 0:
                print(" 분모(IN_QTY)가 0입니다. Loss Rate 계산 불가")
                self.avg_in_qty = 0
                self.total_daily_qty = 0

                # 🔽 이 부분이 사라진 게 아닙니다 — 아래 else 블록에서 처리됩니다
            else:
                self.avg_in_qty = avg_in_qty

                valid_cached = df_cached_3months[df_cached_3months['REJ_GROUP'].notna()]
                total_loss_by_cret = valid_cached.groupby('CRET_CD')['LOSS_QTY'].sum() / total_months

                daily_loss_by_cret = pd.Series(dtype='int64')
                total_daily_qty = 0
                if not df_self_data.empty:
                    valid_daily = df_self_data[df_self_data['REJ_GROUP'].notna()]
                    daily_loss_by_cret = valid_daily.groupby('CRET_CD')['LOSS_QTY'].sum()
                    denominator_daily = df_self_data[df_self_data['REJ_GROUP'] == '분모']
                    total_daily_qty = denominator_daily['IN_QTY'].sum()
                self.total_daily_qty = total_daily_qty

                cret_list = ['FS', 'HG', 'RESC']
                report_table_total = []
                ref_qty_dict = {}
                daily_qty_dict = {}

                for cret_cd in cret_list:
                    ref_qty = total_loss_by_cret.get(cret_cd, 0)
                    daily_qty = daily_loss_by_cret.get(cret_cd, 0)
                    ref_rate = (ref_qty / avg_in_qty) * 100 if avg_in_qty != 0 else 0
                    daily_rate = (daily_qty / total_daily_qty) * 100 if total_daily_qty != 0 else 0
                    gap = ref_rate - daily_rate 

                    report_table_total.append({
                        '구분': cret_cd,
                        'Ref.(3개월)': int(ref_qty),
                        '일': int(daily_qty),
                        'Ref.(3개월)%': f"{ref_rate:+.2f}%",
                        '일%': f"{daily_rate:+.2f}%",
                        'Gap': f"{gap:+.2f}%"
                    })
                    ref_qty_dict[cret_cd] = int(ref_qty)
                    daily_qty_dict[cret_cd] = int(daily_qty)

                ref_qty_dict['모수'] = int(avg_in_qty)
                daily_qty_dict['모수'] = int(total_daily_qty)
                report_table_total.append({
                    '구분': '모수',
                    'Ref.(3개월)': ref_qty_dict['모수'],
                    '일': daily_qty_dict['모수'],
                    'Ref.(3개월)%': "", '일%': "", 'Gap': ""
                })

                details['rc_hg_ref_qty_total'] = ref_qty_dict
                details['rc_hg_daily_qty_total'] = daily_qty_dict
                details['rc_hg_avg_in_qty'] = avg_in_qty
                details['summary'] = pd.DataFrame(report_table_total)

                # ===================================================================
                # 4. 그룹별 비교 표 생성 + 그래프 생성 (모수 제외)
                # ===================================================================
                rej_groups = ['PARTICLE', 'FLATNESS', 'WARP&BOW', 'NANO']
                details['rc_hg_ref_qty_by_group'] = {}
                details['rc_hg_daily_qty_by_group'] = {}
                details['rc_hg_ref_rate_by_group'] = {}    
                details['rc_hg_daily_rate_by_group'] = {}  
                details['rc_hg_gap_data_by_group'] = {}
                details['loss_rate_table_by_group'] = {}
                details['rc_hg_gap_chart_path_by_group'] = {}

                for group in rej_groups:
                    group_data = df_cached_3months[df_cached_3months['REJ_GROUP'] == group]
                    group_loss_by_cret = group_data.groupby('CRET_CD')['LOSS_QTY'].sum() / total_months

                    group_daily_loss_by_cret = pd.Series(dtype='int64')
                    if not df_self_data.empty:
                        group_self_data = df_self_data[df_self_data['REJ_GROUP'] == group]
                        group_daily_loss_by_cret = group_self_data.groupby('CRET_CD')['LOSS_QTY'].sum()

                    group_table = []
                    gap_data = {}
                    ref_qty_dict_group = {}
                    daily_qty_dict_group = {}
                    ref_rate_dict_group = {}     
                    daily_rate_dict_group = {}   

                    for cret_cd in cret_list:
                        ref_qty = group_loss_by_cret.get(cret_cd, 0)
                        daily_qty = group_daily_loss_by_cret.get(cret_cd, 0)
                        ref_rate = (ref_qty / avg_in_qty) * 100 if avg_in_qty != 0 else 0
                        daily_rate = (daily_qty / total_daily_qty) * 100 if total_daily_qty != 0 else 0
                        gap = ref_rate - daily_rate 

                        group_table.append({
                            '구분': cret_cd,
                            'Ref.(3개월)': int(ref_qty),
                            '일': int(daily_qty),
                            'Ref.(3개월)%': f"{ref_rate:+.2f}%",
                            '일%': f"{daily_rate:+.2f}%",
                            'Gap': f"{gap:+.2f}%"
                        })
                        gap_data[cret_cd] = gap
                        ref_qty_dict_group[cret_cd] = int(ref_qty)
                        daily_qty_dict_group[cret_cd] = int(daily_qty)
                        ref_rate_dict_group[cret_cd] = ref_rate      
                        daily_rate_dict_group[cret_cd] = daily_rate 

                    group_table_df = pd.DataFrame(group_table)
                    if group_table_df.empty:
                        group_table_df = pd.DataFrame(columns=['구분', 'Ref.(3개월)', '일', 'Ref.(3개월)%', '일%', 'Gap'])
                    details['loss_rate_table_by_group'][group] = group_table_df

                    details['rc_hg_ref_qty_by_group'][group] = ref_qty_dict_group
                    details['rc_hg_daily_qty_by_group'][group] = daily_qty_dict_group
                    details['rc_hg_ref_rate_by_group'][group] = ref_rate_dict_group      
                    details['rc_hg_daily_rate_by_group'][group] = daily_rate_dict_group 
                    details['rc_hg_gap_data_by_group'][group] = gap_data

                    fig, axes = plt.subplots(1, 3, figsize=(10, 6), dpi=300)
                    categories = ['FS', 'RESC', 'HG']
                    Y_LIM_MAP = {
                        'FS': (0.0, 2.0),
                        'RESC': (-1.0, 1.0),
                        'HG': (-1.0, 1.0)
                    }
                    ref_color = '#0000ff'
                    daily_color = '#ff0000'

                    for idx, cret_cd in enumerate(categories):
                        ax = axes[idx]
                        ref_rate = float(ref_rate_dict_group.get(cret_cd, 0.0))
                        daily_rate = float(daily_rate_dict_group.get(cret_cd, 0.0))
                        gap = float(gap_data.get(cret_cd, 0.0))

                        y_min, y_max = Y_LIM_MAP[cret_cd]
                        y_range = y_max - y_min
                        label_offset = y_range * 0.025
                        gap_offset = y_range * 0.06

                        x_center = 0
                        bar_width = 0.9
                        ref_x = x_center - bar_width/2
                        daily_x = x_center + bar_width/2

                        bar_ref = ax.bar(ref_x, ref_rate, bar_width, color=ref_color)
                        bar_daily = ax.bar(daily_x, daily_rate, bar_width, color=daily_color)
                        ax.set_ylim(y_min, y_max)

                        def add_internal_label(bar, value):
                            for rect in bar:
                                height = rect.get_height()
                                pos_y = height + label_offset if height >= 0 else height - label_offset
                                va = 'bottom' if height >= 0 else 'top'
                                ax.text(rect.get_x() + rect.get_width()/2., pos_y,
                                        f'{value:.2f}%', ha='center', va=va, fontsize=15,
                                        fontweight='bold', color='black', zorder=4)

                        add_internal_label(bar_ref, ref_rate)
                        add_internal_label(bar_daily, daily_rate)

                        if ref_rate >= 0 and daily_rate >= 0:
                            gap_y = max(ref_rate, daily_rate) + gap_offset
                            va_align = 'bottom'
                        elif ref_rate <= 0 and daily_rate <= 0:
                            gap_y = min(ref_rate, daily_rate) - gap_offset
                            va_align = 'top'
                        else:
                            gap_y = (max(ref_rate, daily_rate) + gap_offset) if gap >= 0 else (min(ref_rate, daily_rate) - gap_offset)
                            va_align = 'bottom' if gap >= 0 else 'top'

                        gap_color = '#ff0000' if gap >= 0 else '#0000ff'
                        sign = '+' if gap >= 0 else ''
                        ax.text(x_center, gap_y, f'{sign}{gap:.2f}%', ha='center', va=va_align,
                                fontsize=15, fontweight='bold', color=gap_color)

                        ax.set_xticks([ref_x, daily_x])
                        ax.set_xticklabels(['Ref.', '일'], fontsize=20, fontweight='bold')
                        ax.set_title(cret_cd, fontsize=20, fontweight='bold', pad=10)
                        ax.yaxis.set_major_formatter(plt.FuncFormatter(lambda x, p: f'{x:.1f}%'))
                        ax.tick_params(axis='y', labelsize=14)
                        ax.grid(axis='y', linestyle='--', alpha=0.7)
                        ax.set_axisbelow(True)
                        ax.spines['top'].set_visible(False)
                        ax.spines['right'].set_visible(False)

                    fig.suptitle(f'RC/HG 보상 ({group})', fontsize=14, fontweight='bold', y=1.05)
                    plt.tight_layout(pad=0.8)

                    graph_path = debug_dir / f"RC_HG_보상_{group}.png"
                    if graph_path.exists(): graph_path.unlink()
                    plt.savefig(graph_path, dpi=300, bbox_inches='tight')
                    plt.close()
                    details['rc_hg_gap_chart_path_by_group'][group] = str(graph_path)

                # ===================================================================
                # 7. 전체 RC/HG 보상 그래프 생성
                # ===================================================================
                total_gap_data = {}
                total_ref_data = {}
                total_daily_data = {}
                for row in details['summary'].to_dict('records'):
                    if row['구분'] in ['FS', 'RESC', 'HG']:
                        total_gap_data[row['구분']] = float(row['Gap'].replace('%', '').replace('+', ''))
                        total_ref_data[row['구분']] = float(row['Ref.(3개월)%'].replace('%', '').replace('+', ''))
                        total_daily_data[row['구분']] = float(row['일%'].replace('%', '').replace('+', ''))

                fig, axes = plt.subplots(1, 3, figsize=(10, 6), dpi=300)
                Y_LIM_MAP = {
                    'FS': (4.0, 12.0),
                    'RESC': (-3.0, 1.0),
                    'HG': (-3.0, 1.0)
                }
                for idx, cret_cd in enumerate(['FS', 'RESC', 'HG']):
                    ax = axes[idx]
                    ref_rate = total_ref_data.get(cret_cd, 0)
                    daily_rate = total_daily_data.get(cret_cd, 0)
                    gap = total_gap_data.get(cret_cd, 0)
                    y_min, y_max = Y_LIM_MAP[cret_cd]
                    label_offset = (y_max - y_min) * 0.02

                    x_center = 0
                    bar_width = 0.9
                    ref_x = x_center - bar_width/2
                    daily_x = x_center + bar_width/2

                    ax.bar(ref_x, ref_rate, bar_width, color='#0000ff')
                    ax.bar(daily_x, daily_rate, bar_width, color='#ff0000')
                    ax.set_ylim(y_min, y_max)

                    def add_label(bar, value):
                        for rect in bar:
                            height = rect.get_height()
                            pos_y = height + label_offset if height >= 0 else height - label_offset
                            va = 'bottom' if height >= 0 else 'top'
                            ax.text(rect.get_x() + rect.get_width()/2., pos_y,
                                    f'{value:.2f}%', ha='center', va=va, fontsize=15,
                                    fontweight='bold', color='black', zorder=4)

                    add_label(ax.containers[0], ref_rate)
                    add_label(ax.containers[1], daily_rate)

                    gap_y = (max(ref_rate, daily_rate) + label_offset * 3) if (ref_rate >= 0 or daily_rate >= 0) \
                        else (min(ref_rate, daily_rate) - label_offset * 3)
                    va_align = 'bottom' if gap >= 0 else 'top'
                    gap_color = '#0000ff' if gap >= 0 else '#ff0000'
                    sign = '+' if gap >= 0 else ''
                    ax.text(x_center, gap_y, f'{sign}{gap:.2f}%', ha='center', va=va_align,
                            fontsize=15, fontweight='bold', color=gap_color)

                    ax.set_xticks([ref_x, daily_x])
                    ax.set_xticklabels(['Ref.', '일'], fontsize=20, fontweight='bold')
                    ax.set_title(cret_cd, fontsize=20, fontweight='bold', pad=10)
                    ax.yaxis.set_major_formatter(plt.FuncFormatter(lambda x, p: f'{x:.1f}%'))
                    ax.tick_params(axis='y', labelsize=14)
                    ax.grid(axis='y', linestyle='--', alpha=0.7)
                    ax.set_axisbelow(True)
                    ax.spines['top'].set_visible(False)
                    ax.spines['right'].set_visible(False)

                fig.suptitle(f'RC/HG 보상 (전체)', fontsize=14, fontweight='bold', y=1.05)
                plt.tight_layout(pad=0.8)
                total_graph_path = debug_dir / "RC_HG_보상_전체.png"
                if total_graph_path.exists(): total_graph_path.unlink()
                plt.savefig(total_graph_path, dpi=300, bbox_inches='tight')
                plt.close()
                details['rc_hg_gap_chart_path_total'] = str(total_graph_path)

            # ===================================================================
            # (2) Prime 제품 전용 분석
            # ===================================================================
            if 'GRD_CD_NM_CS' in df_cached_3months.columns:
                df_cached_prime = df_cached_3months[df_cached_3months['GRD_CD_NM_CS'] == 'Prime']
                df_self_prime = df_self_data[df_self_data['GRD_CD_NM_CS'] == 'Prime'] if not df_self_data.empty else pd.DataFrame()

                if not df_cached_prime.empty:
                    denominator_data_prime = df_cached_prime[df_cached_prime['REJ_GROUP'] == '분모']
                    total_in_qty_prime = denominator_data_prime['IN_QTY'].sum()
                    avg_in_qty_prime = total_in_qty_prime / total_months if total_in_qty_prime > 0 else 0

                    if avg_in_qty_prime == 0:
                        print("Prime 분모(IN_QTY)가 0입니다. Loss Rate 계산 불가")
                        details['prime_avg_in_qty'] = 0
                    else:
                        valid_cached_prime = df_cached_prime[df_cached_prime['REJ_GROUP'].notna()]
                        total_loss_by_cret_prime = valid_cached_prime.groupby('CRET_CD')['LOSS_QTY'].sum() / total_months

                        daily_loss_by_cret_prime = pd.Series(dtype='int64')
                        total_daily_qty_prime = 0
                        if not df_self_prime.empty:
                            valid_daily_prime = df_self_prime[df_self_prime['REJ_GROUP'].notna()]
                            daily_loss_by_cret_prime = valid_daily_prime.groupby('CRET_CD')['LOSS_QTY'].sum()
                            denominator_daily_prime = df_self_prime[df_self_prime['REJ_GROUP'] == '분모']
                            total_daily_qty_prime = denominator_daily_prime['IN_QTY'].sum()

                        report_table_prime = []
                        ref_qty_dict_prime = {}
                        daily_qty_dict_prime = {}

                        for cret_cd in cret_list:
                            ref_qty = total_loss_by_cret_prime.get(cret_cd, 0)
                            daily_qty = daily_loss_by_cret_prime.get(cret_cd, 0)
                            ref_rate = (ref_qty / avg_in_qty_prime) * 100 if avg_in_qty_prime != 0 else 0
                            daily_rate = (daily_qty / total_daily_qty_prime) * 100 if total_daily_qty_prime != 0 else 0
                            gap = ref_rate - daily_rate

                            report_table_prime.append({
                                '구분': cret_cd,
                                'Ref.(3개월)': int(ref_qty),
                                '일': int(daily_qty),
                                'Ref.(3개월)%': f"{ref_rate:+.2f}%",
                                '일%': f"{daily_rate:+.2f}%",
                                'Gap': f"{gap:+.2f}%"
                            })
                            ref_qty_dict_prime[cret_cd] = int(ref_qty)
                            daily_qty_dict_prime[cret_cd] = int(daily_qty)

                        ref_qty_dict_prime['모수'] = int(avg_in_qty_prime)
                        daily_qty_dict_prime['모수'] = int(total_daily_qty_prime)
                        report_table_prime.append({
                            '구분': '모수',
                            'Ref.(3개월)': ref_qty_dict_prime['모수'],
                            '일': daily_qty_dict_prime['모수'],
                            'Ref.(3개월)%': "", '일%': "", 'Gap': ""
                        })

                        details['prime_rc_hg_ref_qty_total'] = ref_qty_dict_prime
                        details['prime_rc_hg_daily_qty_total'] = daily_qty_dict_prime
                        details['prime_avg_in_qty'] = avg_in_qty_prime
                        details['prime_summary'] = pd.DataFrame(report_table_prime)

                        details['prime_rc_hg_ref_qty_by_group'] = {}
                        details['prime_rc_hg_daily_qty_by_group'] = {}
                        details['prime_rc_hg_ref_rate_by_group'] = {}
                        details['prime_rc_hg_daily_rate_by_group'] = {}
                        details['prime_rc_hg_gap_data_by_group'] = {}
                        details['prime_loss_rate_table_by_group'] = {}
                        details['prime_rc_hg_gap_chart_path_by_group'] = {}

                        for group in ['PARTICLE', 'FLATNESS', 'WARP&BOW', 'NANO']:
                            group_data = df_cached_prime[df_cached_prime['REJ_GROUP'] == group]
                            group_loss_by_cret = group_data.groupby('CRET_CD')['LOSS_QTY'].sum() / total_months

                            group_daily_loss_by_cret = pd.Series(dtype='int64')
                            if not df_self_prime.empty:
                                group_self_data = df_self_prime[df_self_prime['REJ_GROUP'] == group]
                                group_daily_loss_by_cret = group_self_data.groupby('CRET_CD')['LOSS_QTY'].sum()

                            group_table = []
                            gap_data = {}
                            ref_qty_dict_g = {}
                            daily_qty_dict_g = {}
                            ref_rate_dict_g = {}
                            daily_rate_dict_g = {}

                            for cret_cd in cret_list:
                                ref_qty = group_loss_by_cret.get(cret_cd, 0)
                                daily_qty = group_daily_loss_by_cret.get(cret_cd, 0)
                                ref_rate = (ref_qty / avg_in_qty_prime) * 100 if avg_in_qty_prime != 0 else 0
                                daily_rate = (daily_qty / total_daily_qty_prime) * 100 if total_daily_qty_prime != 0 else 0
                                gap = ref_rate - daily_rate

                                group_table.append({
                                    '구분': cret_cd,
                                    'Ref.(3개월)': int(ref_qty),
                                    '일': int(daily_qty),
                                    'Ref.(3개월)%': f"{ref_rate:+.2f}%",
                                    '일%': f"{daily_rate:+.2f}%",
                                    'Gap': f"{gap:+.2f}%"
                                })
                                gap_data[cret_cd] = gap
                                ref_qty_dict_g[cret_cd] = int(ref_qty)
                                daily_qty_dict_g[cret_cd] = int(daily_qty)
                                ref_rate_dict_g[cret_cd] = ref_rate
                                daily_rate_dict_g[cret_cd] = daily_rate

                            group_table_df = pd.DataFrame(group_table)
                            if group_table_df.empty:
                                group_table_df = pd.DataFrame(columns=['구분', 'Ref.(3개월)', '일', 'Ref.(3개월)%', '일%', 'Gap'])
                            details['prime_loss_rate_table_by_group'][group] = group_table_df

                            details['prime_rc_hg_ref_qty_by_group'][group] = ref_qty_dict_g
                            details['prime_rc_hg_daily_qty_by_group'][group] = daily_qty_dict_g
                            details['prime_rc_hg_ref_rate_by_group'][group] = ref_rate_dict_g
                            details['prime_rc_hg_daily_rate_by_group'][group] = daily_rate_dict_g
                            details['prime_rc_hg_gap_data_by_group'][group] = gap_data

                            fig, axes = plt.subplots(1, 3, figsize=(10, 6), dpi=300)
                            Y_LIM_MAP = {
                                'FS': (0.0, 2.0),
                                'RESC': (-1.0, 1.0),
                                'HG': (-1.0, 1.0)
                            }
                            for idx, cret_cd in enumerate(['FS', 'RESC', 'HG']):
                                ax = axes[idx]
                                ref_rate = float(ref_rate_dict_g.get(cret_cd, 0.0))
                                daily_rate = float(daily_rate_dict_g.get(cret_cd, 0.0))
                                gap = float(gap_data.get(cret_cd, 0.0))

                                y_min, y_max = Y_LIM_MAP[cret_cd]
                                label_offset = (y_max - y_min) * 0.025
                                gap_offset = (y_max - y_min) * 0.06

                                x_center = 0
                                bar_width = 0.9
                                ref_x = x_center - bar_width/2
                                daily_x = x_center + bar_width/2

                                ax.bar(ref_x, ref_rate, bar_width, color='#0000ff')
                                ax.bar(daily_x, daily_rate, bar_width, color='#ff0000')
                                ax.set_ylim(y_min, y_max)

                                def add_label(bar, value):
                                    for rect in bar:
                                        height = rect.get_height()
                                        pos_y = height + label_offset if height >= 0 else height - label_offset
                                        va = 'bottom' if height >= 0 else 'top'
                                        ax.text(rect.get_x() + rect.get_width()/2., pos_y,
                                                f'{value:.2f}%', ha='center', va=va, fontsize=15,
                                                fontweight='bold', color='black', zorder=4)

                                add_label(ax.containers[0], ref_rate)
                                add_label(ax.containers[1], daily_rate)

                                if ref_rate >= 0 and daily_rate >= 0:
                                    gap_y = max(ref_rate, daily_rate) + gap_offset
                                    va_align = 'bottom'
                                elif ref_rate <= 0 and daily_rate <= 0:
                                    gap_y = min(ref_rate, daily_rate) - gap_offset
                                    va_align = 'top'
                                else:
                                    gap_y = (max(ref_rate, daily_rate) + gap_offset) if gap >= 0 else (min(ref_rate, daily_rate) - gap_offset)
                                    va_align = 'bottom' if gap >= 0 else 'top'

                                gap_color = '#ff0000' if gap >= 0 else '#0000ff'
                                sign = '+' if gap >= 0 else ''
                                ax.text(x_center, gap_y, f'{sign}{gap:.2f}%', ha='center', va=va_align,
                                        fontsize=15, fontweight='bold', color=gap_color)

                                ax.set_xticks([ref_x, daily_x])
                                ax.set_xticklabels(['Ref.', '일'], fontsize=20, fontweight='bold')
                                ax.set_title(cret_cd, fontsize=20, fontweight='bold', pad=10)
                                ax.yaxis.set_major_formatter(plt.FuncFormatter(lambda x, p: f'{x:.1f}%'))
                                ax.tick_params(axis='y', labelsize=14)
                                ax.grid(axis='y', linestyle='--', alpha=0.7)
                                ax.set_axisbelow(True)
                                ax.spines['top'].set_visible(False)
                                ax.spines['right'].set_visible(False)

                            fig.suptitle(f'RC/HG 보상 ({group}) - Prime', fontsize=14, fontweight='bold', y=1.05)
                            plt.tight_layout(pad=0.8)
                            graph_path = debug_dir / f"RC_HG_보상_{group}_Prime.png"
                            if graph_path.exists(): graph_path.unlink()
                            plt.savefig(graph_path, dpi=300, bbox_inches='tight')
                            plt.close()
                            details['prime_rc_hg_gap_chart_path_by_group'][group] = str(graph_path)

                        total_gap_data_p = {}
                        total_ref_data_p = {}
                        total_daily_data_p = {}
                        for row in details['prime_summary'].to_dict('records'):
                            if row['구분'] in ['FS', 'RESC', 'HG']:
                                total_gap_data_p[row['구분']] = float(row['Gap'].replace('%', '').replace('+', ''))
                                total_ref_data_p[row['구분']] = float(row['Ref.(3개월)%'].replace('%', '').replace('+', ''))
                                total_daily_data_p[row['구분']] = float(row['일%'].replace('%', '').replace('+', ''))

                        fig, axes = plt.subplots(1, 3, figsize=(10, 6), dpi=300)
                        Y_LIM_MAP = {
                            'FS': (4.0, 12.0),
                            'RESC': (-3.0, 1.0),
                            'HG': (-3.0, 1.0)
                        }
                        for idx, cret_cd in enumerate(['FS', 'RESC', 'HG']):
                            ax = axes[idx]
                            ref_rate = total_ref_data_p.get(cret_cd, 0)
                            daily_rate = total_daily_data_p.get(cret_cd, 0)
                            gap = total_gap_data_p.get(cret_cd, 0)
                            y_min, y_max = Y_LIM_MAP[cret_cd]
                            label_offset = (y_max - y_min) * 0.02

                            x_center = 0
                            bar_width = 0.9
                            ref_x = x_center - bar_width/2
                            daily_x = x_center + bar_width/2

                            ax.bar(ref_x, ref_rate, bar_width, color='#0000ff')
                            ax.bar(daily_x, daily_rate, bar_width, color='#ff0000')
                            ax.set_ylim(y_min, y_max)

                            def add_label(bar, value):
                                for rect in bar:
                                    height = rect.get_height()
                                    pos_y = height + label_offset if height >= 0 else height - label_offset
                                    va = 'bottom' if height >= 0 else 'top'
                                    ax.text(rect.get_x() + rect.get_width()/2., pos_y,
                                            f'{value:.2f}%', ha='center', va=va, fontsize=15,
                                            fontweight='bold', color='black', zorder=4)

                            add_label(ax.containers[0], ref_rate)
                            add_label(ax.containers[1], daily_rate)

                            gap_y = (max(ref_rate, daily_rate) + label_offset * 3) if (ref_rate >= 0 or daily_rate >= 0) \
                                else (min(ref_rate, daily_rate) - label_offset * 3)
                            va_align = 'bottom' if gap >= 0 else 'top'
                            gap_color = '#0000ff' if gap >= 0 else '#ff0000'
                            sign = '+' if gap >= 0 else ''
                            ax.text(x_center, gap_y, f'{sign}{gap:.2f}%', ha='center', va=va_align,
                                    fontsize=15, fontweight='bold', color=gap_color)

                            ax.set_xticks([ref_x, daily_x])
                            ax.set_xticklabels(['Ref.', '일'], fontsize=20, fontweight='bold')
                            ax.set_title(cret_cd, fontsize=20, fontweight='bold', pad=10)
                            ax.yaxis.set_major_formatter(plt.FuncFormatter(lambda x, p: f'{x:.1f}%'))
                            ax.tick_params(axis='y', labelsize=14)
                            ax.grid(axis='y', linestyle='--', alpha=0.7)
                            ax.set_axisbelow(True)
                            ax.spines['top'].set_visible(False)
                            ax.spines['right'].set_visible(False)

                        fig.suptitle(f'RC/HG 보상 (전체) - Prime', fontsize=14, fontweight='bold', y=1.05)
                        plt.tight_layout(pad=0.8)
                        total_graph_path_p = debug_dir / "RC_HG_보상_전체_Prime.png"
                        if total_graph_path_p.exists(): total_graph_path_p.unlink()
                        plt.savefig(total_graph_path_p, dpi=300, bbox_inches='tight')
                        plt.close()
                        details['prime_rc_hg_gap_chart_path_total'] = str(total_graph_path_p)
                else:
                    print("Prime 제품에 대한 3개월 데이터가 없습니다.")
            else:
                print("GRD_CD_NM_CS 컬럼이 데이터에 없습니다.")

        else:
            # 🔴 이 블록은 절대 삭제하지 마세요 — Total 데이터 없을 때 fallback
            details['rc_hg_ref_qty_total'] = {}
            details['rc_hg_daily_qty_total'] = {}
            details['rc_hg_ref_qty_by_group'] = {}
            details['rc_hg_daily_qty_by_group'] = {}
            details['rc_hg_avg_in_qty'] = 0
            details['rc_hg_gap_chart_path_by_group'] = {}
            details['rc_hg_gap_chart_path_total'] = ""
            self.avg_in_qty = 0
            self.total_daily_qty = 0

        # ===================================================================
        # 8. 기본 정보 추가
        # ===================================================================
        details['cache_data_available'] = not df_cached_3months.empty
        details['self_data_available'] = not df_self_data.empty
        details['cache_total_count'] = len(df_cached_3months) if not df_cached_3months.empty else 0
        details['self_data_count'] = len(df_self_data) if not df_self_data.empty else 0

        return details




    
    # def _create_DATA_LOT_3210_wafering_300(self):
    #     """3210 LOT 상세 분석 - 캐시된 3개월 데이터 + self.data의 당일 데이터 모두 활용"""

    #     details = {}

    #     # ===================================================================
    #     # 1. [신규] data_cache에서 3개월 데이터 직접 로드 (장기 분석용)
    #     # ===================================================================
    #     PROJECT_ROOT = Path(__file__).parent.parent  

    #     # 어제 날짜 폴더 생성
    #     base_date = (datetime.now().date() - timedelta(days=1))
    #     date_folder_name = base_date.strftime("%Y%m%d")  # 예: 20260204

    #     # 출력 폴더: daily_reports_debug/YYYYMMDD
    #     debug_dir = PROJECT_ROOT / "daily_reports_debug" / date_folder_name
    #     debug_dir.mkdir(exist_ok=True, parents=True)  # 폴더 생성

    #     target_months = []
    #     current = base_date.replace(day=1)
    #     for _ in range(3):
    #         # 전월로 이동
    #         current = (current - timedelta(days=1)).replace(day=1)
    #         month_str = current.strftime("%Y%m")
    #         target_months.append(month_str)

    #     # 역순 정렬 (과거 → 최근)
    #     target_months = sorted(target_months)

    #     # print(f"[캐시 필터링] 최근 3개월 대상 월: {target_months}")

    #     cache_dir = PROJECT_ROOT / "data_cache"
    #     pattern = "DATA_LOT_3210_wafering_300_*.parquet"
    #     parquet_files = list(cache_dir.glob(pattern))

    #     df_cached_3months = pd.DataFrame()

    #     if parquet_files:
    #         valid_files = []
    #         for file_path in parquet_files:
    #             try:
    #                 stem = file_path.stem  # 전체 이름 (확장자 제외)
    #                 date_part = stem.split('_')[-1]  # '202506'

    #                 if len(date_part) != 6 or not date_part.isdigit():
    #                     continue  # 형식 맞지 않으면 건너뜀

    #                 file_ym = date_part  # '202506' 형식
    #             except Exception as e:
    #                 print(f"[캐시] {file_path.name}에서 월 정보 추출 실패 → 건너뜀: {e}")
    #                 continue

    #             if file_ym in target_months:
    #                 valid_files.append(file_path)

    #         # print(f"[캐시 필터링] 전체 {len(parquet_files)}개 중 대상 {len(valid_files)}개 파일 선정: {[f.name for f in valid_files]}")

    #         dfs = []
    #         for file_path in valid_files:
    #             try:
    #                 df_part = pd.read_parquet(file_path)
    #                 # print(f"[캐시] {file_path.name} 로드 완료: {len(df_part):,} 건")
    #                 dfs.append(df_part)
    #             except Exception as e:
    #                 print(f"[캐시] {file_path.name} 읽기 실패: {e}")

    #         if dfs:
    #             df_cached_3months = pd.concat(dfs, ignore_index=True)
    #             # print(f"[캐시] 총 {len(df_cached_3months):,} 건 데이터 병합 완료")
    #         else:
    #             print("[캐시] 모든 파일 로드 실패 → 3개월 데이터 없음")
    #     else:
    #         print("[캐시] data_cache에 DATA_LOT_3210_wafering_300_*.parquet 파일 없음")

    #     # ===================================================================
    #     # 2. [기존] self.data에서 당일 데이터 사용 (실시간 리포트용)
    #     # ===================================================================
    #     df_self_data = pd.DataFrame()
    #     if 'DATA_LOT_3210_wafering_300' in self.data and not self.data['DATA_LOT_3210_wafering_300'].empty:
    #         df_self_data = self.data['DATA_LOT_3210_wafering_300']
    #         # print(f"[self.data] DATA_LOT_3210_wafering_300 데이터 건수: {len(df_self_data):,} 건")
    #     else:
    #         print("[self.data] DATA_LOT_3210_wafering_300 없거나 빈 데이터")


    #     # ===================================================================
    #     # [핵심] MS6 기반 PRODUCT_TYPE 병합
    #     # ===================================================================
    #     if not df_cached_3months.empty:
    #         df_cached_3months = self._merge_product_type(df_cached_3months)

    #     if not df_self_data.empty:
    #         df_self_data = self._merge_product_type(df_self_data)

    #     # print(f"PRODUCT_TYPE 병합 완료: 3개월 {df_cached_3months['PRODUCT_TYPE'].notna().sum()}건, 당일 {df_self_data['PRODUCT_TYPE'].notna().sum()}건")

    #     # ===================================================================
    #     # 3. [핵심] 3개월 데이터 기반 Loss Rate 분석
    #     # ===================================================================
    #     if not df_cached_3months.empty:
    #         # 3개월 수량 합계 → 평균으로 변환 (3으로 나눔)
    #         total_months = 3

    #         # ===================================================================
    #         # (1) 전체 (Total) 데이터 기준 분석
    #         # ===================================================================

    #         # 분모: REJ_GROUP == "분모" 인 IN_QTY 합계
    #         denominator_data = df_cached_3months[df_cached_3months['REJ_GROUP'] == '분모']
    #         total_in_qty = denominator_data['IN_QTY'].sum() 
    #         avg_in_qty = total_in_qty / total_months  # 3개월 평균 전체 분모

    #         if avg_in_qty == 0:
    #             print(" 분모(IN_QTY)가 0입니다. Loss Rate 계산 불가")
    #             self.avg_in_qty = 0 # 인스턴스 변수에 0 저장
    #             self.total_daily_qty = 0 # 인스턴스 변수에 0 저장
    #             return details

    #         self.avg_in_qty = avg_in_qty # 인스턴스 변수에 저장 → WAF 분석에서 사용

    #         # ===================================================================
    #         #  1. 전체 (Total) CRET_CD별 Loss Rate
    #         # ===================================================================

    #         valid_cached = df_cached_3months[df_cached_3months['REJ_GROUP'].notna()]
    #         total_loss_by_cret = valid_cached.groupby('CRET_CD')['LOSS_QTY'].sum() / total_months #FS/HG/RESC 별 loss_qty 3개월 평균

    #         # ===================================================================
    #         #  2. 당일 CRET_CD별 LOSS_QTY
    #         # ===================================================================
    #         daily_loss_by_cret = pd.Series(dtype='int64')
    #         total_daily_qty = 0

    #         if not df_self_data.empty:
    #             valid_daily = df_self_data[df_self_data['REJ_GROUP'].notna()]
    #             daily_loss_by_cret = valid_daily.groupby('CRET_CD')['LOSS_QTY'].sum()

    #             denominator_daily = df_self_data[df_self_data['REJ_GROUP'] == '분모']
    #             total_daily_qty = denominator_daily['IN_QTY'].sum()  #  정의 추가
    #         else:
    #             print("[self.data] DATA_LOT_3210_wafering_300 없거나 빈 데이터")

    #         self.total_daily_qty = total_daily_qty # 인스턴스 변수에 저장

    #         # ===================================================================
    #         #  3.  전체 비교 표 생성 (모수 포함)
    #         # ===================================================================
    #         cret_list = ['FS', 'HG', 'RESC']
    #         report_table_total = []

    #         #  원시 데이터 저장용
    #         ref_qty_dict = {}
    #         daily_qty_dict = {}

    #         for cret_cd in cret_list:
    #             ref_qty = total_loss_by_cret.get(cret_cd, 0)
    #             daily_qty = daily_loss_by_cret.get(cret_cd, 0)

    #             ref_rate = (ref_qty / avg_in_qty) * 100 if avg_in_qty != 0 else 0
    #             daily_rate = (daily_qty / total_daily_qty) * 100 if total_daily_qty != 0 else 0
    #             gap = ref_rate - daily_rate 

    #             report_table_total.append({
    #                 '구분': cret_cd,
    #                 'Ref.(3개월)': int(ref_qty),
    #                 '일': int(daily_qty),
    #                 'Ref.(3개월)%': f"{ref_rate:+.2f}%",
    #                 '일%': f"{daily_rate:+.2f}%",
    #                 'Gap': f"{gap:+.2f}%"
    #             })
    #             #  원시 데이터 저장
    #             ref_qty_dict[cret_cd] = int(ref_qty)
    #             daily_qty_dict[cret_cd] = int(daily_qty)

    #         #  모수 저장
    #         ref_qty_dict['모수'] = int(avg_in_qty) #3개월 평균 분모 -> ref 분모
    #         daily_qty_dict['모수'] = int(total_daily_qty) #일 분모

    #         report_table_total.append({
    #             '구분': '모수',
    #             'Ref.(3개월)': ref_qty_dict['모수'],
    #             '일': daily_qty_dict['모수'],
    #             'Ref.(3개월)%': "",
    #             '일%': "",
    #             'Gap': ""
    #         })

    #         #  details에 저장 (표 X, 값 O)
    #         details['rc_hg_ref_qty_total'] = ref_qty_dict
    #         details['rc_hg_daily_qty_total'] = daily_qty_dict
    #         details['rc_hg_avg_in_qty'] = avg_in_qty

    #         report_table_total_df = pd.DataFrame(report_table_total)
    #         details['summary'] = report_table_total_df

    #         # ===================================================================
    #         #  4. 그룹별 비교 표 생성 + 그래프 생성 (모수 제외)
    #         # ===================================================================
    #         rej_groups = ['PARTICLE', 'FLATNESS', 'WARP&BOW', 'NANO']
    #         details['rc_hg_ref_qty_by_group'] = {}
    #         details['rc_hg_daily_qty_by_group'] = {}
    #         details['rc_hg_ref_rate_by_group'] = {}    
    #         details['rc_hg_daily_rate_by_group'] = {}  
    #         details['rc_hg_gap_data_by_group'] = {}
    #         details['loss_rate_table_by_group'] = {}
    #         details['rc_hg_gap_chart_path_by_group'] = {}

    #         for group in rej_groups:
    #             # 각 그룹별 3개월 데이터 필터링
    #             group_data = df_cached_3months[df_cached_3months['REJ_GROUP'] == group]
    #             group_loss_by_cret = group_data.groupby('CRET_CD')['LOSS_QTY'].sum() / total_months
    #             # 각 그룹별 당일 데이터 필터링
    #             group_daily_loss_by_cret = pd.Series(dtype='int64')
    #             if not df_self_data.empty:
    #                 group_self_data = df_self_data[df_self_data['REJ_GROUP'] == group]
    #                 group_daily_loss_by_cret = group_self_data.groupby('CRET_CD')['LOSS_QTY'].sum()

    #             group_table = []
    #             gap_data = {}
    #             ref_qty_dict_group = {}
    #             daily_qty_dict_group = {}
    #             ref_rate_dict_group = {}     
    #             daily_rate_dict_group = {}   

    #             for cret_cd in cret_list:
    #                 ref_qty = group_loss_by_cret.get(cret_cd, 0)
    #                 daily_qty = group_daily_loss_by_cret.get(cret_cd, 0)

    #                 ref_rate = (ref_qty / avg_in_qty) * 100 if avg_in_qty != 0 else 0
    #                 daily_rate = (daily_qty / total_daily_qty) * 100 if total_daily_qty != 0 else 0
    #                 gap = ref_rate - daily_rate 

    #                 group_table.append({
    #                     '구분': cret_cd,
    #                     'Ref.(3개월)': int(ref_qty),
    #                     '일': int(daily_qty),
    #                     'Ref.(3개월)%': f"{ref_rate:+.2f}%",
    #                     '일%': f"{daily_rate:+.2f}%",
    #                     'Gap': f"{gap:+.2f}%"
    #                 })

    #                 gap_data[cret_cd] = gap
    #                 ref_qty_dict_group[cret_cd] = int(ref_qty)
    #                 daily_qty_dict_group[cret_cd] = int(daily_qty)
    #                 ref_rate_dict_group[cret_cd] = ref_rate      
    #                 daily_rate_dict_group[cret_cd] = daily_rate 


    #             # 기존 방식과 동일하게 DataFrame으로 저장
    #             group_table_df = pd.DataFrame(group_table)
    #             if group_table_df.empty:
    #                 group_table_df = pd.DataFrame(columns=['구분', 'Ref.(3개월)', '일', 'Ref.(3개월)%', '일%', 'Gap'])
    #             details['loss_rate_table_by_group'][group] = group_table_df

    #             #  저장
    #             details['rc_hg_ref_qty_by_group'][group] = ref_qty_dict_group
    #             details['rc_hg_daily_qty_by_group'][group] = daily_qty_dict_group
    #             details['rc_hg_ref_rate_by_group'][group] = ref_rate_dict_group      
    #             details['rc_hg_daily_rate_by_group'][group] = daily_rate_dict_group 
    #             details['rc_hg_gap_data_by_group'][group] = gap_data  # 그래프용

    #             fig, axes = plt.subplots(1, 3, figsize=(10, 6), dpi=300)
    #             categories = ['FS', 'RESC', 'HG']

    #             # 각 공정별 Y축 범위를 고정
    #             Y_LIM_MAP = {
    #                 'FS': (0.0, 2.0),   # FS: 0% \~ 2%
    #                 'RESC': (-1.0, 1.0), # RESC: -3% \~ 1%
    #                 'HG': (-1.0, 1.0)    # HG: -3% \~ 1%
    #             }

    #             # 색상
    #             ref_color = '#0000ff'   # 쨍한 파랑
    #             daily_color = '#ff0000' # 쨍한 빨강

    #             for idx, cret_cd in enumerate(categories):
    #                 ax = axes[idx]
                    
    #                 # 데이터 추출
    #                 ref_rate = float(details['rc_hg_ref_rate_by_group'][group].get(cret_cd, 0.0))
    #                 daily_rate = float(details['rc_hg_daily_rate_by_group'][group].get(cret_cd, 0.0))
    #                 gap = float(details['rc_hg_gap_data_by_group'][group].get(cret_cd, 0.0))

    #                 # Y축 범위 기반 동적 오프셋 계산
    #                 y_min, y_max = Y_LIM_MAP[cret_cd]
    #                 y_range = y_max - y_min
                    
    #                 label_offset = y_range * 0.025  # 막대 값 라벨: 전체 높이의 2.5%
    #                 gap_offset   = y_range * 0.06   # Gap 라벨: 전체 높이의 6% (더 멀리 배치)

    #                 # 막대 그리기
    #                 x_center = 0
    #                 bar_width = 0.9
    #                 ref_x = x_center - bar_width/2
    #                 daily_x = x_center + bar_width/2

    #                 bar_ref = ax.bar(ref_x, ref_rate, bar_width, color=ref_color)
    #                 bar_daily = ax.bar(daily_x, daily_rate, bar_width, color=daily_color)
                    
    #                 # 통일된 Y축 범위 적용
    #                 y_min, y_max = Y_LIM_MAP[cret_cd]
    #                 ax.set_ylim(y_min, y_max)
                                        
    #                 # ✅ 막대 값 라벨 함수: 양수/음수 동일한 거리로
    #                 def add_internal_label(bar, value):
    #                     for rect in bar:
    #                         height = rect.get_height()
    #                         if height >= 0:
    #                             pos_y = float(height + label_offset)
    #                             va = 'bottom'
    #                         else:
    #                             pos_y = float(height - label_offset)
    #                             va = 'top'
    #                         ax.text(
    #                             rect.get_x() + rect.get_width() / 2.,
    #                             pos_y,
    #                             f'{value:.2f}%',
    #                             ha='center', va=va,
    #                             fontsize=15, fontweight='bold',
    #                             color='black',
    #                             zorder=4
    #                         )
                    
    #                 add_internal_label(bar_ref, ref_rate)
    #                 add_internal_label(bar_daily, daily_rate)

    #                 max_height = max(ref_rate, daily_rate)
    #                 min_height = min(ref_rate, daily_rate)

    #                 # ✅ Gap 라벨 위치 결정: 막대 값의 부호 기준 (Gap 부호 무시)
    #                 if ref_rate >= 0 and daily_rate >= 0:
    #                     # 두 값 모두 양수 → Gap 라벨은 무조건 위쪽 배치
    #                     gap_y = float(max_height + gap_offset)
    #                     va_align = 'bottom'
    #                 elif ref_rate <= 0 and daily_rate <= 0:
    #                     # 두 값 모두 음수 → Gap 라벨은 무조건 아래쪽 배치
    #                     gap_y = float(min_height - gap_offset)
    #                     va_align = 'top'
    #                 else:
    #                     # 혼합 (양수/음수) → 기존 로직 유지 (Gap 부호 기준)
    #                     if gap >= 0:
    #                         gap_y = float(max_height + gap_offset)
    #                         va_align = 'bottom'
    #                     else:
    #                         gap_y = float(min_height - gap_offset)
    #                         va_align = 'top'

    #                 gap_color = '#ff0000' if gap >= 0 else '#0000ff'
    #                 sign = '+' if gap >= 0 else ''
    #                 ax.text(x_center, gap_y, f'{sign}{gap:.2f}%', ha='center', va=va_align,
    #                         fontsize=15, fontweight='bold', color=gap_color)

    #                 # X 축
    #                 ax.set_xticks([ref_x, daily_x])
    #                 ax.set_xticklabels(['Ref.', '일'], fontsize=20, fontweight='bold')
                    
    #                 # 제목
    #                 ax.set_title(cret_cd, fontsize=20, fontweight='bold', pad=10)
    #                 ax.yaxis.set_major_formatter(plt.FuncFormatter(lambda x, p: f'{x:.1f}%'))
    #                 ax.tick_params(axis='y', labelsize=14)
    #                 #  그리드 및 축 설정 유지
    #                 ax.grid(axis='y', linestyle='--', alpha=0.7, zorder=0)
    #                 ax.set_axisbelow(True)
    #                 ax.spines['top'].set_visible(False)
    #                 ax.spines['right'].set_visible(False)

    #             # 전체 제목
    #             fig.suptitle(f'RC/HG 보상 ({group})', fontsize=14, fontweight='bold', y=1.05)
                
    #             # 여백 최적화
    #             plt.tight_layout(pad=0.8)

    #             graph_path = debug_dir / f"RC_HG_보상_{group}.png"
    #             if graph_path.exists():
    #                 graph_path.unlink()
    #                 print(f"기존 그래프 파일 삭제됨: {graph_path}")

    #             plt.savefig(graph_path, dpi=300, bbox_inches='tight')
    #             plt.close()

    #             details['rc_hg_gap_chart_path_by_group'][group] = str(graph_path)

    #         # ===================================================================
    #         #  7. 전체 RC/HG 보상 그래프 생성
    #         # ===================================================================
    #         total_gap_data = {}
    #         total_ref_data = {}
    #         total_daily_data = {}

    #         for row in report_table_total:
    #             if row['구분'] in ['FS', 'RESC', 'HG']:
    #                 gap_str = row['Gap'].replace('%', '').replace('+', '')
    #                 total_gap_data[row['구분']] = float(gap_str)
    #                 ref_str = row['Ref.(3개월)%'].replace('%', '').replace('+', '')
    #                 daily_str = row['일%'].replace('%', '').replace('+', '')
    #                 total_ref_data[row['구분']] = float(ref_str)
    #                 total_daily_data[row['구분']] = float(daily_str)

    #         fig, axes = plt.subplots(1, 3, figsize=(10, 6), dpi=300)
    #         categories = ['FS', 'RESC', 'HG']
    #         # ──────────────────────────────────────────────────
    #         # 전체 데이터 기준 Y축 범위 계산
    #         # ──────────────────────────────────────────────────
    #         all_vals = []
    #         for cret_cd in categories:
    #             ref_rate = total_ref_data.get(cret_cd, 0)
    #             daily_rate = total_daily_data.get(cret_cd, 0)
    #             all_vals.extend([ref_rate, daily_rate, 0])

    #         # 색상
    #         ref_color = '#0000ff'   # 쨍한 파랑
    #         daily_color = '#ff0000' # 쨍한 빨강

    #         # 각 공정별 Y축 범위 정의 (고정)
    #         Y_LIM_MAP = {
    #             'FS': (4.0, 12.0),   # 4% \~ 12%
    #             'RESC': (-3.0, 1.0), # -3% \~ 1%
    #             'HG': (-3.0, 1.0)    # -3% \~ 1%
    #         }

    #         GAP_LABEL_RATIO = 0.05  # 3% 오프셋 (사용자 요구사항 반영)

    #         for idx, cret_cd in enumerate(categories):
    #             ax = axes[idx]
                
    #             # 데이터 추출
    #             ref_rate = float(total_ref_data.get(cret_cd, 0))
    #             daily_rate = float(total_daily_data.get(cret_cd, 0))
    #             gap = float(total_gap_data.get(cret_cd, 0.0))
                                
    #             # Gap 라벨 위치 조정 (음수/양수 대응)
    #             y_min, y_max = Y_LIM_MAP[cret_cd]
                
    #             label_offset = (y_max - y_min) * 0.02  # 막대 값 라벨용

    #             # 막대 그리기
    #             x_center = 0
    #             bar_width = 0.9
    #             ref_x = x_center - bar_width/2
    #             daily_x = x_center + bar_width/2

    #             bar_ref = ax.bar(ref_x, ref_rate, bar_width, color=ref_color)
    #             bar_daily = ax.bar(daily_x, daily_rate, bar_width, color=daily_color)
                
    #             def add_internal_label(bar, value):
    #                 for rect in bar:
    #                     height = rect.get_height()
    #                     if height >= 0:
    #                         pos_y = float(height + label_offset)
    #                         va = 'bottom'
    #                     else:
    #                         pos_y = float(height - label_offset)
    #                         va = 'top'
    #                     ax.text(
    #                         rect.get_x() + rect.get_width() / 2.,
    #                         pos_y,
    #                         f'{value:.2f}%',
    #                         ha='center', va=va,
    #                         fontsize=15, fontweight='bold',
    #                         color='black',
    #                         zorder=4
    #                     )
                
    #             add_internal_label(bar_ref, ref_rate)
    #             add_internal_label(bar_daily, daily_rate)

    #             if ref_rate >= 0 and daily_rate >= 0:
    #                 # 양수 막대 (FS) → max 값 위에 Gap 라벨
    #                 max_height = max(ref_rate, daily_rate)
    #                 gap_offset_value = max_height * GAP_LABEL_RATIO
    #                 gap_y = max_height + gap_offset_value + label_offset
    #                 va_align = 'bottom'
    #             else:
    #                 # 음수 막대 (RESC, HG) → min 값 아래에 Gap 라벨
    #                 min_height = min(ref_rate, daily_rate)
    #                 gap_offset_value = abs(min_height) * GAP_LABEL_RATIO
    #                 gap_y = min_height - gap_offset_value - label_offset
    #                 va_align = 'top'

    #             gap_color = '#0000ff' if gap >= 0 else '#ff0000'
    #             sign = '+' if gap >= 0 else ''
    #             ax.text(x_center, gap_y, f'{sign}{gap:.2f}%', ha='center', va=va_align,
    #                     fontsize=15, fontweight='bold', color=gap_color)

    #             # X 축
    #             ax.set_xticks([ref_x, daily_x])
    #             ax.set_xticklabels(['Ref.', '일'], fontsize=20, fontweight='bold')
                
    #             # 제목
    #             ax.set_title(cret_cd, fontsize=20, fontweight='bold', pad=10)
    #             y_min, y_max = Y_LIM_MAP[cret_cd] # Y축 범위 고정
    #             ax.set_ylim(y_min, y_max)
    #             ax.yaxis.set_major_formatter(plt.FuncFormatter(lambda x, p: f'{x:.1f}%'))
    #             ax.tick_params(axis='y', labelsize=14)
    #             # 그리드
    #             ax.grid(axis='y', linestyle='--', alpha=0.7, zorder=0)
    #             ax.set_axisbelow(True)
    #             ax.spines['top'].set_visible(False)
    #             ax.spines['right'].set_visible(False)
                
    #         # 전체 제목
    #         fig.suptitle(f'RC/HG 보상 ({group})', fontsize=14, fontweight='bold', y=1.05)
            
    #         # 여백 최적화
    #         plt.tight_layout(pad=0.8)
            
    #         # 저장
    #         total_graph_path = debug_dir / "RC_HG_보상_전체.png"
    #         if total_graph_path.exists():
    #             total_graph_path.unlink()
    #             print(f"기존 전체 그래프 파일 삭제됨: {total_graph_path}")
            
    #         plt.savefig(total_graph_path, dpi=300, bbox_inches=None, pad_inches=0)
    #         plt.close()
            
    #         details['rc_hg_gap_chart_path_total'] = str(total_graph_path)

    #     # ===================================================================
    #     # (2) Prime 제품 전용 분석
    #     # ===================================================================
    #     if 'GRD_CD_NM_CS' in df_cached_3months.columns:
    #         df_cached_prime = df_cached_3months[df_cached_3months['GRD_CD_NM_CS'] == 'Prime']
    #         df_self_prime = df_self_data[df_self_data['GRD_CD_NM_CS'] == 'Prime'] if not df_self_data.empty else pd.DataFrame()

    #         if not df_cached_prime.empty:
    #             denominator_data_prime = df_cached_prime[df_cached_prime['REJ_GROUP'] == '분모']
    #             total_in_qty_prime = denominator_data_prime['IN_QTY'].sum()
    #             avg_in_qty_prime = total_in_qty_prime / total_months if total_in_qty_prime > 0 else 0

    #             if avg_in_qty_prime == 0:
    #                 print("Prime 분모(IN_QTY)가 0입니다. Loss Rate 계산 불가")
    #                 details['prime_avg_in_qty'] = 0
    #             else:
    #                 valid_cached_prime = df_cached_prime[df_cached_prime['REJ_GROUP'].notna()]
    #                 total_loss_by_cret_prime = valid_cached_prime.groupby('CRET_CD')['LOSS_QTY'].sum() / total_months

    #                 daily_loss_by_cret_prime = pd.Series(dtype='int64')
    #                 total_daily_qty_prime = 0
    #                 if not df_self_prime.empty:
    #                     valid_daily_prime = df_self_prime[df_self_prime['REJ_GROUP'].notna()]
    #                     daily_loss_by_cret_prime = valid_daily_prime.groupby('CRET_CD')['LOSS_QTY'].sum()
    #                     denominator_daily_prime = df_self_prime[df_self_prime['REJ_GROUP'] == '분모']
    #                     total_daily_qty_prime = denominator_daily_prime['IN_QTY'].sum()

    #                 report_table_prime = []
    #                 ref_qty_dict_prime = {}
    #                 daily_qty_dict_prime = {}

    #                 for cret_cd in cret_list:
    #                     ref_qty = total_loss_by_cret_prime.get(cret_cd, 0)
    #                     daily_qty = daily_loss_by_cret_prime.get(cret_cd, 0)
    #                     ref_rate = (ref_qty / avg_in_qty_prime) * 100 if avg_in_qty_prime != 0 else 0
    #                     daily_rate = (daily_qty / total_daily_qty_prime) * 100 if total_daily_qty_prime != 0 else 0
    #                     gap = ref_rate - daily_rate

    #                     report_table_prime.append({
    #                         '구분': cret_cd,
    #                         'Ref.(3개월)': int(ref_qty),
    #                         '일': int(daily_qty),
    #                         'Ref.(3개월)%': f"{ref_rate:+.2f}%",
    #                         '일%': f"{daily_rate:+.2f}%",
    #                         'Gap': f"{gap:+.2f}%"
    #                     })
    #                     ref_qty_dict_prime[cret_cd] = int(ref_qty)
    #                     daily_qty_dict_prime[cret_cd] = int(daily_qty)

    #                 ref_qty_dict_prime['모수'] = int(avg_in_qty_prime)
    #                 daily_qty_dict_prime['모수'] = int(total_daily_qty_prime)
    #                 report_table_prime.append({
    #                     '구분': '모수',
    #                     'Ref.(3개월)': ref_qty_dict_prime['모수'],
    #                     '일': daily_qty_dict_prime['모수'],
    #                     'Ref.(3개월)%': "", '일%': "", 'Gap': ""
    #                 })

    #                 details['prime_rc_hg_ref_qty_total'] = ref_qty_dict_prime
    #                 details['prime_rc_hg_daily_qty_total'] = daily_qty_dict_prime
    #                 details['prime_avg_in_qty'] = avg_in_qty_prime
    #                 details['prime_summary'] = pd.DataFrame(report_table_prime)

    #                 details['prime_rc_hg_ref_qty_by_group'] = {}
    #                 details['prime_rc_hg_daily_qty_by_group'] = {}
    #                 details['prime_rc_hg_ref_rate_by_group'] = {}
    #                 details['prime_rc_hg_daily_rate_by_group'] = {}
    #                 details['prime_rc_hg_gap_data_by_group'] = {}
    #                 details['prime_loss_rate_table_by_group'] = {}
    #                 details['prime_rc_hg_gap_chart_path_by_group'] = {}

    #                 for group in ['PARTICLE', 'FLATNESS', 'WARP&BOW', 'NANO']:
    #                     group_data = df_cached_prime[df_cached_prime['REJ_GROUP'] == group]
    #                     group_loss_by_cret = group_data.groupby('CRET_CD')['LOSS_QTY'].sum() / total_months

    #                     group_daily_loss_by_cret = pd.Series(dtype='int64')
    #                     if not df_self_prime.empty:
    #                         group_self_data = df_self_prime[df_self_prime['REJ_GROUP'] == group]
    #                         group_daily_loss_by_cret = group_self_data.groupby('CRET_CD')['LOSS_QTY'].sum()

    #                     group_table = []
    #                     gap_data = {}
    #                     ref_qty_dict_g = {}
    #                     daily_qty_dict_g = {}
    #                     ref_rate_dict_g = {}
    #                     daily_rate_dict_g = {}

    #                     for cret_cd in cret_list:
    #                         ref_qty = group_loss_by_cret.get(cret_cd, 0)
    #                         daily_qty = group_daily_loss_by_cret.get(cret_cd, 0)
    #                         ref_rate = (ref_qty / avg_in_qty_prime) * 100 if avg_in_qty_prime != 0 else 0
    #                         daily_rate = (daily_qty / total_daily_qty_prime) * 100 if total_daily_qty_prime != 0 else 0
    #                         gap = ref_rate - daily_rate

    #                         group_table.append({
    #                             '구분': cret_cd,
    #                             'Ref.(3개월)': int(ref_qty),
    #                             '일': int(daily_qty),
    #                             'Ref.(3개월)%': f"{ref_rate:+.2f}%",
    #                             '일%': f"{daily_rate:+.2f}%",
    #                             'Gap': f"{gap:+.2f}%"
    #                         })
    #                         gap_data[cret_cd] = gap
    #                         ref_qty_dict_g[cret_cd] = int(ref_qty)
    #                         daily_qty_dict_g[cret_cd] = int(daily_qty)
    #                         ref_rate_dict_g[cret_cd] = ref_rate
    #                         daily_rate_dict_g[cret_cd] = daily_rate

    #                     group_table_df = pd.DataFrame(group_table)
    #                     if group_table_df.empty:
    #                         group_table_df = pd.DataFrame(columns=['구분', 'Ref.(3개월)', '일', 'Ref.(3개월)%', '일%', 'Gap'])
    #                     details['prime_loss_rate_table_by_group'][group] = group_table_df

    #                     details['prime_rc_hg_ref_qty_by_group'][group] = ref_qty_dict_g
    #                     details['prime_rc_hg_daily_qty_by_group'][group] = daily_qty_dict_g
    #                     details['prime_rc_hg_ref_rate_by_group'][group] = ref_rate_dict_g
    #                     details['prime_rc_hg_daily_rate_by_group'][group] = daily_rate_dict_g
    #                     details['prime_rc_hg_gap_data_by_group'][group] = gap_data

    #                     fig, axes = plt.subplots(1, 3, figsize=(10, 6), dpi=300)
    #                     Y_LIM_MAP = {
    #                         'FS': (0.0, 2.0),
    #                         'RESC': (-1.0, 1.0),
    #                         'HG': (-1.0, 1.0)
    #                     }
    #                     for idx, cret_cd in enumerate(['FS', 'RESC', 'HG']):
    #                         ax = axes[idx]
    #                         ref_rate = float(ref_rate_dict_g.get(cret_cd, 0.0))
    #                         daily_rate = float(daily_rate_dict_g.get(cret_cd, 0.0))
    #                         gap = float(gap_data.get(cret_cd, 0.0))

    #                         y_min, y_max = Y_LIM_MAP[cret_cd]
    #                         label_offset = (y_max - y_min) * 0.025
    #                         gap_offset = (y_max - y_min) * 0.06

    #                         x_center = 0
    #                         bar_width = 0.9
    #                         ref_x = x_center - bar_width/2
    #                         daily_x = x_center + bar_width/2

    #                         ax.bar(ref_x, ref_rate, bar_width, color='#0000ff')
    #                         ax.bar(daily_x, daily_rate, bar_width, color='#ff0000')
    #                         ax.set_ylim(y_min, y_max)

    #                         def add_label(bar, value):
    #                             for rect in bar:
    #                                 height = rect.get_height()
    #                                 pos_y = height + label_offset if height >= 0 else height - label_offset
    #                                 va = 'bottom' if height >= 0 else 'top'
    #                                 ax.text(rect.get_x() + rect.get_width()/2., pos_y,
    #                                         f'{value:.2f}%', ha='center', va=va, fontsize=15,
    #                                         fontweight='bold', color='black', zorder=4)

    #                         add_label(ax.containers[0], ref_rate)
    #                         add_label(ax.containers[1], daily_rate)

    #                         if ref_rate >= 0 and daily_rate >= 0:
    #                             gap_y = max(ref_rate, daily_rate) + gap_offset
    #                             va_align = 'bottom'
    #                         elif ref_rate <= 0 and daily_rate <= 0:
    #                             gap_y = min(ref_rate, daily_rate) - gap_offset
    #                             va_align = 'top'
    #                         else:
    #                             gap_y = (max(ref_rate, daily_rate) + gap_offset) if gap >= 0 else (min(ref_rate, daily_rate) - gap_offset)
    #                             va_align = 'bottom' if gap >= 0 else 'top'

    #                         gap_color = '#ff0000' if gap >= 0 else '#0000ff'
    #                         sign = '+' if gap >= 0 else ''
    #                         ax.text(x_center, gap_y, f'{sign}{gap:.2f}%', ha='center', va=va_align,
    #                                 fontsize=15, fontweight='bold', color=gap_color)

    #                         ax.set_xticks([ref_x, daily_x])
    #                         ax.set_xticklabels(['Ref.', '일'], fontsize=20, fontweight='bold')
    #                         ax.set_title(cret_cd, fontsize=20, fontweight='bold', pad=10)
    #                         ax.yaxis.set_major_formatter(plt.FuncFormatter(lambda x, p: f'{x:.1f}%'))
    #                         ax.tick_params(axis='y', labelsize=14)
    #                         ax.grid(axis='y', linestyle='--', alpha=0.7)
    #                         ax.set_axisbelow(True)
    #                         ax.spines['top'].set_visible(False)
    #                         ax.spines['right'].set_visible(False)

    #                     fig.suptitle(f'RC/HG 보상 ({group}) - Prime', fontsize=14, fontweight='bold', y=1.05)
    #                     plt.tight_layout(pad=0.8)
    #                     graph_path = debug_dir / f"RC_HG_보상_{group}_Prime.png"
    #                     if graph_path.exists(): graph_path.unlink()
    #                     plt.savefig(graph_path, dpi=300, bbox_inches='tight')
    #                     plt.close()
    #                     details['prime_rc_hg_gap_chart_path_by_group'][group] = str(graph_path)

    #                 total_gap_data_p = {}
    #                 total_ref_data_p = {}
    #                 total_daily_data_p = {}
    #                 for row in details['prime_summary'].to_dict('records'):
    #                     if row['구분'] in ['FS', 'RESC', 'HG']:
    #                         total_gap_data_p[row['구분']] = float(row['Gap'].replace('%', '').replace('+', ''))
    #                         total_ref_data_p[row['구분']] = float(row['Ref.(3개월)%'].replace('%', '').replace('+', ''))
    #                         total_daily_data_p[row['구분']] = float(row['일%'].replace('%', '').replace('+', ''))

    #                 fig, axes = plt.subplots(1, 3, figsize=(10, 6), dpi=300)
    #                 Y_LIM_MAP = {
    #                     'FS': (4.0, 12.0),
    #                     'RESC': (-3.0, 1.0),
    #                     'HG': (-3.0, 1.0)
    #                 }
    #                 for idx, cret_cd in enumerate(['FS', 'RESC', 'HG']):
    #                     ax = axes[idx]
    #                     ref_rate = total_ref_data_p.get(cret_cd, 0)
    #                     daily_rate = total_daily_data_p.get(cret_cd, 0)
    #                     gap = total_gap_data_p.get(cret_cd, 0)
    #                     y_min, y_max = Y_LIM_MAP[cret_cd]
    #                     label_offset = (y_max - y_min) * 0.02

    #                     x_center = 0
    #                     bar_width = 0.9
    #                     ref_x = x_center - bar_width/2
    #                     daily_x = x_center + bar_width/2

    #                     ax.bar(ref_x, ref_rate, bar_width, color='#0000ff')
    #                     ax.bar(daily_x, daily_rate, bar_width, color='#ff0000')
    #                     ax.set_ylim(y_min, y_max)

    #                     def add_label(bar, value):
    #                         for rect in bar:
    #                             height = rect.get_height()
    #                             pos_y = height + label_offset if height >= 0 else height - label_offset
    #                             va = 'bottom' if height >= 0 else 'top'
    #                             ax.text(rect.get_x() + rect.get_width()/2., pos_y,
    #                                     f'{value:.2f}%', ha='center', va=va, fontsize=15,
    #                                     fontweight='bold', color='black', zorder=4)

    #                     add_label(ax.containers[0], ref_rate)
    #                     add_label(ax.containers[1], daily_rate)

    #                     gap_y = (max(ref_rate, daily_rate) + label_offset * 3) if (ref_rate >= 0 or daily_rate >= 0) \
    #                         else (min(ref_rate, daily_rate) - label_offset * 3)
    #                     va_align = 'bottom' if gap >= 0 else 'top'
    #                     gap_color = '#0000ff' if gap >= 0 else '#ff0000'
    #                     sign = '+' if gap >= 0 else ''
    #                     ax.text(x_center, gap_y, f'{sign}{gap:.2f}%', ha='center', va=va_align,
    #                             fontsize=15, fontweight='bold', color=gap_color)

    #                     ax.set_xticks([ref_x, daily_x])
    #                     ax.set_xticklabels(['Ref.', '일'], fontsize=20, fontweight='bold')
    #                     ax.set_title(cret_cd, fontsize=20, fontweight='bold', pad=10)
    #                     ax.yaxis.set_major_formatter(plt.FuncFormatter(lambda x, p: f'{x:.1f}%'))
    #                     ax.tick_params(axis='y', labelsize=14)
    #                     ax.grid(axis='y', linestyle='--', alpha=0.7)
    #                     ax.set_axisbelow(True)
    #                     ax.spines['top'].set_visible(False)
    #                     ax.spines['right'].set_visible(False)

    #                 fig.suptitle(f'RC/HG 보상 (전체) - Prime', fontsize=14, fontweight='bold', y=1.05)
    #                 plt.tight_layout(pad=0.8)
    #                 total_graph_path_p = debug_dir / "RC_HG_보상_전체_Prime.png"
    #                 if total_graph_path_p.exists(): total_graph_path_p.unlink()
    #                 plt.savefig(total_graph_path_p, dpi=300, bbox_inches='tight')
    #                 plt.close()
    #                 details['prime_rc_hg_gap_chart_path_total'] = str(total_graph_path_p)
    #         else:
    #             print("Prime 제품에 대한 3개월 데이터가 없습니다.")
    #     else:
    #         print("GRD_CD_NM_CS 컬럼이 데이터에 없습니다.")

    # else:
    #     # 🔴 이 블록은 절대 삭제하지 마세요 — Total 데이터 없을 때 fallback
    #     details['rc_hg_ref_qty_total'] = {}
    #     details['rc_hg_daily_qty_total'] = {}
    #     details['rc_hg_ref_qty_by_group'] = {}
    #     details['rc_hg_daily_qty_by_group'] = {}
    #     details['rc_hg_avg_in_qty'] = 0
    #     details['rc_hg_gap_chart_path_by_group'] = {}
    #     details['rc_hg_gap_chart_path_total'] = ""
    #     self.avg_in_qty = 0
    #     self.total_daily_qty = 0

    # # ===================================================================
    # # 8. 기본 정보 추가
    # # ===================================================================
    # details['cache_data_available'] = not df_cached_3months.empty
    # details['self_data_available'] = not df_self_data.empty
    # details['cache_total_count'] = len(df_cached_3months) if not df_cached_3months.empty else 0
    # details['self_data_count'] = len(df_self_data) if not df_self_data.empty else 0

    # return details


    def _load_waf_60days_from_mixed_cache(self):
        """
        BASE_DT 기준으로 월별 + 일별 캐시에서 데이터를 로드만 함
        → 이후 _create_DATA_WAF_3210_wafering_300 내에서 
        공정별 REG_DTTM 컬럼 기준으로 재분리
        """
        PROJECT_ROOT = Path(__file__).parent.parent
        cache_dir = PROJECT_ROOT / "data_cache"

        if not cache_dir.exists():
            print(f"[WAF 혼합 로드] 캐시 디렉토리 없음: {cache_dir}")
            return pd.DataFrame()

        # ===================================================================
        # [1] 기준일 설정: 어제 기준 최근 90일 (여유)
        # ===================================================================
        base_date = self.target_date_obj #이미 date 객체
        print(f"base_date:", base_date)
        target_dates_60days = {
            (base_date - timedelta(days=i)).strftime("%Y%m%d")
            for i in range(70)  # 여유 있게 70일 (일별 파일용)
        }
        target_months_90days = {
            (base_date - timedelta(days=i)).strftime("%Y%m")
            for i in range(70)  # 월별 파일용
        }

        # ===================================================================
        # [2] 월별 파일 로드 (예: DATA_WAF_3210_wafering_300_202601.parquet)
        # ===================================================================
        monthly_pattern = "DATA_WAF_3210_wafering_300_*.parquet"
        monthly_files = list(cache_dir.glob(monthly_pattern))

        valid_monthly_files = []
        for file_path in monthly_files:
            try:
                # 파일명에서 마지막 언더스코어 뒤의 6자리 추출 (YYYYMM)
                stem = file_path.stem  # 전체 이름
                parts = stem.split('_')
                if len(parts) < 1:
                    continue
                date_part = parts[-1]  # '202601' or '20260301'

                # 길이로 구분: 6자리 → 월별, 8자리 → 일별
                if len(date_part) == 6 and date_part.isdigit():
                    ym = date_part  # '202601'
                    if ym in target_months_90days:
                        valid_monthly_files.append((file_path, ym))
            except Exception as e:
                print(f"[월별 파일 파싱 실패] {file_path.name}: {e}")

        dfs_monthly = []
        for file_path, ym in valid_monthly_files:
            try:
                df = pd.read_parquet(file_path)
                # BASE_DT가 있는 경우, 해당 월 데이터만 필터링
                if 'BASE_DT' in df.columns:
                    df['BASE_DT'] = df['BASE_DT'].astype(str)
                    df = df[df['BASE_DT'].str.startswith(ym)].copy()

                dfs_monthly.append(df)
            except Exception as e:
                print(f"[월별 로드 실패] {file_path.name}: {e}")

        # ===================================================================
        # [3] 일별 파일 로드 (예: DATA_WAF_3210_wafering_300_20260301.parquet)
        # ===================================================================
        daily_files = []
        for date_str in target_dates_60days:
            file_path = cache_dir / f"DATA_WAF_3210_wafering_300_{date_str}.parquet"
            if file_path.exists():
                daily_files.append(file_path)

        dfs_daily = []
        for file_path in daily_files:
            try:
                df = pd.read_parquet(file_path)
                dfs_daily.append(df)
            except Exception as e:
                print(f"[일별 로드 실패] {file_path.name}: {e}")

        # ===================================================================
        # [4] 병합
        # ===================================================================
        all_dfs = dfs_monthly + dfs_daily
        if not all_dfs:
            print("[WAF 혼합 로드] 사용 가능한 데이터 없음")
            return pd.DataFrame()

        df_combined = pd.concat(all_dfs, ignore_index=True)
        df_fs = df_combined[df_combined['CRET_CD'] == 'FS'].copy() # CRET_CD가 FS인 데이터만 사용.

        df_combined = df_fs

        # ===================================================================
        # 🔍 BASE_DT 기준 데이터 기간 출력
        # ===================================================================
        if 'BASE_DT' in df_combined.columns:
            base_dt_series = pd.to_datetime(df_combined['BASE_DT'], format='%Y%m%d', errors='coerce')
            valid_dates = base_dt_series.dropna()

            if len(valid_dates) > 0:
                start_date = valid_dates.min().strftime('%Y-%m-%d')
                end_date = valid_dates.max().strftime('%Y-%m-%d')
                print(f"[WAF 데이터 기간] BASE_DT 기준: {start_date} \\\~ {end_date}")
            else:
                print("[WAF 데이터 기간] BASE_DT에서 유효한 날짜 없음")
        else:
            print("[WAF 데이터 기간] BASE_DT 컬럼 없음")

        return df_combined

    def _create_DATA_WAF_3210_wafering_300(self):
        """
        3210 WAF 상세 분석 - 캐시된 3개월 데이터 + self.data의 당일 데이터 모두 활용
        """
        details = {}
        PROJECT_ROOT = Path(__file__).parent.parent

        # ===================================================================
        # 1. [신규] data_cache에서 3개월 데이터 직접 로드 (장기 분석용)
        # ===================================================================
        # base_date = (datetime.now().date() - timedelta(days=1))
        base_date = self.target_date_obj
        target_months = []
        current = base_date.replace(day=1)
        for _ in range(3):
            current = (current - timedelta(days=1)).replace(day=1)
            month_str = current.strftime("%Y%m")
            target_months.append(month_str)
        target_months = sorted(target_months)  # 과거 → 최근

        # print(f"[WAF 캐시 필터링] 최근 3개월 대상 월: {target_months}")

        cache_dir = PROJECT_ROOT / "data_cache"
        pattern = "DATA_WAF_3210_wafering_300_*.parquet"
        parquet_files = list(cache_dir.glob(pattern))

        df_cached_3months = pd.DataFrame()

        if parquet_files:
            valid_files = []
            for file_path in parquet_files:
                try:
                    stem = file_path.stem
                    date_part = stem.split('_')[-1]  # '202506'

                    if len(date_part) != 6 or not date_part.isdigit():
                        continue

                    file_ym = date_part
                except Exception as e:
                    print(f"[WAF 캐시] {file_path.name}에서 월 정보 추출 실패 → 건너뜀: {e}")
                    continue

                if file_ym in target_months:
                    valid_files.append(file_path)

            # print(f"[WAF 캐시 필터링] 전체 {len(parquet_files)}개 중 대상 {len(valid_files)}개 파일 선정: {[f.name for f in valid_files]}")

            dfs = []
            for file_path in valid_files:
                try:
                    df_part = pd.read_parquet(file_path)
                    print(f"[WAF 캐시] {file_path.name} 로드 완료: {len(df_part):,} 건")
                    dfs.append(df_part)
                except Exception as e:
                    print(f"[WAF 캐시] {file_path.name} 읽기 실패: {e}")

            if dfs:
                df_cached_3months = pd.concat(dfs, ignore_index=True)
                print(f"[WAF 캐시] 총 {len(df_cached_3months):,} 건 데이터 병합 완료")
            else:
                print("[WAF 캐시] 모든 파일 로드 실패 → 3개월 데이터 없음")
        else:
            print("[WAF 캐시] data_cache에 DATA_WAF_3210_wafering_300_*.parquet 파일 없음")

        # ===================================================================
        # 2. [기존] self.data에서 당일 데이터 사용
        # ===================================================================
        df_self_data = pd.DataFrame()
        if 'DATA_WAF_3210_wafering_300' in self.data and not self.data['DATA_WAF_3210_wafering_300'].empty:
            df_self_data = self.data['DATA_WAF_3210_wafering_300'].copy()
            print(f"[self.data] DATA_WAF_3210_wafering_300 데이터 건수: {len(df_self_data):,} 건")
        else:
            print("[self.data] DATA_WAF_3210_wafering_300 없거나 빈 데이터")

        # ===================================================================
        # 3. [핵심] PRODUCT_TYPE 병합
        # ===================================================================
        if not df_cached_3months.empty:
            df_cached_3months = self._merge_product_type(df_cached_3months)

        if not df_self_data.empty:
            df_self_data = self._merge_product_type(df_self_data)

        print(f"[WAF] PRODUCT_TYPE 병합 완료: 3개월 {df_cached_3months['PRODUCT_TYPE'].notna().sum()}건, 당일 {df_self_data['PRODUCT_TYPE'].notna().sum()}건")

        # ===================================================================
        # 3. [핵심] 3개월 데이터 기반 Loss Rate 분석
        # ===================================================================

        avg_in_qty = getattr(self, 'avg_in_qty', 0)
        total_daily_qty = getattr(self, 'total_daily_qty', 0)

        # 강제 float 변환 (Decimal → float)
        try:
            avg_in_qty = float(avg_in_qty)
            total_daily_qty = float(total_daily_qty)
        except (TypeError, ValueError) as e:
            return details  # 분석 중단

        if avg_in_qty == 0:
            print(" 분모(IN_QTY)가 0입니다. Loss Rate 계산 불가")
            return details
        
        if total_daily_qty == 0:
            print("total_daily_qty = 0 → Daily 분석 생략 (Ref만 분석)")

        def calculate_loss_metrics(group_data, eqp_col, denominator):
            """
            실제 LOSS_QTY 합계 (Count) 와 불량률 (Rate) 을 모두 반환
            """
            if eqp_col not in group_data.columns or denominator == 0:
                return {}
            
            valid = group_data.dropna(subset=[eqp_col]).copy()
            if valid.empty:
                return {}
            
            # LOSS_QTY numeric 변환
            valid.loc[:,'LOSS_QTY'] = pd.to_numeric(valid['LOSS_QTY'], errors='coerce').fillna(0.0).astype(float)
            
            # 장비별 LOSS_QTY 합계
            loss_sum = valid.groupby(eqp_col)['LOSS_QTY'].sum()
            
            #  count (int) 와 rate (float, %) 를 모두 반환
            return {
                eqp: {
                    'count': int(qty), 
                    'rate': round(qty / denominator * 100, 4)  # % 단위 (예: 3.05)
                } 
                for eqp, qty in loss_sum.items()
            }

        # ===================================================================
        # [분석] Ref(3개월) 장비별 불량률 계산
        # ===================================================================

        ref_results = {}
        daily_results = {}

        # 1) PIT
        df_pit = df_cached_3months[df_cached_3months['REJ_GROUP'] == 'PIT']
        if not df_pit.empty:
            eqp_col = ['EQP_NM_300_WF_3670']
            pit_ref = {}
            for eqp in eqp_col:
                res = calculate_loss_metrics(df_pit, eqp, avg_in_qty)
                if res:
                    pit_ref[eqp] = res
            ref_results['PIT'] = pit_ref


        df_pit_d = df_self_data[df_self_data['REJ_GROUP'] == 'PIT']
        if not df_pit_d.empty:
            eqp_col = ['EQP_NM_300_WF_3670']
            pit_daily = {}
            for eqp in eqp_col:
                res = calculate_loss_metrics(df_pit_d, eqp, avg_in_qty)
                if res:
                    pit_daily[eqp] = res
            daily_results['PIT'] = pit_daily

        # 2) SCRATCH
        df_scratch = df_cached_3months[df_cached_3months['REJ_GROUP'] == 'SCRATCH']
        if not df_scratch.empty:
            eqps_scratch = ['EQP_NM_300_WF_3670', 'EQP_NM_300_WF_6100', 'EQP_NM_300_WF_7000'] #3670(LAP), 6100(DSP), 7000(EBIS)
            scratch_ref  = {}
            for eqp in eqps_scratch:
                res = calculate_loss_metrics(df_scratch, eqp, avg_in_qty)
                if res:
                    scratch_ref[eqp] = res
            ref_results['SCRATCH'] = scratch_ref

        df_scratch_d = df_self_data[df_self_data['REJ_GROUP'] == 'SCRATCH']
        if not df_scratch_d.empty:
            eqps_scratch = ['EQP_NM_300_WF_3670', 'EQP_NM_300_WF_6100', 'EQP_NM_300_WF_7000']
            scratch_daily = {}
            for eqp in eqps_scratch:
                res = calculate_loss_metrics(df_scratch_d, eqp, total_daily_qty)
                if res:
                    scratch_daily[eqp] = res
            daily_results['SCRATCH'] = scratch_daily


        # 3) EDGE
        df_edge = df_cached_3months[df_cached_3months['REJ_GROUP'] == 'EDGE']
        if not df_edge.empty:
            eqps = ['EQP_NM_300_WF_3335', 'EQP_NM_300_WF_3696', 'EQP_NM_300_WF_7000']
            edge_ref = {}
            for eqp in eqps:
                res = calculate_loss_metrics(df_edge, eqp, avg_in_qty)
                if res:
                    edge_ref[eqp] = res
            ref_results['EDGE'] = edge_ref

        df_edge_d = df_self_data[df_self_data['REJ_GROUP'] == 'EDGE']
        if not df_edge_d.empty:
            eqps = ['EQP_NM_300_WF_3335', 'EQP_NM_300_WF_3696', 'EQP_NM_300_WF_7000']
            edge_daily = {}
            for eqp in eqps:
                res = calculate_loss_metrics(df_edge_d, eqp, total_daily_qty)
                if res:
                    edge_daily[eqp] = res
            daily_results['EDGE'] = edge_daily


        # 4) BROKEN
        # AFT_BAD_RSN_CD → MID_GROUP 매핑
        BROKEN_MID_MAPPING = REJ_GROUP_TO_MID_MAPPING.get('BROKEN', {})

        # 결과 저장
        ref_results['BROKEN'] = {}
        daily_results['BROKEN'] = {}


        # 중간 결과 저장 (전체 비교용)
        broken_ref_list = []
        broken_daily_list = []

        # df_broken 필터링
        df_broken = df_cached_3months[df_cached_3months['REJ_GROUP'] == 'BROKEN']
        df_broken_d = df_self_data[df_self_data['REJ_GROUP'] == 'BROKEN']

        # ===================================================================
        # [1] REF 데이터 처리 (각 MID_GROUP 내 장비별 상위 3개)
        # ===================================================================
        if not df_broken.empty:
            # MID_GROUP 생성
            df_broken['MID_GROUP'] = df_broken['AFT_BAD_RSN_CD'].map(BROKEN_MID_MAPPING)
            df_broken = df_broken.dropna(subset=['MID_GROUP'])
            df_broken['MID_GROUP'] = df_broken['MID_GROUP'].astype(str).str.strip().str.upper()
            df_broken = df_broken.reset_index(drop=True)  # 중복 인덱스 방지

            for mid_group, group_df in df_broken.groupby('MID_GROUP'):
                if mid_group not in MID_TO_EQP:
                    continue

                eqp_col = MID_TO_EQP[mid_group]

                if eqp_col not in group_df.columns:
                    # 컬럼 없으면 전체 LOSS_QTY 합산 (비상)
                    loss_qty = pd.to_numeric(group_df['LOSS_QTY'], errors='coerce').sum()
                    total_rate = (loss_qty / avg_in_qty * 100)
                    broken_ref_list.append({
                        'EQP': f"{mid_group}_UNKNOWN",
                        'count': float(loss_qty),
                        'in_qty': avg_in_qty,
                        'rate': round(total_rate, 4),
                        'MID_GROUP': mid_group,
                        'AFT_BAD_RSN_CD': f"{mid_group}_ALL"
                    })
                else:
                    # 장비별 분해
                    res_dict = calculate_loss_metrics(group_df, eqp_col, avg_in_qty)
                    if not res_dict:
                        continue

                    for eqp_name, metrics in res_dict.items():
                        if pd.isna(eqp_name) or str(eqp_name).strip() == '':
                            continue
                        broken_ref_list.append({
                            'EQP': eqp_name,
                            'count': metrics['count'],
                            'in_qty': avg_in_qty,
                            'rate': metrics['rate'],
                            'MID_GROUP': mid_group,
                            'AFT_BAD_RSN_CD': f"{mid_group}_{eqp_name}"
                        })

        # ===================================================================
        # [2] Daily 데이터 처리 (장비별 상위 3개)
        # ===================================================================
        if not df_broken_d.empty:
            df_broken_d['MID_GROUP'] = df_broken_d['AFT_BAD_RSN_CD'].map(BROKEN_MID_MAPPING)
            df_broken_d = df_broken_d.dropna(subset=['MID_GROUP'])
            df_broken_d['MID_GROUP'] = df_broken_d['MID_GROUP'].astype(str).str.strip().str.upper()
            df_broken_d = df_broken_d.reset_index(drop=True)

            for mid_group, group_df in df_broken_d.groupby('MID_GROUP'):
                if mid_group not in MID_TO_EQP:
                    continue

                eqp_col = MID_TO_EQP[mid_group]

                if eqp_col not in group_df.columns:
                    loss_qty = pd.to_numeric(group_df['LOSS_QTY'], errors='coerce').sum()
                    total_rate = (loss_qty / total_daily_qty * 100)
                    broken_daily_list.append({
                        'EQP': f"{mid_group}_UNKNOWN",
                        'count': float(loss_qty),
                        'in_qty': total_daily_qty,
                        'rate': round(total_rate, 4),
                        'MID_GROUP': mid_group,
                        'AFT_BAD_RSN_CD': f"{mid_group}_ALL"
                    })
                else:
                    res_dict = calculate_loss_metrics(group_df, eqp_col, total_daily_qty)
                    if not res_dict:
                        continue

                    for eqp_name, metrics in res_dict.items():
                        if pd.isna(eqp_name) or str(eqp_name).strip() == '':
                            continue
                        broken_daily_list.append({
                            'EQP': eqp_name,
                            'count': metrics['count'],
                            'in_qty': total_daily_qty,
                            'rate': metrics['rate'],
                            'MID_GROUP': mid_group,
                            'AFT_BAD_RSN_CD': f"{mid_group}_{eqp_name}"
                        })

        # ===================================================================
        # [4] 상위 3개 장비 선정 + ref_results/daily_results 분리 저장
        # ===================================================================
        # 기존 'BROKEN' 제거
        if 'BROKEN' in ref_results:
            del ref_results['BROKEN']
        if 'BROKEN' in daily_results:
            del daily_results['BROKEN']

        # 분리 저장용 딕셔너리
        ref_results['LAP_BROKEN'] = {}
        ref_results['EP_BROKEN'] = {}
        ref_results['DSP_BROKEN'] = {}
        ref_results['FP_BROKEN'] = {}

        daily_results['LAP_BROKEN'] = {}
        daily_results['EP_BROKEN'] = {}
        daily_results['DSP_BROKEN'] = {}
        daily_results['FP_BROKEN'] = {}

        # REF: MID_GROUP별 상위 3개 추출
        if broken_ref_list:
            df_all = pd.DataFrame(broken_ref_list)

            for mid_group, group_df in df_all.groupby('MID_GROUP'):
                top3 = group_df.sort_values('rate', ascending=False).head(3)
                key = f"{mid_group}_BROKEN"

                if key in ref_results:
                    for _, row in top3.iterrows():
                        ref_results[key][row['EQP']] = {
                            'count': row['count'],
                            'in_qty': row['in_qty'],
                            'rate': row['rate'],
                            'MID_GROUP': row['MID_GROUP'],
                            'AFT_BAD_RSN_CD': row['AFT_BAD_RSN_CD']
                        }
 
        # Daily: MID_GROUP별 상위 3개 추출
        if broken_daily_list:
            df_all_d = pd.DataFrame(broken_daily_list)

            for mid_group, group_df in df_all_d.groupby('MID_GROUP'):
                top3 = group_df.sort_values('rate', ascending=False).head(3)
                key = f"{mid_group}_BROKEN"

                if key in daily_results:
                    for _, row in top3.iterrows():
                        daily_results[key][row['EQP']] = {
                            'count': row['count'],
                            'in_qty': row['in_qty'],
                            'rate': row['rate'],
                            'MID_GROUP': row['MID_GROUP'],
                            'AFT_BAD_RSN_CD': row['AFT_BAD_RSN_CD']
                        }

        # 5) CHIP
        df_chip = df_cached_3months[df_cached_3months['REJ_GROUP'] == 'CHIP']

        if not df_chip.empty:
            chip_ref = {}
            # 조건 정의
            cond_edge = df_chip['AFT_BAD_RSN_CD'] == 'EDGE-CHIP'
            cond_lap = df_chip['AFT_BAD_RSN_CD'] == 'CHIP-LAP'
            cond_eg1af = df_chip['AFT_BAD_RSN_CD'] == 'CHIP_EG1AF'
            cond_eg1bf = df_chip['AFT_BAD_RSN_CD'] == 'CHIP_EG1BF'
            cond_echip = df_chip['AFT_BAD_RSN_CD'] == 'E_CHIP'  # 추가

            # 그룹별 처리 (그룹명, 조건, 장비 리스트)
            groups = [
                ('EDGE-CHIP', cond_edge, ['EQP_NM_300_WF_3335', 'EQP_NM_300_WF_3696']),
                ('CHIP-LAP', cond_lap, ['EQP_NM_300_WF_3670']),
                ('CHIP_EG1AF', cond_eg1af, ['EQP_NM_300_WF_3335']),  # 수정: 리스트로 통일
                ('CHIP_EG1BF', cond_eg1bf, ['EQP_NM_300_WF_3300']),
                ('EDGE-ECHIP', cond_echip, ['EQP_NM_300_WF_3335', 'EQP_NM_300_WF_3696', 'EQP_NM_300_WF_7000'])  # 추가
            ]

            for group_name, condition, eqp_list in groups:
                group_df = df_chip[condition]
                if group_df.empty:
                    continue

                # 그룹 초기화
                if group_name not in ref_results:
                    ref_results[group_name] = {}

                has_data = False  # 데이터 유무 플래그

                for eqp in eqp_list:
                    res = calculate_loss_metrics(group_df, eqp, avg_in_qty)
                    if not res:
                        continue
                
                    key = f'{group_name}_{eqp}'
                    chip_ref[key] = res

                    # BROKEN 방식과 동일하게 ref_results[group][equip_id] 저장
                    for equip_id, metrics in res.items():
                        if pd.isna(equip_id) or str(equip_id).strip() == '':
                            continue
                        ref_results[group_name][equip_id] = {
                            'count': metrics['count'],
                            'in_qty': avg_in_qty,
                            'rate': metrics['rate'],
                            'MID_GROUP': group_name,
                            'AFT_BAD_RSN_CD': f"{group_name}_{equip_id}"
                        }
                        has_data = True  # 데이터 있음 표시
                # 데이터가 없으면 ref_results에서 제거
                if not has_data:
                    ref_results.pop(group_name, None)

            # 전체 CHIP 저장 (옵션)
            if chip_ref:
                ref_results['CHIP'] = chip_ref

        df_chip_d = df_self_data[df_self_data['REJ_GROUP'] == 'CHIP']

        if not df_chip_d.empty:
            chip_daily = {}
            cond_edge = df_chip_d['AFT_BAD_RSN_CD'] == 'EDGE-CHIP'
            cond_lap = df_chip_d['AFT_BAD_RSN_CD'] == 'CHIP-LAP'
            cond_eg1af = df_chip_d['AFT_BAD_RSN_CD'] == 'CHIP_EG1AF'
            cond_eg1bf = df_chip_d['AFT_BAD_RSN_CD'] == 'CHIP_EG1BF'
            cond_echip = df_chip_d['AFT_BAD_RSN_CD'] == 'E_CHIP'

            groups = [
                ('EDGE-CHIP', cond_edge, ['EQP_NM_300_WF_3335', 'EQP_NM_300_WF_3696']),
                ('CHIP-LAP', cond_lap, ['EQP_NM_300_WF_3670']),
                ('CHIP_EG1AF', cond_eg1af, ['EQP_NM_300_WF_3335']),
                ('CHIP_EG1BF', cond_eg1bf, ['EQP_NM_300_WF_3300']),
                ('EDGE-ECHIP', cond_echip, ['EQP_NM_300_WF_3335', 'EQP_NM_300_WF_3696', 'EQP_NM_300_WF_7000'])
            ]

            for group_name, condition, eqp_list in groups:
                group_df = df_chip_d[condition]
                if group_df.empty:
                    continue

                if group_name not in daily_results:
                    daily_results[group_name] = {}

                has_data = False
                for eqp in eqp_list:
                    res = calculate_loss_metrics(group_df, eqp, total_daily_qty)
                    if not res:
                        continue

                    key = f'{group_name}_{eqp}'
                    chip_daily[key] = res

                    # daily_results도 동일 구조
                    for equip_id, metrics in res.items():
                        if pd.isna(equip_id) or str(equip_id).strip() == '':
                            continue
                        daily_results[group_name][equip_id] = {
                            'count': metrics['count'],
                            'in_qty': total_daily_qty,
                            'rate': metrics['rate'],
                            'MID_GROUP': group_name,
                            'AFT_BAD_RSN_CD': f"{group_name}_{equip_id}"
                        }
                        has_data = True

                if not has_data:
                    daily_results.pop(group_name, None)

            if chip_daily:
                daily_results['CHIP'] = chip_daily

        # 6) VISUAL
        df_visual = df_cached_3months[df_cached_3months['REJ_GROUP'] == 'VISUAL']
        if not df_visual.empty:
            cond = df_visual['AFT_BAD_RSN_CD'].isin(['B_PARTICLE', 'B_PAR2'])
            visual_filtered = df_visual[cond]
            eqp_col = ['EQP_NM_300_WF_6100']  
            visual_ref = {}
            for eqp in eqp_col:
                if eqp in visual_filtered.columns:
                    res = calculate_loss_metrics(visual_filtered, eqp, avg_in_qty)
                    if res:
                        visual_ref[eqp] = res
            if visual_ref:
                ref_results['VISUAL'] = visual_ref

        # 2) VISUAL (daily_results)
        df_visual_d = df_self_data[df_self_data['REJ_GROUP'] == 'VISUAL']
        if not df_visual_d.empty:
            cond = df_visual_d['AFT_BAD_RSN_CD'].isin(['B_PARTICLE', 'B_PAR2'])
            visual_filtered_d = df_visual_d[cond]
            eqp_col = ['EQP_NM_300_WF_6100']  
            visual_daily = {}
            for eqp in eqp_col:
                if eqp in visual_filtered_d.columns:
                    res = calculate_loss_metrics(visual_filtered_d, eqp, total_daily_qty)
                    if res:
                        visual_daily[eqp] = res
            if visual_daily:
                daily_results['VISUAL'] = visual_daily

        # ===================================================================
        # 8. Gap 계산 (각 공정별 상위 3개 장비만)
        # ===================================================================
        gap_results = {}

        # BROKEN, CHIP 등 분해된 그룹 정의
        BROKEN_SUBGROUPS = ['LAP_BROKEN', 'EP_BROKEN', 'DSP_BROKEN', 'FP_BROKEN']
        CHIP_SUBGROUPS = ['EDGE-CHIP', 'CHIP_EG1BF', 'CHIP_EG1AF', 'CHIP-LAP', 'E_CHIP']

        def extract_process(eqp_col):
            match = re.search(r'(\d{4})$', eqp_col)
            return match.group(1) if match else eqp_col

        # ===================================================================
        # [1] BROKEN: LAP_BROKEN, EP_BROKEN 등 처리
        # ===================================================================
        for group in BROKEN_SUBGROUPS:
            if group not in ref_results:
                print(f"[WARN] ref_results에 {group} 없음")
                continue
            if group not in daily_results:
                print(f"[WARN] daily_results에 {group} 없음")
                continue

            # 공정 추출: 'LAP_BROKEN' → 'LAP'
            proc = group.split('_')[0]
            eqp_col = MID_TO_EQP.get(proc)
            if not eqp_col:
                continue

            ref_dict = ref_results[group]
            daily_dict = daily_results[group]

            gap_col = {}
            for eqp, ref_data in ref_dict.items():
                if not isinstance(ref_data, dict):
                    continue
                # ref_rate: % → 소수 변환
                ref_rate_decimal = ref_data.get('rate', 0.0) / 100.0  # 0.0012 → 0.000012

                # daily_rate: 존재하면 가져오기, 없으면 0
                daily_rate_decimal = 0.0
                if eqp in daily_dict:
                    daily_data = daily_dict[eqp]
                    if isinstance(daily_data, dict):
                        daily_rate_decimal = daily_data.get('rate', 0.0) / 100.0  # 0.0163 → 0.000163

                # gap = daily - ref
                gap = daily_rate_decimal - ref_rate_decimal
                if ref_data or eqp in daily_dict:  #  데이터 존재 여부 확인
                    gap_col[eqp] = round(gap, 6)  # 0도 포함

            # 장비별 gap이 있으면 저장
            if gap_col:
                gap_results[group] = {eqp_col: gap_col}

        for group in CHIP_SUBGROUPS:
            if group not in ref_results or group not in daily_results:
                continue

            ref_dict = ref_results[group] # ex: {'EDGE-CHIP_EQP_NM_300_WF_3335': {'BSEG02': {...}, ...}}
            daily_dict = daily_results[group]
            # group에 따라 해당 장비 컬럼 결정 → mapping.py 기반
            gap_sub = {}
            # 그룹 → 장비 컬럼 목록 매핑
            if group in ['CHIP-LAP', 'EDGE-CHIP', 'CHIP_EG1AF']:
                eqp_cols = ['EQP_NM_300_WF_3335', 'EQP_NM_300_WF_3696']
            elif group == 'CHIP_EG1BF':
                eqp_cols = ['EQP_NM_300_WF_3300']
            elif group == 'E_CHIP':
                eqp_cols = ['EQP_NM_300_WF_3335', 'EQP_NM_300_WF_3696', 'EQP_NM_300_WF_7000']
            else:
                continue

            # 각 장비 컬럼별 gap_col 생성
            for eqp_col in eqp_cols:
                gap_col = {}

                # ref_dict는 equip_id 기준 → 바로 순회
                for equip_id, ref_item in ref_dict.items():
                    if not isinstance(ref_item, dict):
                        continue

                    ref_rate = ref_item.get('rate', 0.0) / 100.0

                    daily_rate = 0.0
                    if equip_id in daily_dict:
                        daily_item = daily_dict[equip_id]
                        if isinstance(daily_item, dict):
                            daily_rate = daily_item.get('rate', 0.0) / 100.0

                    gap = daily_rate - ref_rate
                    gap_col[equip_id] = round(gap, 6)

                if gap_col:
                    gap_sub[eqp_col] = gap_col

            if gap_sub:
                gap_results[group] = gap_sub
                print(f"[INFO] {group} → gap_results 저장됨: {list(gap_sub.keys())}")
            else:
                print(f"[WARN] {group} → gap_sub 비어있음")

        # ===================================================================
        # [3] 나머지 그룹: SCRATCH, EDGE, PIT, VISUAL 등
        # ===================================================================
        SIMPLE_GROUPS = ['SCRATCH', 'EDGE', 'PIT', 'VISUAL']

        for group in SIMPLE_GROUPS:
            if group not in ref_results or group not in daily_results:
                continue

            ref_dict = ref_results[group]
            daily_dict = daily_results[group]

            if not isinstance(ref_dict, dict):
                continue

            gap_sub = {}
            for eqp_col, eqp_rates in ref_dict.items():
                if eqp_col not in daily_dict:
                    daily_rates = {}
                else:
                    daily_rates = daily_dict[eqp_col]

                if not isinstance(eqp_rates, dict):
                    continue

                gap_col = {}
                for eqp, data in eqp_rates.items():
                    if not isinstance(data, dict):
                        continue
                    ref_rate = data.get('rate', 0.0) / 100.0

                    daily_rate = 0.0
                    if isinstance(daily_rates, dict) and eqp in daily_rates:
                        daily_data = daily_rates[eqp]
                        if isinstance(daily_data, dict):
                            daily_rate = daily_data.get('rate', 0.0) / 100.0

                    gap = daily_rate - ref_rate
                    gap_col[eqp] = round(gap, 6)

                if gap_col:
                    gap_sub[eqp_col] = gap_col

            if gap_sub:
                gap_results[group] = gap_sub
            

        # [핵심] 60일 데이터 로드 (BASE_DT 기준)
        df_60days_raw = self._load_waf_60days_from_mixed_cache()
        if df_60days_raw.empty:
            print("[WAF] 60일 원본 데이터 없음")
            # 기존 로직 계속
        else:
            # print(f"[WAF] 60일 원본 데이터 로드 완료: {len(df_60days_raw):,} 건")
            # 공정별 데이터셋 생성
            process_datasets = {}
            # 1. 3300
            eqp_col_3300 = 'EQP_NM_300_WF_3300'
            reg_col_3300 = 'REG_DTTM_300_WF_3300'
            if reg_col_3300 in df_60days_raw.columns:
                df_3300 = df_60days_raw.dropna(subset=[reg_col_3300, eqp_col_3300]).copy()
                df_3300['process_datetime'] = pd.to_datetime(
                    df_3300[reg_col_3300],
                    format='%Y%m%d%H%M%S%f',
                    errors='coerce'
                )

                # 시간 기준 보정: 07:00 \~ 익일 06:59 까지를 전날 기준일로 묶기 위해 -7시간 이동
                df_3300['base_dt_7hours'] = (df_3300['process_datetime'] - pd.Timedelta(hours=7)).dt.strftime('%Y%m%d') 

                selected_cols = ['process_datetime','base_dt_7hours','REJ_GROUP', 'BASE_DT','LOSS_QTY', eqp_col_3300]
                missing_cols = [c for c in selected_cols if c not in df_3300.columns]
                if missing_cols:
                    print(f"[3300] 누락된 컬럼: {missing_cols} → 건너뜀")
                else:
                    df_3300 = df_3300[selected_cols].rename(columns={eqp_col_3300: 'EQP_NM'})
                    process_datasets['3300'] = df_3300

            # 2. 3335
            eqp_col_3335 = 'EQP_NM_300_WF_3335'
            reg_col_3335 = 'REG_DTTM_300_WF_3335'
            if reg_col_3335 in df_60days_raw.columns:
                df_3335 = df_60days_raw.dropna(subset=[reg_col_3335, eqp_col_3335]).copy()
                df_3335['process_datetime'] = pd.to_datetime(
                    df_3335[reg_col_3335],
                    format='%Y%m%d%H%M%S%f',
                    errors='coerce'
                )

                # 시간 기준 보정: 07:00 \~ 익일 06:59 까지를 전날 기준일로 묶기 위해 -7시간 이동
                df_3335['base_dt_7hours'] = (df_3335['process_datetime'] - pd.Timedelta(hours=7)).dt.strftime('%Y%m%d') 

                selected_cols = ['process_datetime','base_dt_7hours','REJ_GROUP', 'BASE_DT','LOSS_QTY', eqp_col_3335]
                missing_cols = [c for c in selected_cols if c not in df_3335.columns]
                if missing_cols:
                    print(f"[3335] 누락된 컬럼: {missing_cols} → 건너뜀")
                else:
                    df_3335 = df_3335[selected_cols].rename(columns={eqp_col_3335: 'EQP_NM'})
                    process_datasets['3335'] = df_3335

            # 3. 3670
            eqp_col_3670 = 'EQP_NM_300_WF_3670'
            reg_col_3670 = 'REG_DTTM_300_WF_3670'
            if reg_col_3670 in df_60days_raw.columns:
                df_3670 = df_60days_raw.dropna(subset=[reg_col_3670, eqp_col_3670]).copy()
                df_3670['process_datetime'] = pd.to_datetime(
                    df_3670[reg_col_3670],
                    format='%Y%m%d%H%M%S%f',
                    errors='coerce'
                )

                # 시간 기준 보정: 07:00 \~ 익일 06:59 까지를 전날 기준일로 묶기 위해 -7시간 이동
                df_3670['base_dt_7hours'] = (df_3670['process_datetime'] - pd.Timedelta(hours=7)).dt.strftime('%Y%m%d') 

                # 3개 컬럼만 선택 + EQP 컬럼 이름 통일
                selected_cols = ['process_datetime','base_dt_7hours','REJ_GROUP', 'BASE_DT','LOSS_QTY', eqp_col_3670]
                missing_cols = [c for c in selected_cols if c not in df_3670.columns]
                if missing_cols:
                    print(f"[3670] 누락된 컬럼: {missing_cols} → 건너뜀")
                else:
                    df_3670 = df_3670[selected_cols].rename(columns={eqp_col_3670: 'EQP_NM'})
                    process_datasets['3670'] = df_3670


            # 4. 3696
            eqp_col_3696 = 'EQP_NM_300_WF_3696'
            reg_col_3696 = 'REG_DTTM_300_WF_3696'
            if reg_col_3696 in df_60days_raw.columns:
                df_3696 = df_60days_raw.dropna(subset=[reg_col_3696, eqp_col_3696]).copy()
                df_3696['process_datetime'] = pd.to_datetime(
                    df_3696[reg_col_3696],
                    format='%Y%m%d%H%M%S%f',
                    errors='coerce'
                )

                # 시간 기준 보정: 07:00 \~ 익일 06:59 까지를 전날 기준일로 묶기 위해 -7시간 이동
                df_3696['base_dt_7hours'] = (df_3696['process_datetime'] - pd.Timedelta(hours=7)).dt.strftime('%Y%m%d') 

                selected_cols = ['process_datetime','base_dt_7hours','REJ_GROUP','BASE_DT', 'LOSS_QTY', eqp_col_3696]
                missing_cols = [c for c in selected_cols if c not in df_3696.columns]
                if missing_cols:
                    print(f"[3696] 누락된 컬럼: {missing_cols} → 건너뜀")
                else:
                    df_3696 = df_3696[selected_cols].rename(columns={eqp_col_3696: 'EQP_NM'})
                    process_datasets['3696'] = df_3696


            # 5. 6100
            eqp_col_6100 = 'EQP_NM_300_WF_6100'
            reg_col_6100 = 'REG_DTTM_300_WF_6100'
            if reg_col_6100 in df_60days_raw.columns:
                df_6100 = df_60days_raw.dropna(subset=[reg_col_6100, eqp_col_6100]).copy()
                df_6100['process_datetime'] = pd.to_datetime(
                    df_6100[reg_col_6100],
                    format='%Y%m%d%H%M%S%f',
                    errors='coerce'
                )

                # 시간 기준 보정: 07:00 \~ 익일 06:59 까지를 전날 기준일로 묶기 위해 -7시간 이동
                df_6100['base_dt_7hours'] = (df_6100['process_datetime'] - pd.Timedelta(hours=7)).dt.strftime('%Y%m%d') 

                selected_cols = ['process_datetime','base_dt_7hours','REJ_GROUP', 'BASE_DT','LOSS_QTY', eqp_col_6100]
                missing_cols = [c for c in selected_cols if c not in df_6100.columns]
                if missing_cols:
                    print(f"[6100] 누락된 컬럼: {missing_cols} → 건너뜀")
                else:
                    df_6100 = df_6100[selected_cols].rename(columns={eqp_col_6100: 'EQP_NM'})
                    process_datasets['6100'] = df_6100

            # 6. 6210
            eqp_col_6210 = 'EQP_NM_300_WF_6210'
            reg_col_6210 = 'REG_DTTM_300_WF_6210'
            if reg_col_6210 in df_60days_raw.columns:
                df_6210 = df_60days_raw.dropna(subset=[reg_col_6210, eqp_col_6210]).copy()
                df_6210['process_datetime'] = pd.to_datetime(
                    df_6210[reg_col_6210],
                    format='%Y%m%d%H%M%S%f',
                    errors='coerce'
                )

                # 시간 기준 보정: 07:00 \~ 익일 06:59 까지를 전날 기준일로 묶기 위해 -7시간 이동
                df_6210['base_dt_7hours'] = (df_6210['process_datetime'] - pd.Timedelta(hours=7)).dt.strftime('%Y%m%d') 

                selected_cols = ['process_datetime','base_dt_7hours','REJ_GROUP', 'BASE_DT','LOSS_QTY', eqp_col_6210]
                missing_cols = [c for c in selected_cols if c not in df_6210.columns]
                if missing_cols:
                    print(f"[6210] 누락된 컬럼: {missing_cols} → 건너뜀")
                else:
                    df_6210 = df_6210[selected_cols].rename(columns={eqp_col_6210: 'EQP_NM'})
                    process_datasets['6210'] = df_6210

            # 7. 6500
            eqp_col_6500 = 'EQP_NM_300_WF_6500'
            reg_col_6500 = 'REG_DTTM_300_WF_6500'
            if reg_col_6500 in df_60days_raw.columns:
                df_6500 = df_60days_raw.dropna(subset=[reg_col_6500, eqp_col_6500]).copy()
                df_6500['process_datetime'] = pd.to_datetime(
                    df_6500[reg_col_6500],
                    format='%Y%m%d%H%M%S%f',
                    errors='coerce'
                )

                # 시간 기준 보정: 07:00 \~ 익일 06:59 까지를 전날 기준일로 묶기 위해 -7시간 이동
                df_6500['base_dt_7hours'] = (df_6500['process_datetime'] - pd.Timedelta(hours=7)).dt.strftime('%Y%m%d') 

                selected_cols = ['process_datetime','base_dt_7hours','REJ_GROUP', 'BASE_DT','LOSS_QTY', eqp_col_6500]
                missing_cols = [c for c in selected_cols if c not in df_6500.columns]
                if missing_cols:
                    print(f"[6500] 누락된 컬럼: {missing_cols} → 건너뜀")
                else:
                    df_6500 = df_6500[selected_cols].rename(columns={eqp_col_6500: 'EQP_NM'})
                    process_datasets['6500'] = df_6500

            # 8. 7000
            eqp_col_7000 = 'EQP_NM_300_WF_7000'
            reg_col_7000 = 'REG_DTTM_300_WF_7000'
            if reg_col_7000 in df_60days_raw.columns:
                df_7000 = df_60days_raw.dropna(subset=[reg_col_7000, eqp_col_7000]).copy()
                df_7000['process_datetime'] = pd.to_datetime(
                    df_7000[reg_col_7000],
                    format='%Y%m%d%H%M%S%f',
                    errors='coerce'
                )

                # 시간 기준 보정: 07:00 \~ 익일 06:59 까지를 전날 기준일로 묶기 위해 -7시간 이동
                df_7000['base_dt_7hours'] = (df_7000['process_datetime'] - pd.Timedelta(hours=7)).dt.strftime('%Y%m%d') 

                selected_cols = ['process_datetime','base_dt_7hours','REJ_GROUP', 'BASE_DT','LOSS_QTY', eqp_col_7000]
                missing_cols = [c for c in selected_cols if c not in df_7000.columns]
                if missing_cols:
                    print(f"[7000] 누락된 컬럼: {missing_cols} → 건너뜀")
                else:
                    df_7000 = df_7000[selected_cols].rename(columns={eqp_col_7000: 'EQP_NM'})
                    process_datasets['7000'] = df_7000

            # 저장
            self.data['WAF_PROCESS_DATASETS'] = process_datasets

        # ===================================================================
        # 9. details에 저장
        # ===================================================================

        ref_results = ref_results if isinstance(ref_results, dict) else {}
        daily_results = daily_results if isinstance(daily_results, dict) else {}
        gap_results = gap_results if isinstance(gap_results, dict) else {}

        details['waf_analysis_ref'] = ref_results
        details['waf_analysis_daily'] = daily_results
        details['waf_analysis_gap'] = gap_results
        details['df_cached_3months'] = df_cached_3months
        details['df_self_data'] = df_self_data
        details['avg_in_qty'] = avg_in_qty
        details['total_daily_qty'] = total_daily_qty

        self.data['waf_analysis_gap'] = gap_results
        self.data['DATA_3210_wafering_300_details'] = details  # top3_rej_groups 포함

        return details
    

    def _create_DATA_1511_SMAX_wafering_300(self):
        """
        self.data['DATA_1511_SMAX_wafering_300'] 에서
        공정별, 장비별, 일별 IN_QTY 추출 → 분모용 데이터셋 생성
        """
        # ===================================================================
        # [1] 원본 데이터 확인
        # ===================================================================
        df_raw = self.data.get('DATA_1511_SMAX_wafering_300')
        if df_raw is None or df_raw.empty:
            print("[SMAX] 원본 데이터 없음: DATA_1511_SMAX_wafering_300")
            return


        # ===================================================================
        # [2] 기준일 설정: 어제 기준 최근 60일
        # ===================================================================
        # base_date = datetime.now().date() - timedelta(days=1)
        base_date = self.target_date_obj  # 이미 date 객체
        target_dates = {
            (base_date - timedelta(days=i)).strftime("%Y%m%d")
            for i in range(60)
        }
        # print("SMAX 컬럼:", df_raw.columns.tolist())
        df_filtered = df_raw[df_raw['base_dt'].astype(str).isin(target_dates)].copy()
        if df_filtered.empty:
            print("[SMAX] 필터링 후 데이터 없음 (최근 60일)")
            return

        # ===================================================================
        # [3] 필수 컬럼 확인
        # ===================================================================
        required_cols = ['base_dt', 'oper_id', 'eqp_name', 'prodc_qty']
        missing_cols = [c for c in required_cols if c not in df_filtered.columns]
        if missing_cols:
            print(f"[SMAX] 필수 컬럼 누락: {missing_cols}")
            return

        # 숫자형 변환
        df_filtered['prodc_qty'] = pd.to_numeric(df_filtered['prodc_qty'], errors='coerce').fillna(0)

        # ===================================================================
        # [4] 공정별 데이터셋 분리
        # ===================================================================
        inqty_datasets = {}
        target_procs = ['3300', '3335', '3670', '3696', '6100', '6210', '6500', '7000']


        for proc in target_procs:
            df_proc = df_filtered[df_filtered['oper_id'] == proc].copy()
            if df_proc.empty:
                print(f"[SMAX] 공정 {proc} 데이터 없음")
                continue

            # 필요한 컬럼만 유지
            df_proc = df_proc[['base_dt', 'eqp_name', 'prodc_qty']].copy()
            inqty_datasets[proc] = df_proc

        # ===================================================================
        # [5] 저장: self.data에 저장
        # ===================================================================
        self.data['SMAX_INQTY_DATASETS'] = inqty_datasets


    def _calculate_loss_rate_by_process(self):
        """
        WAF_PROCESS_DATASETS (분자) + SMAX_INQTY_DATASETS (분모)
        → 공정별, 장비별, 일별 LOSS_RATE 계산
        """
        if 'WAF_PROCESS_DATASETS' not in self.data:
            print("[LOSS_RATE] 분자 데이터 없음: WAF_PROCESS_DATASETS")
            return {}

        if 'SMAX_INQTY_DATASETS' not in self.data:
            print("[LOSS_RATE] 분모 데이터 없음: SMAX_INQTY_DATASETS")
            return {}

        waf_datasets = self.data['WAF_PROCESS_DATASETS']  # 분자
        smax_datasets = self.data['SMAX_INQTY_DATASETS']  # 분모
        loss_rate_results = {}

        for proc in waf_datasets.keys():
            if proc not in smax_datasets:
                print(f"[LOSS_RATE] 공정 {proc}: 분모 데이터 없음")
                continue

            # ===================================================================
            # [1] 분자 데이터 준비 (LOSS_QTY)
            # ===================================================================
            df_num = waf_datasets[proc].copy()
            df_num = df_num.rename(columns={
                'base_dt_7hours' : 'base_dt_7hours',
                'EQP_NM': 'eqp_name'
            })
            df_num['base_dt'] = df_num['base_dt_7hours'].astype(str)
            df_num['eqp_name'] = df_num['eqp_name'].astype(str)

            # REJ_GROUP 기준으로 LOSS_QTY 합계
            df_num_grouped = df_num.groupby(['base_dt', 'eqp_name', 'REJ_GROUP'])['LOSS_QTY'].sum().reset_index()

            # ===================================================================
            # [2] 분모 데이터 준비 (IN_QTY)
            # ===================================================================
            df_denom = smax_datasets[proc].copy()
            df_denom['base_dt'] = df_denom['base_dt'].astype(str)
            df_denom['eqp_name'] = df_denom['eqp_name'].astype(str)

            # 장비별 투입량 합계 (일별)
            df_denom_grouped = df_denom.groupby(['base_dt', 'eqp_name'])['prodc_qty'].sum().reset_index()
            df_denom_grouped = df_denom_grouped.rename(columns={'prodc_qty': 'IN_QTY'})

            # ===================================================================
            # [3] 병합: 분자 + 분모
            # ===================================================================
            df_merge = pd.merge(
                df_denom_grouped,
                df_num_grouped,
                on=['base_dt', 'eqp_name'],
                how='outer'
            )

            # ===================================================================
            # [4] LOSS_RATE 계산
            # ===================================================================
            df_merge['LOSS_QTY'] = df_merge['LOSS_QTY'].fillna(0).astype(float)
            df_merge['IN_QTY'] = df_merge['IN_QTY'].fillna(0).astype(float)


            # ===================================================================
            # [4] REJ_GROUP Pivot (행 → 컬럼)
            # ===================================================================
            pivot_cols = ['base_dt', 'eqp_name', 'IN_QTY', 'REJ_GROUP', 'LOSS_QTY']
            df_pivot = df_merge[pivot_cols].copy()
            
            # Pivot 수행
            df_pivot = df_pivot.pivot_table(
                index=['base_dt', 'eqp_name', 'IN_QTY'],
                columns='REJ_GROUP',
                values='LOSS_QTY',
                aggfunc='sum',
                fill_value=0  # 불량 없는 경우 0
            ).reset_index()
            
            # 컬럼명 정리 (MultiIndex 제거)
            df_pivot.columns.name = None

        # ===================================================================
            # [5] LOSS_RATE 계산 (전체 불량률)
            # ===================================================================
            # 전체 불량 수량 (모든 REJ_GROUP 합계)
            rej_cols = [c for c in df_pivot.columns if c not in ['base_dt', 'eqp_name', 'IN_QTY']]
            df_pivot['TOTAL_LOSS_QTY'] = df_pivot[rej_cols].sum(axis=1)
            
            # 각 불량 유형별로 LOSS_RATE 컬럼 생성
            for rej in rej_cols:
                rate_col = f"{rej}_RATE"
                df_pivot[rate_col] = (
                    df_pivot[rej] / (df_pivot['IN_QTY'] + 1e-9)
                ) * 100
                df_pivot[rate_col] = df_pivot[rate_col].round(4)

            # 날짜 정렬
            df_pivot = df_pivot.sort_values(['base_dt', 'eqp_name']).reset_index(drop=True)

            loss_rate_results[proc] = df_pivot

        # 최종 저장
        self.data['LOSS_RATE_BY_EQP'] = loss_rate_results
        return loss_rate_results


    def _plot_rej_group_top3_eqp_trend(self, output_dir="./daily_reports_debug"):
        """
        REJ_GROUP별 상위 3개 장비에 대해 IN_QTY(막대) + LOSS_RATE(선) 이중축 그래프 생성 → PNG 저장
        → 생성된 파일 경로를 반환 (엑셀 삽입용)
        """

        # ===================================================================
        # [1] LOSS_RATE_BY_EQP 존재 확인
        # ===================================================================
        if 'LOSS_RATE_BY_EQP' not in self.data:
            print("[ERROR] self.data에 'LOSS_RATE_BY_EQP' 없음")
            return {}
        else:
            print(f"[OK] LOSS_RATE_BY_EQP 존재 → {len(self.data['LOSS_RATE_BY_EQP'])}개 공정")

        # ===================================================================
        # [2] top3_rej_groups 및 valid_groups 확인
        # ===================================================================
        top3_rej_groups = self.data.get('DATA_3210_wafering_300', {}).get('top3_rej_groups', [])
        # top3_rej_groups = ['BROKEN', 'EDGE', 'CHIP']
        # 지원되는 상위 그룹 정의 (기존 조건 반영)
        SUPPORTED_TOP_GROUPS = ['PIT', 'SCRATCH', 'EDGE', 'BROKEN', 'CHIP', 'VISUAL']

        # 1. expanded_groups 생성 (BROKEN, CHIP 확장)
        expanded_groups = []
                    
        for g in top3_rej_groups:
            if g == 'BROKEN' and g in SUPPORTED_TOP_GROUPS:
                expanded_groups.extend(['LAP_BROKEN', 'EP_BROKEN', 'DSP_BROKEN', 'FP_BROKEN'])
            elif g == 'CHIP' and g in SUPPORTED_TOP_GROUPS:
                expanded_groups.extend([
                    'CHIP-LAP', 
                    'EDGE-CHIP', 
                    'CHIP_EG1AF', 
                    'CHIP_EG1BF',
                    'E_CHIP'
                ])
            elif g in SUPPORTED_TOP_GROUPS:
                expanded_groups.append(g)
            # else: 지원 안 함 → 무시

        valid_groups = [g for g in expanded_groups]

        if not valid_groups:
            print("[WARNING] 유효한 REJ_GROUP 없음 → 그래프 생성 중단")
            return {}

        # ===================================================================
        # [3] waf_analysis_gap 확인
        # ===================================================================
        waf_gap_data = self.data.get('waf_analysis_gap', {})
        if not waf_gap_data:
            print("[ERROR] self.data에 'waf_analysis_gap' 없음")
            return {}
        else:
            print(f"[OK] waf_analysis_gap 존재 → {len(waf_gap_data)}개 그룹")

        # ===================================================================
        # [4] LOSS_RATE_BY_EQP 통합
        # ===================================================================
        try:
            df_list = []
            for proc, df in self.data['LOSS_RATE_BY_EQP'].items():
                if df.empty:
                    continue
                df = df.copy()
                df['PROC_CD'] = proc
                df_list.append(df)
            
            df_all_loss = pd.concat(df_list, ignore_index=True)
            
            if df_all_loss.empty:
                print("[ERROR] LOSS_RATE_BY_EQP 통합 결과가 빈 데이터프레임")
                return {}

            # base_dt를 datetime으로 변환
            df_all_loss['base_dt_dt'] = pd.to_datetime(df_all_loss['base_dt'], format='%Y%m%d', errors='coerce')

        except Exception as e:
            print(f"[ERROR] df_all_loss 생성 실패: {e}")
            traceback.print_exc()
            return {}

        # ===================================================================
        # [5] 출력 디렉토리 생성 및 확인
        # ===================================================================
        PROJECT_ROOT = Path(__file__).parent.parent
        # base_date = (datetime.now().date() - timedelta(days=1)).strftime("%Y%m%d")
        base_date = self.target_date_obj
        debug_dir = PROJECT_ROOT / output_dir / self.target_date 
        
        if not debug_dir.exists():
            print(f"[ERROR] 디렉토리가 실제로 생성되지 않음: {debug_dir}")
            return {}
        else:
            print(f"[OK] 디렉토리 존재 확인")

        # ===================================================================
        # [6] 각 REJ_GROUP별 그래프 생성
        # ===================================================================
        graph_paths = {}  # { 'SCRATCH': [path1, path2, ...], 'BROKEN': [...], ... }

        for rej_group in valid_groups:
            gap_data = waf_gap_data.get(rej_group, {})
            if not gap_data:
                graph_paths[rej_group] = []
                continue

            # 장비별 그래프 생성
            eqp_graph_paths = []
            if isinstance(gap_data, dict):
                for eqp_col, rates in gap_data.items():
                    if not isinstance(rates, dict):
                        continue

                    top3_eqps_in_col = sorted(rates.items(), key=lambda x: abs(x[1]), reverse=True)[:3]
                    top3_eqps_in_col = [eqp for eqp, _ in top3_eqps_in_col]

                    for eqp in top3_eqps_in_col:
                        df_eqp_raw = df_all_loss[df_all_loss['eqp_name'] == eqp].copy()
                        if df_eqp_raw.empty:
                            continue

                        # LOSS_RATE 컬럼명
                        rate_col = f"{rej_group}_RATE"

                        # base_dt 기준 집계
                        agg_dict = {'IN_QTY': 'sum'}
                        if rate_col in df_eqp_raw.columns:
                            agg_dict[rate_col] = 'max'
                        df_eqp = df_eqp_raw.groupby('base_dt').agg(agg_dict).reset_index()

                        # 날짜 보정
                        latest_date = df_eqp['base_dt'].max()
                        latest_dt = pd.to_datetime(latest_date, format='%Y%m%d')
                        start_dt = latest_dt - timedelta(days=59)
                        all_dates = pd.date_range(start=start_dt, end=latest_dt, freq='D')
                        all_dates_str = all_dates.strftime('%Y%m%d')

                        df_plot = df_eqp.set_index('base_dt').reindex(all_dates_str, fill_value=0).reset_index()
                        df_plot['base_dt_dt'] = pd.to_datetime(df_plot['index'], format='%Y%m%d')
                        df_plot = df_plot.sort_values('base_dt_dt')

                        df_plot['IN_QTY'] = df_plot['IN_QTY'].fillna(0)
                        loss_rate_series = df_plot[rate_col].fillna(0.0).values if rate_col in df_plot.columns else np.zeros(len(df_plot))

                        # 그래프 생성
                        plt.figure(figsize=(12, 6))
                        ax1 = plt.gca()

                        # 막대: IN_QTY
                        ax1.bar(df_plot['base_dt_dt'], df_plot['IN_QTY'],
                                color='lightgray', alpha=0.7, label='IN_QTY', width=0.8)
                        ax1.set_xlabel('Date', fontsize=12, fontweight='bold')
                        ax1.set_ylabel('IN 수량', color='lightgray', fontsize=12, fontweight='bold')
                        ax1.tick_params(axis='y', labelcolor='lightgray')
                        ax1.set_ylim(0, max(df_plot['IN_QTY'].max() * 1.5, 4000))
                        ax1.grid(axis='y', linestyle='--', alpha=0.3)

                        # 선: LOSS_RATE
                        ax2 = ax1.twinx()
                        ax2.plot(df_plot['base_dt_dt'], loss_rate_series,
                                marker='o', linestyle='-', linewidth=2, markersize=4,
                                color='darkred', label='LOSS_RATE')
                        ax2.set_ylabel('불량률(%)', color='darkred', fontsize=12, fontweight='bold')
                        ax2.tick_params(axis='y', labelcolor='darkred')

                        max_rate = loss_rate_series.max()
                        ax2.set_ylim(0, max(max_rate * 1.5, 0.1) if max_rate > 0 else 1)

                        # 제목에 eqp_col 추가 (디버깅 용이)
                        plt.title(f'{rej_group} - {eqp} ({eqp_col[-4:]})',
                                fontsize=14, fontweight='bold', pad=20)

                        lines1, labels1 = ax1.get_legend_handles_labels()
                        lines2, labels2 = ax2.get_legend_handles_labels()
                        ax1.legend(lines1 + lines2, labels1 + labels2,
                                loc='upper left', fontsize=10, framealpha=0.9)

                        plt.xticks(rotation=45, ha='right')
                        ax1.set_xlim(all_dates[0], all_dates[-1])
                        plt.tight_layout()

                        # 파일명: 불량_장비_base_date.png
                        safe_rej = "".join(c if c.isalnum() else "_" for c in rej_group)
                        safe_eqp = "".join(c if c.isalnum() else "_" for c in eqp)
                        filename = f"loss_rate_{safe_rej}_{safe_eqp}_{base_date}.png"
                        filepath = debug_dir / filename

                        if filepath.exists():
                            filepath.unlink()
                        plt.savefig(filepath, dpi=300, bbox_inches='tight')
                        plt.close()

                        eqp_graph_paths.append(str(filepath))
                        print(f"[SUCCESS] 개별 그래프 생성: {filepath}")
            else:
                # gap_data가 flat dict인 경우 (예: SCRATCH)
                sorted_rates = sorted(gap_data.items(), key=lambda x: abs(x[1]), reverse=True)[:3]
                top3_eqps = [eqp for eqp, _ in sorted_rates]
                for eqp in top3_eqps:

                    df_eqp_raw = df_all_loss[df_all_loss['eqp_name'] == eqp].copy()
                if df_eqp_raw.empty:
                    continue

                # LOSS_RATE 컬럼명 확인
                rate_col = f"{rej_group}_RATE"

                # base_dt 기준으로 집계 (중복 제거)
                agg_dict = {'IN_QTY': 'sum'}
                if rate_col in df_eqp_raw.columns:
                    agg_dict[rate_col] = 'max'  # 또는 'mean'
                
                df_eqp = df_eqp_raw.groupby('base_dt').agg(agg_dict).reset_index()

                # 날짜 기준 reindex (모든 날짜 포함, 누락 시 0)
                df_plot = df_eqp.set_index('base_dt').reindex(all_dates_str, fill_value=0).reset_index()
                df_plot['base_dt_dt'] = pd.to_datetime(df_plot['index'], format='%Y%m%d')
                df_plot = df_plot.sort_values('base_dt_dt')

                # IN_QTY: 누락 시 0 (이미 fill_value=0 처리됨)
                df_plot['IN_QTY'] = df_plot['IN_QTY'].fillna(0)

                # LOSS_RATE: 해당 REJ_GROUP 의 _RATE 컬럼 사용
                if rate_col in df_plot.columns:
                    loss_rate_series = df_plot[rate_col].fillna(0.0).values
                else:
                    loss_rate_series = np.zeros(len(df_plot))


                # 그래프 생성
                plt.figure(figsize=(12, 6))
                ax1 = plt.gca()

                # 막대: IN_QTY
                ax1.bar(df_plot['base_dt_dt'], df_plot['IN_QTY'],
                        color='lightgray', alpha=0.7, label='IN_QTY', width=0.8)
                ax1.set_xlabel('Date', fontsize=12, fontweight='bold')
                ax1.set_ylabel('IN 수량', color='lightgray', fontsize=12, fontweight='bold')
                ax1.tick_params(axis='y', labelcolor='lightgray')
                ax1.set_ylim(0, max(df_plot['IN_QTY'].max() * 1.5, 4000))
                ax1.grid(axis='y', linestyle='--', alpha=0.3)

                # 선: LOSS_RATE
                ax2 = ax1.twinx()
                ax2.plot(df_plot['base_dt_dt'], loss_rate_series,
                        marker='o', linestyle='-', linewidth=2, markersize=4,
                        color='darkred', label='LOSS_RATE')
                ax2.set_ylabel('불량률(%)', color='darkred', fontsize=12, fontweight='bold')
                ax2.tick_params(axis='y', labelcolor='darkred')

                # Y 축 범위 조정 (0.03 → 실제 값 기반)
                max_rate = loss_rate_series.max()
                if max_rate > 0:
                    ax2.set_ylim(0, max(max_rate * 1.5, 0.1))  # 최소 0.1%
                else:
                    ax2.set_ylim(0, 1)

                # 제목
                plt.title(f'{rej_group} - {eqp} 불량률',
                        fontsize=14, fontweight='bold', pad=20)

                # 범례
                lines1, labels1 = ax1.get_legend_handles_labels()
                lines2, labels2 = ax2.get_legend_handles_labels()
                ax1.legend(lines1 + lines2, labels1 + labels2,
                        loc='upper left', fontsize=10, framealpha=0.9)

                # X축
                plt.xticks(rotation=45, ha='right')
                ax1.set_xlim(all_dates[0], all_dates[-1])
                plt.tight_layout()

                # 파일 저장
                safe_rej = "".join(c if c.isalnum() else "_" for c in rej_group)
                safe_eqp = "".join(c if c.isalnum() else "_" for c in eqp)
                filename = f"loss_rate_{safe_rej}_{safe_eqp}_{base_date}.png"
                filepath = debug_dir / filename

                if filepath.exists():
                    filepath.unlink()
                    print(f"[INFO] 기존 파일 삭제: {filepath}")
                plt.savefig(filepath, dpi=300, bbox_inches='tight')
                plt.close()

                eqp_graph_paths.append(str(filepath))
                print(f"[SUCCESS] 개별 그래프 생성: {filepath}")

            graph_paths[rej_group] = eqp_graph_paths

        # ===================================================================
        # [7] 결과 저장
        # ===================================================================
        if not hasattr(self, 'report'):
            self.report = {}
        self.data['EQP_TREND_GRAPHS'] = graph_paths
        print(f"[OK] EQP_TREND_GRAPHS 저장: {list(graph_paths.keys())}")

        return graph_paths


    def _export_to_excel(self, report, output_dir="./daily_reports_debug"):
        """Excel 보고서 생성 (기존 출력 형식과 동일하게 정확히 재현)"""
        try:

            # self.data → report 복사
            if hasattr(self, 'data'):
                eqp_graphs = self.data.get('EQP_TREND_GRAPHS')
                if eqp_graphs:
                    report['EQP_TREND_GRAPHS'] = eqp_graphs
                    print(f"[INFO] EQP_TREND_GRAPHS 복사됨: {list(eqp_graphs.keys())}")

            PROJECT_ROOT = Path(__file__).parent.parent
            # base_date = (datetime.now().date() - timedelta(days=1))
            # date_folder_name = base_date.strftime("%Y%m%d")
            # debug_dir = PROJECT_ROOT / output_dir / date_folder_name
            debug_dir = PROJECT_ROOT / "daily_reports_debug" / self.target_date
            debug_dir.mkdir(exist_ok=True, parents=True)

            excel_path = debug_dir / f"Daily_Report_{self.target_date}.xlsx"

            # 기존 파일 삭제
            if excel_path.exists():
                try:
                    excel_path.unlink()
                    print(f"기존 파일 삭제됨: {excel_path}")
                except PermissionError:
                    raise PermissionError(f"엑셀을 닫고 다시 시도하세요: {excel_path}")

            wb = Workbook()
            ws = wb.active
            ws.title = "Daily 불량분석"
            ws.sheet_view.showGridLines = False # 눈금선 끄기 추가

            # ──────────────────────────────────────────────────
            # 1. [3010 수율 분석] 제목 및 그래프 삽입 (가장 위)
            # ──────────────────────────────────────────────────
            title_cell = ws.cell(row=1, column=1, value="[ WF 수율 비교 (월/일) ]")
            title_cell.font = Font(size=12, bold=True)
            title_cell.alignment = Alignment(horizontal='left')

            data_3010_details = report.get('DATA_3010_wafering_300', {})
            chart_path_3010 = data_3010_details.get('chart_path')
            table_df_3010 = data_3010_details.get('table_df')

            # ──────────────────────────────────────────────────
            # 1. Gap 데이터 추출 (일 기준, RTY & OAY, Total & Prime)
            # ──────────────────────────────────────────────────
            gap_data = {}
            if table_df_3010 is not None and not table_df_3010.empty:
                for yld_type in ['RTY', 'OAY']:
                    for grade in ['Total', 'Prime']:
                        mask = (table_df_3010['yld_type'] == yld_type) & (table_df_3010['grade'] == grade)
                        filtered = table_df_3010[mask]
                        if not filtered.empty:
                            gap_val = filtered.iloc[0].get('daily_gap')  # 일 기준
                            if gap_val is not None:
                                try:
                                    gap_data[f'{yld_type}_{grade}'] = float(gap_val)
                                except:
                                    gap_data[f'{yld_type}_{grade}'] = None
                            else:
                                gap_data[f'{yld_type}_{grade}'] = None
                        else:
                            gap_data[f'{yld_type}_{grade}'] = None

            # ──────────────────────────────────────────────────
            # 2. 상태 및 색상 판정 함수
            # ──────────────────────────────────────────────────
            def get_status_and_color(gap):
                if gap is None:
                    return "N/A", "999999"
                if gap < 0:
                    return "미달", "FF0000"  # 빨강
                elif gap > 0:
                    return "달성", "0000FF"  # 파랑
                else:
                    return "달성", "000000"  # 검정

            # ──────────────────────────────────────────────────
            # 3. Row 2: 일 기준 RTY 요약 (한 줄)
            # ──────────────────────────────────────────────────
            # A2: -. 일 : Total RTY
            ws.cell(row=2, column=1, value="-. 일 : Total RTY")
            ws.cell(row=2, column=1).font = Font(size=10, bold=False, color="000000")
            ws.cell(row=2, column=1).alignment = Alignment(horizontal='right')

            # B2: Gap + 상태 (Total RTY)
            gap_rty_total = gap_data.get('RTY_Total')
            if gap_rty_total is not None:
                status, color = get_status_and_color(gap_rty_total)
                ws.cell(row=2, column=2, value=f"{gap_rty_total:.2f}%p {status}")
                ws.cell(row=2, column=2).font = Font(size=10, bold=True, color=color)
                ws.cell(row=2, column=2).alignment = Alignment(horizontal='center')
            else:
                ws.cell(row=2, column=2, value="N/A").font = Font(size=10, color="999999")

            # C2: Prime RTY
            ws.cell(row=2, column=3, value="Prime RTY")
            ws.cell(row=2, column=3).font = Font(size=10, bold=False, color="000000")
            ws.cell(row=2, column=3).alignment = Alignment(horizontal='left')

            # D2: Gap + 상태 (Prime RTY)
            gap_rty_prime = gap_data.get('RTY_Prime')
            if gap_rty_prime is not None:
                status, color = get_status_and_color(gap_rty_prime)
                ws.cell(row=2, column=4, value=f"{gap_rty_prime:.2f}%p {status}")
                ws.cell(row=2, column=4).font = Font(size=10, bold=True, color=color)
                ws.cell(row=2, column=4).alignment = Alignment(horizontal='center')
            else:
                ws.cell(row=2, column=4, value="N/A").font = Font(size=10, color="999999")

            # ──────────────────────────────────────────────────
            # 4. Row 3: OAY 요약 (다음 줄, 들여쓰기 없이)
            # ──────────────────────────────────────────────────
            # A3: Total OAY
            ws.cell(row=3, column=1, value="Total OAY")
            ws.cell(row=3, column=1).font = Font(size=10, bold=False, color="000000")
            ws.cell(row=3, column=1).alignment = Alignment(horizontal='right')

            # B3: Gap + 상태 (Total OAY)
            gap_oay_total = gap_data.get('OAY_Total')
            if gap_oay_total is not None:
                status, color = get_status_and_color(gap_oay_total)
                ws.cell(row=3, column=2, value=f"{gap_oay_total:.2f}%p {status}")
                ws.cell(row=3, column=2).font = Font(size=10, bold=True, color=color)
                ws.cell(row=3, column=2).alignment = Alignment(horizontal='center')
            else:
                ws.cell(row=3, column=2, value="N/A").font = Font(size=10, color="999999")

            # C3: Prime OAY
            ws.cell(row=3, column=3, value="Prime OAY")
            ws.cell(row=3, column=3).font = Font(size=10, bold=False, color="000000")
            ws.cell(row=3, column=3).alignment = Alignment(horizontal='left')

            # D3: Gap + 상태 (Prime OAY)
            gap_oay_prime = gap_data.get('OAY_Prime')
            if gap_oay_prime is not None:
                status, color = get_status_and_color(gap_oay_prime)
                ws.cell(row=3, column=4, value=f"{gap_oay_prime:.2f}%p {status}")
                ws.cell(row=3, column=4).font = Font(size=10, bold=True, color=color)
                ws.cell(row=3, column=4).alignment = Alignment(horizontal='center')
            else:
                ws.cell(row=3, column=4, value="N/A").font = Font(size=10, color="999999")

            if not chart_path_3010:
                ws.cell(row=4, column=1, value="[차트 없음: chart_path 없음]").font = Font(size=10, color="FF0000")
                print("3010: 삽입할 chart_path 없음")
            else:
                chart_path_3010 = Path(chart_path_3010)
                if not chart_path_3010.exists():
                    ws.cell(row=4, column=1, value=f"[차트 파일 없음: {chart_path_3010.name}]").font = Font(size=10, color="FF0000")
                    print(f"3010: 차트 파일 없음: {chart_path_3010}")
                else:
                    try:
                        img = ExcelImage(str(chart_path_3010))
                        img.width = 1000
                        img.height = 400
                        ws.add_image(img, 'A4')
                    except Exception as e:
                        ws.cell(row=4, column=1, value=f"[이미지 삽입 실패: {e}]").font = Font(size=10, color="FF0000")


            if table_df_3010 is not None and not table_df_3010.empty:
                start_row = 23
                start_col = 1  # A열

                cell_a = ws.cell(row=start_row, column=1, value='구분')
                ws.merge_cells(f'A{start_row}:B{start_row}')
                cell_a.font = HEADER_FONT
                cell_a.alignment = CENTER_WRAP
                cell_a.fill = HEADER_FILL
                for col in range(1, 3):  # A=1, B=2
                    ws.cell(row=start_row, column=col).border = Border(
                        left=Side(style='medium'),
                        right=Side(style='medium'),
                        top=Side(style='medium')
                        # ,  
                        # bottom=Side(style='thin')
                    )

                for col in range(1, 3):  # A=1, B=2
                    ws.cell(row=start_row, column=col).border = Border(
                        left=Side(style='medium'),
                        right=Side(style='medium'),
                        top=Side(style='medium')
                        # ,  
                        # bottom=Side(style='thin')
                    )

                # C19: '월'
                ws.merge_cells(f'C{start_row}:E{start_row}')
                cell_c = ws.cell(row=start_row, column=3)
                cell_c.font = HEADER_FONT
                cell_c.alignment = CENTER_WRAP
                cell_c.fill = HEADER_FILL
                # ✅ C19:E19 병합된 범위의 각 셀에 border 적용
                for col in range(3, 6):  # C=3, D=4, E=5
                    ws.cell(row=start_row, column=col).border = Border(
                        left=Side(style='medium'),
                        right=Side(style='medium'),
                        top=Side(style='medium'),  # ✅ top medium
                        bottom=Side(style='thin')
                    )
                cell_c.value = '월'

                # F19: '일'
                ws.merge_cells(f'F{start_row}:H{start_row}')
                cell_f = ws.cell(row=start_row, column=6)
                cell_f.font = HEADER_FONT
                cell_f.alignment = CENTER_WRAP
                cell_f.fill = HEADER_FILL
                for col in range(6, 9):  # F=6, G=7, H=8
                    ws.cell(row=start_row, column=col).border = Border(
                        left=Side(style='medium'),
                        right=Side(style='medium'),
                        top=Side(style='medium'),  # ✅ top medium
                        bottom=Side(style='thin')
                    )
                cell_f.value = '일'

                # 2. 하단 헤더
                sub_headers = ['', '', '목표', '수율', 'Gap', '목표', '수율', 'Gap']

                for c_idx, h in enumerate(sub_headers, start_col):
                    cell = ws.cell(row=start_row+1, column=c_idx)
                    cell.value = h
                    cell.font = HEADER_FONT
                    cell.alignment = CENTER_WRAP
                    cell.fill = HEADER_FILL

                    # 병합된 셀(A20:B20)의 경우, B20은 내부 셀 → 테두리 조정
                    if c_idx == 1:  # A20
                        cell.border = Border(
                            left=Side(style='medium'),
                            right=Side(style='thin'),    # 내부선
                            top=Side(style='thin'),
                            bottom=Side(style='medium')
                        )
                    elif c_idx == 2:  # B20 (내부 셀)
                        cell.border = Border(
                            left=Side(style='thin'),     # 내부선
                            right=Side(style='medium'),  # 병합 외곽 오른쪽
                            top=Side(style='thin'),
                            bottom=Side(style='medium')
                        )
                    else:  # C20 \~ H20
                        # D20, G20: left/right = thin ✅
                        left_style = 'thin' if c_idx in [4, 7] else 'medium'
                        right_style = 'thin' if c_idx in [4, 7] else 'medium'

                        # C20, F20 (그룹 왼쪽 끝)
                        if c_idx in [3, 6]:
                            cell.border = Border(
                                left=Side(style='medium'),
                                right=Side(style='thin'),
                                top=Side(style='thin'),
                                bottom=Side(style='medium')
                            )
                        # D20, E20, G20, H20
                        elif c_idx in [4, 5, 7, 8]:
                            cell.border = Border(
                                left=Side(style='thin'),
                                right=Side(style='medium') if c_idx in [5, 8] else Side(style='thin'),
                                top=Side(style='thin'),
                                bottom=Side(style='medium')
                            )

                    # ==========================================
                    # 3. 데이터 행
                    # ==========================================

                    row_idx = 0
                    last_row_num = None  # 마지막 행 저장용
                    for yld_type in ['RTY', 'OAY']:
                        # ✅ [수정] yld_type 마다 2 행씩 증가 (Total, Prime 각 1 행)
                        main_cat_row = start_row + 2 + (row_idx * 2)
                        
                        # A 열 병합 (세로)
                        ws.merge_cells(f'A{main_cat_row}:A{main_cat_row+1}')
                        cell_main = ws.cell(row=main_cat_row, column=1)
                        cell_main.value = f'WF {yld_type}'
                        cell_main.font = Font(bold=True, size=9)
                        cell_main.alignment = CENTER_WRAP
                        cell_main.border = Border(left=Side(style='medium'), right=Side(style='medium'), top=Side(style='thin'), bottom=Side(style='thin'))
                        
                        for grade_idx, grade in enumerate(['Total', 'Prime']):
                            row_num = main_cat_row + grade_idx
                            data_row = table_df_3010[
                                (table_df_3010['yld_type'] == yld_type) &
                                (table_df_3010['grade'] == grade)
                            ].iloc[0]
                            grade_label = "Total(P+N)" if grade == 'Total' else "Prime"
                            # B 열: 소분류
                            cell_b = ws.cell(row=row_num, column=2, value=grade_label)
                            cell_b.alignment = CENTER_WRAP
                            cell_b.border = Border(left=Side(style='medium'), right=Side(style='medium'), top=Side(style='thin'), bottom=Side(style='thin'))
                            
                            # C\~H 열: 월/일 데이터
                            # 월
                            cell_c = ws.cell(row=row_num, column=3, value=data_row['monthly_plan'] / 100.0)
                            cell_c.number_format = '0.00%'
                            cell_c.font = Font(bold=False, size=9)
                            cell_c.alignment = CENTER_WRAP
                            cell_c.border = Border(left=Side(style='medium'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
                            
                            cell_d = ws.cell(row=row_num, column=4, value=data_row['monthly_actual'] / 100.0)
                            cell_d.number_format = '0.00%'
                            cell_d.font = Font(bold=False, size=9)
                            cell_d.alignment = CENTER_WRAP
                            cell_d.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
                            
                            cell_e = ws.cell(row=row_num, column=5, value=data_row['monthly_gap'] / 100.0)
                            cell_e.number_format = '+0.00%;-0.00%;0.00%'
                            cell_e.alignment = CENTER_WRAP
                            cell_e.border = Border(left=Side(style='thin'), right=Side(style='medium'), top=Side(style='thin'), bottom=Side(style='thin'))
                            if data_row['monthly_gap'] > 0:
                                cell_e.font = Font(color="0000FF", bold=True, size=9)
                            elif data_row['monthly_gap'] < 0:
                                cell_e.font = Font(color="FF0000", bold=True, size=9)
                            # 일
                            cell_f = ws.cell(row=row_num, column=6, value=data_row['daily_plan'] / 100.0)
                            cell_f.number_format = '0.00%'
                            cell_f.font = Font(bold=False, size=9)
                            cell_f.alignment = CENTER_WRAP
                            cell_f.border = Border(left=Side(style='medium'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
                            
                            cell_g = ws.cell(row=row_num, column=7, value=data_row['daily_actual'] / 100.0)
                            cell_g.number_format = '0.00%'
                            cell_g.font = Font(bold=False, size=9)
                            cell_g.alignment = CENTER_WRAP
                            cell_g.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
                            
                            cell_h = ws.cell(row=row_num, column=8, value=data_row['daily_gap'] / 100.0)
                            cell_h.number_format = '+0.00%;-0.00%;0.00%'
                            cell_h.alignment = CENTER_WRAP
                            cell_h.border = Border(left=Side(style='thin'), right=Side(style='medium'), top=Side(style='thin'), bottom=Side(style='thin'))
                            if data_row['daily_gap'] > 0:
                                cell_h.font = Font(color="0000FF", bold=True, size=9)
                            elif data_row['daily_gap'] < 0:
                                cell_h.font = Font(color="FF0000", bold=True, size=9)

                        last_row_num = main_cat_row + 1  # 마지막 행 번호 저장 (Prime 행)
                        row_idx += 1

                    # ✅ 표의 맨 밑줄 (마지막 데이터 행) bottom 테두리 medium 적용
                    if last_row_num is not None:
                        for col in range(1, 9):  # A\~H 열
                            cell = ws.cell(row=last_row_num, column=col)
                            # 기존 border를 유지하면서 bottom만 medium으로
                            current_border = cell.border
                            cell.border = Border(
                                left=current_border.left,
                                right=current_border.right,
                                top=current_border.top,
                                bottom=Side(style='medium')  # ✅ bottom만 medium
                            )

            else:
                ws.cell(row=2, column=8, value="표 없음").font = Font(size=10, color="FF0000")

            # ──────────────────────────────────────────────────
            # 4. [RC/HG 보상 영향성 분석] 섹션
            # ──────────────────────────────────────────────────
            title_cell = ws.cell(row=30, column=1, value="[ RC/HG 보상 영향성 분석(Ref. 3개월 比) ]")
            title_cell.font = Font(size=12, bold=True)
            title_cell.alignment = Alignment(horizontal='left')

            current_date = (datetime.now().date() - timedelta(days=1)).strftime("%Y%m%d")
            debug_dir = PROJECT_ROOT / "daily_reports_debug" / current_date

            total_chart_path = debug_dir / "RC_HG_보상_전체.png"
            group_chart_paths = {
                'PARTICLE': debug_dir / "RC_HG_보상_PARTICLE.png",
                'FLATNESS': debug_dir / "RC_HG_보상_FLATNESS.png",
                'WARP&BOW': debug_dir / "RC_HG_보상_WARP&BOW.png",
                'NANO': debug_dir / "RC_HG_보상_NANO.png"
            }

            data_lot_details = report.get('DATA_LOT_3210_wafering_300_details', {})
            loss_rate_table_total = data_lot_details.get('summary')  # 전체 표
            loss_rate_table_by_group = data_lot_details.get('loss_rate_table_by_group', {})  # 그룹별 표

            current_row = 32
            SECTION_HEIGHT = 9

            def safe_pct_to_float(x):
                try:
                    if pd.isna(x) or x == '' or x is None:
                        return 0.0
                    # %만 제거, 부호는 유지
                    cleaned = str(x).strip().replace('%', '')
                    if cleaned == '':
                        return 0.0
                    # float 변환 시 자동으로 +, - 처리
                    return float(cleaned) / 100.0
                except Exception as e:
                    print(f"[ERROR] safe_pct_to_float 변환 실패: {x} → {e}")
                    return 0.0

            # 값 변환 함수 (% 문자열 제거 → float)
            def parse_pct_value(val):
                if val is None or (isinstance(val, float) and pd.isna(val)):
                    return None
                # 문자열이면 % 제거
                if isinstance(val, str):
                    val = val.replace('%', '').strip()
                try:
                    return float(val)
                except:
                    return None

            def create_group_comment(ws, row_addr, label, table_row):
                """그룹별 코멘트 생성 (RichText 대신 셀 분리 방식)"""
                try:
                    raw_daily = table_row.iloc[0]['일%']
                    raw_ref = table_row.iloc[0]['Ref.(3개월)%']
                    raw_gap = table_row.iloc[0]['Gap']

                    daily_rate = parse_pct_value(raw_daily)
                    ref_rate = parse_pct_value(raw_ref)
                    gap_val = parse_pct_value(raw_gap)

                    if any(v is None for v in [daily_rate, ref_rate, gap_val]):
                        raise ValueError("값 파싱 실패")

                    prefix = f" - {label} 보상률 {daily_rate:+.2f}%, Ref {ref_rate:+.2f}%, 比 "

                    # gap_val 기준으로 텍스트와 상태 결정
                    if gap_val > 0:
                        status = "양호"
                        gap_color = "0000FF"  # 빨간색 : FF0000, 파란색 : 0000FF
                    elif gap_val < 0:
                        status = "열위"
                        gap_color = "FF0000"  # 파란색
                    else:
                        status = "변화없음"
                        gap_color = "000000"  # 검은색      

                    gap_text = f"{gap_val:+.2f}%p {status}"

                    # A 열: 접두사
                    ws[f'A{row_addr}'] = prefix
                    ws[f'A{row_addr}'].font = Font(size=9, color="000000")

                    # B 열: Gap 값 (색상 적용)
                    ws[f'C{row_addr}'] = gap_text
                    gap_color = "0000FF" if gap_val > 0 else "FF0000" if gap_val < 0 else "000000"
                    ws[f'C{row_addr}'].font = Font(size=9, color=gap_color, bold=True)
                    return True
                except Exception as e:
                    print(f"[ERROR] {label} 코멘트 생성 실패: {e}")
                    ws[f'A{row_addr}'] = f" - [{label}: {str(e)[:20]}]"
                    ws[f'A{row_addr}'].font = Font(size=9, color="FF0000")
                    return False

            # 전체 그래프 + 표
            if total_chart_path.exists():
                # 코멘트 삽입: 그래프 바로 위
                comment_row = current_row - 1  # 그래프는 current_row 에 삽입 → 그 위에 코멘트
                ws[f'A{comment_row}'] = "Total"
                ws[f'A{comment_row}'].font = Font(size=10, bold=True)

                if isinstance(loss_rate_table_total, pd.DataFrame) and not loss_rate_table_total.empty:
                    fs_row = loss_rate_table_total[loss_rate_table_total['구분'] == 'FS']
                    resc_row = loss_rate_table_total[loss_rate_table_total['구분'] == 'RESC']
                    hg_row = loss_rate_table_total[loss_rate_table_total['구분'] == 'HG']

                    # 코멘트 시작 행
                    comment_row += 1

                    # FS 처리
                    if not resc_row.empty:
                        create_group_comment(ws, comment_row, 'FS', fs_row)
                        comment_row += 1
                    else:
                        ws[f'A{comment_row}'] = " - [FS 데이터 없음]"
                        ws[f'A{comment_row}'].font = Font(size=9, color="808080")
                        comment_row += 1
                    # RESC 처리
                    if not resc_row.empty:
                        create_group_comment(ws, comment_row, 'RC', resc_row)
                        comment_row += 1
                    else:
                        ws[f'A{comment_row}'] = " - [RESC 데이터 없음]"
                        ws[f'A{comment_row}'].font = Font(size=9, color="808080")
                        comment_row += 1
                    # HG 처리
                    if not hg_row.empty:
                        create_group_comment(ws, comment_row, 'HG', hg_row)
                        comment_row += 1
                    else:
                        ws[f'A{comment_row}'] = " - [HG 데이터 없음]"
                        ws[f'A{comment_row}'].font = Font(size=9, color="808080")
                        comment_row += 1

                graph_row = comment_row + 1
                if total_chart_path.exists():
                    try:
                        img = ExcelImage(str(total_chart_path))
                        img.width = 600
                        img.height = 300
                        ws.add_image(img, f'A{current_row+3}')
                    except Exception as e:
                        ws[f'A{current_row+3}'] = f"[RC/HG 전체 그래프 삽입 실패: {e}]"
                        ws[f'A{current_row+3}'].font = Font(size=10, color="FF0000")

                else:
                    ws[f'A{graph_row}'] = "[전체 그래프 파일 없음]"
                    ws[f'A{graph_row}'].font = Font(size=10, color="FF0000")

                if isinstance(loss_rate_table_total, pd.DataFrame) and not loss_rate_table_total.empty:
                    headers = ['구분', 'Ref.(3개월)', '일', 'Ref.(3개월)%', '일%', 'Gap']
                    start_row = graph_row
                    start_col = 7

                    for c_idx, header in enumerate(headers, start_col):
                        cell = ws.cell(row=start_row, column=c_idx, value=header)
                        cell.font = Font(bold=True, size=10)
                        cell.fill = PatternFill("solid", fgColor="D3D3D3")
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                        cell.border = Border(left=Side(style='thin'), right=Side(style='thin'),
                                            top=Side(style='thin'), bottom=Side(style='thin'))

                    table_total_fmt = loss_rate_table_total.copy()
                    pct_columns = ['Ref.(3개월)%', '일%', 'Gap']
                    for col in pct_columns:
                        if col in table_total_fmt.columns:
                            table_total_fmt[col] = table_total_fmt[col].apply(safe_pct_to_float)

                    for r_idx, row in enumerate(dataframe_to_rows(table_total_fmt, index=False, header=False), start_row + 1):
                        for c_idx, value in enumerate(row, start_col):
                            cell = ws.cell(row=r_idx, column=c_idx, value=value)
                            cell.font = Font(size=9)
                            cell.alignment = Alignment(horizontal='center', vertical='center')
                            cell.border = Border(left=Side(style='thin'), right=Side(style='thin'),
                                                top=Side(style='thin'), bottom=Side(style='thin'))
                            
                            # Ref.(3 개월), 일은 #,### 형식
                            if c_idx in [8, 9]:
                                cell.number_format = '#,###'
                            if c_idx in [10, 11, 12]:
                                cell.number_format = '0.00%'
                            if c_idx == 12:
                                try:
                                    gap_val = float(value)
                                    if gap_val < 0:
                                        cell.fill = PatternFill("solid", fgColor="FFCCCC") 
                                        cell.font = Font(color="FF0000", bold=False, size=9) # 빨간색 : FF0000,FFCCCC 파란색 : 0000FF, CCE5FF
                                    elif gap_val > 0:
                                        cell.fill = PatternFill("solid", fgColor="CCE5FF")
                                        cell.font = Font(color="0000FF", bold=False, size=9)
                                except:
                                    pass

                    table_height = len(loss_rate_table_total) + 1                        

                else:
                    ws.cell(row=graph_row, column=7, value="[RC/HG 전체 표 없음]").font = Font(size=10, color="FF0000")
                    table_height = 1

                current_row = graph_row + 13

            # ──────────────────────────────────────────────────
            # 4-2. [RC/HG 보상 영향성 분석 - Prime] 섹션 추가
            # ──────────────────────────────────────────────────
            prime_chart_path = debug_dir / "RC_HG_보상_전체_Prime.png"
            prime_data_details = report.get('DATA_LOT_3210_wafering_300_details', {})
            prime_loss_rate_table_total = prime_data_details.get('prime_summary')  # Prime 전체 표
            prime_loss_rate_table_by_group = prime_data_details.get('prime_loss_rate_table_by_group', {})  # 그룹별 표

            # # Prime 섹션 시작
            # ws.cell(row=current_row, column=1, value="[ RC/HG 보상 영향성 분석(Ref. 3개월 比) - Prime ]")
            # ws.cell(row=current_row, column=1).font = Font(size=12, bold=True)
            # ws.cell(row=current_row, column=1).alignment = Alignment(horizontal='left')
            # current_row += 2  # 여백

            if prime_chart_path.exists():
                # 코멘트 삽입: 그래프 바로 위
                comment_row = current_row
                ws[f'A{comment_row}'] = "Prime"
                ws[f'A{comment_row}'].font = Font(size=10, bold=True)
                comment_row += 1

                if isinstance(prime_loss_rate_table_total, pd.DataFrame) and not prime_loss_rate_table_total.empty:
                    fs_row = prime_loss_rate_table_total[prime_loss_rate_table_total['구분'] == 'FS']
                    resc_row = prime_loss_rate_table_total[prime_loss_rate_table_total['구분'] == 'RESC']
                    hg_row = prime_loss_rate_table_total[prime_loss_rate_table_total['구분'] == 'HG']

                    if not fs_row.empty:
                        create_group_comment(ws, comment_row, 'FS', fs_row)
                        comment_row += 1
                    else:
                        ws[f'A{comment_row}'] = " - [FS 데이터 없음]"
                        ws[f'A{comment_row}'].font = Font(size=9, color="808080")
                        comment_row += 1

                    if not resc_row.empty:
                        create_group_comment(ws, comment_row, 'RC', resc_row)
                        comment_row += 1
                    else:
                        ws[f'A{comment_row}'] = " - [RESC 데이터 없음]"
                        ws[f'A{comment_row}'].font = Font(size=9, color="808080")
                        comment_row += 1

                    if not hg_row.empty:
                        create_group_comment(ws, comment_row, 'HG', hg_row)
                        comment_row += 1
                    else:
                        ws[f'A{comment_row}'] = " - [HG 데이터 없음]"
                        ws[f'A{comment_row}'].font = Font(size=9, color="808080")
                        comment_row += 1
                else:
                    ws[f'A{comment_row}'] = " - [Prime 분석 데이터 없음]"
                    ws[f'A{comment_row}'].font = Font(size=9, color="808080")
                    comment_row += 1

                graph_row = comment_row + 1
                try:
                    img = ExcelImage(str(prime_chart_path))
                    img.width = 600
                    img.height = 300
                    ws.add_image(img, f'A{graph_row}')
                except Exception as e:
                    ws[f'A{graph_row}'] = f"[RC/HG Prime 그래프 삽입 실패: {e}]"
                    ws[f'A{graph_row}'].font = Font(size=10, color="FF0000")

                # Prime 표 삽입 (H열)
                if isinstance(prime_loss_rate_table_total, pd.DataFrame) and not prime_loss_rate_table_total.empty:
                    headers = ['구분', 'Ref.(3개월)', '일', 'Ref.(3개월)%', '일%', 'Gap']
                    start_row = graph_row
                    start_col = 7

                    for c_idx, header in enumerate(headers, start_col):
                        cell = ws.cell(row=start_row, column=c_idx, value=header)
                        cell.font = Font(bold=True, size=10)
                        cell.fill = PatternFill("solid", fgColor="D3D3D3")
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                        cell.border = Border(left=Side(style='thin'), right=Side(style='thin'),
                                            top=Side(style='thin'), bottom=Side(style='thin'))

                    table_fmt = prime_loss_rate_table_total.copy()
                    pct_columns = ['Ref.(3개월)%', '일%', 'Gap']
                    for col in pct_columns:
                        if col in table_fmt.columns:
                            table_fmt[col] = table_fmt[col].apply(safe_pct_to_float)

                    for r_idx, row in enumerate(dataframe_to_rows(table_fmt, index=False, header=False), start_row + 1):
                        for c_idx, value in enumerate(row, start_col):
                            cell = ws.cell(row=r_idx, column=c_idx, value=value)
                            cell.font = Font(size=9)
                            cell.alignment = Alignment(horizontal='center', vertical='center')
                            cell.border = Border(left=Side(style='thin'), right=Side(style='thin'),
                                                top=Side(style='thin'), bottom=Side(style='thin'))
                            
                            if c_idx in [8, 9]:
                                cell.number_format = '#,###'
                            if c_idx in [10, 11, 12]:
                                cell.number_format = '0.00%'
                            if c_idx == 12:
                                try:
                                    gap_val = float(value)
                                    if gap_val < 0:
                                        cell.fill = PatternFill("solid", fgColor="FFCCCC") 
                                        cell.font = Font(color="FF0000", bold=False, size=9)
                                    elif gap_val > 0:
                                        cell.fill = PatternFill("solid", fgColor="CCE5FF")
                                        cell.font = Font(color="0000FF", bold=False, size=9)
                                except:
                                    pass
                else:
                    ws.cell(row=graph_row, column=7, value="[RC/HG Prime 표 없음]").font = Font(size=10, color="FF0000")

                current_row = graph_row + 13
            else:
                ws.cell(row=current_row, column=1, value="[RC/HG Prime 그래프 없음: RC_HG_보상_전체_Prime.png]")
                ws.cell(row=current_row, column=1).font = Font(size=10, color="FF0000")
                current_row += 15


            # ──────────────────────────────────────────────────
            # 5. [ 제품 영향성 분석 ] 섹션
            # ──────────────────────────────────────────────────
            current_row = current_row + 1
            ws[f'A{current_row}'] = "[ 제품 영향성 분석(Ref. 6개월 比) ]"
            ws[f'A{current_row}'].font = Font(size=12, bold=True)
            current_row += 1

            total_loss_gap = report.get('total_loss_gap')

            if not isinstance(total_loss_gap, pd.DataFrame) or total_loss_gap.empty or 'PRODUCT_TYPE' not in total_loss_gap.columns:
                ws.cell(row=current_row, column=1, value="[제품 영향성 분석: 데이터 없음]").font = Font(size=10, color="FF0000")
                current_row += 10
            else:
                display_df = total_loss_gap.rename(columns={'PRODUCT_TYPE': '제품',
                                                            '물량비_GAP(%)' : 'Ref. 比 물량비 Gap',
                                                            'Ref_전체_불량률(%)' : 'Ref_전체_불량률',
                                                            'Daily_물량비(%)' : '일 물량비',
                                                            'Ref_물량비(%)' : 'Ref_물량비'}).copy()

                pct_columns = ['Ref_전체_불량률', '일 물량비', 'Ref_물량비', 'Ref. 比 물량비 Gap', '제품 Mix비 변동']

                for col in pct_columns:
                    if col in display_df.columns:
                        display_df[col] = display_df[col] / 100

                # ──────────────────────────────────────────────────
                # Total 코멘트 추가 (제목 바로 아래)
                # ──────────────────────────────────────────────────
                overall_stats = report.get('overall_stats', {})
                total_val = overall_stats.get('total_volume_defect_change', 0.0)
                total_val = total_val

                # Total 코멘트 생성
                total_comment = f"-. Total 제품 Mix 비 기인 변동 {total_val:+.2f}%p"

                ws[f'A{current_row}'] = total_comment
                ws[f'A{current_row}'].font = Font(size=10, color="000000", bold=False)  
                ws[f'A{current_row}'].alignment = Alignment(horizontal='left')
                current_row += 1

                # ──────────────────────────────────────────────────
                # 제품별 코멘트 생성 (상위 3개 제품)
                # ──────────────────────────────────────────────────
                comment_row = current_row  # 코멘트는 그래프 전에 출력
                for _, row in display_df.iterrows():
                    pt = row['제품']
                    gap_loss = row['제품 Mix비 변동']       # 0.01% → 0.01%p
                    ref_ratio = row['Ref_물량비']           # 5.74%
                    daily_ratio = row['일 물량비']          # 12.96%
                    vol_gap = row['Ref. 比 물량비 Gap']     # 7.22%p

                    # 물량비 불량 변동이 양수면 "열위", 음수면 "양호"
                    impact = "열위" if gap_loss > 0 else "양호"

                    # 코멘트 생성
                    comment = (
                        f"-. {pt} 제품 Mix비 기인 {gap_loss*100:+.2f}%p {impact}"
                        f"(Ref. {ref_ratio*100:.2f}% -> {daily_ratio*100:.2f}%, "
                        f"{vol_gap*100:+.2f}%p 물량 변동)"
                    )

                    ws[f'A{comment_row}'] = comment
                    ws[f'A{comment_row}'].font = Font(size=10, color="000000", bold=False)
                    ws[f'A{comment_row}'].alignment = Alignment(horizontal='left')
                    comment_row += 1

                current_row = comment_row + 1  # 코멘트 후 여백

                # 전체 평균 행 추가
                total_ref_qty = overall_stats.get('total_ref_qty', 0)
                total_daily_qty = overall_stats.get('total_daily_qty', 0)
                overall_ref_loss_rate = overall_stats.get('overall_ref_loss_rate', 0.0)
                overall_daily_loss_rate = overall_stats.get('overall_daily_loss_rate', 0.0)
                total_volume_defect_change = overall_stats.get('total_volume_defect_change', 0.0) #제품 Mix비 변동 전체 sum

                # 첫 번째 그래프: 제품 Mix비 변동
                chart1_path = debug_dir / "제품 Mix비 변동.png"
                try:
                    # Total 데이터 준비
                    total_val = overall_stats.get('total_volume_defect_change', 0.0)
                    total_val = total_val / 100 #%로 변경
                    # display_df에서 제품명과 제품 Mix비 변동 추출
                    x_labels = ['Total'] + [str(pt) for pt in display_df['제품']]
                    y_values = [float(total_val)] + display_df['제품 Mix비 변동'].astype(float).tolist()

                    # 색상 설정: 값이 양수면 빨간색, 음수면 파란색
                    colors = ['#ff0000' if val >= 0 else '#0000ff' for val in y_values]

                    def wrap_label(label, max_width=15):
                        """긴 라벨을 지정된 너비만큼 줄바꿈"""
                        return '\n'.join(textwrap.wrap(label, width=max_width))

                    # X 축 라벨 줄바꿈 적용
                    wrapped_labels = [wrap_label(str(label), max_width=10) for label in x_labels]

                    # 그래프 생성
                    fig1, ax1 = plt.subplots(figsize=(14, 6))  # 6개 제품 기준
                    bars = ax1.bar(x_labels, y_values, color=colors)

                    # Y축 범위 설정 (기존 방식 유지)
                    y_min_val = min(y_values + [0])
                    y_max_val = max(y_values + [0])
                    y_range = y_max_val - y_min_val
                    
                    # 범위가 0 인 경우 방지
                    if y_range == 0:
                        y_range = 0.001  # 최소 범위 설정
                    # Y축 상/하단 20% 여백 추가
                    y_min = y_min_val - y_range * 0.2
                    y_max = y_max_val + y_range * 0.2
                    ax1.set_ylim(y_min, y_max)
                    ax1.yaxis.set_major_formatter(PercentFormatter(1.0))  # 1.0 = 100%

                    # 제목 및 레이블
                    ax1.set_title("제품별 물량비 불량 변동 (Ref 기준)", fontsize=20, fontweight='bold')
                    ax1.set_xlabel('제품', fontsize=14, fontweight='bold')
                    ax1.set_ylabel('물량비 불량 변동', fontsize=14, fontweight='bold')
                    ax1.set_xticklabels(wrapped_labels, rotation=0, ha='center', fontsize=18)
                    ax1.grid(axis='y', linestyle='--', alpha=0.7, zorder=0)
                    # ax1.tick_params(axis='x', labelsize=18, rotation=0, pad=8)      # X축 눈금 글씨 크기
                    ax1.tick_params(axis='y', labelsize=14, pad=8)                  # Y축 눈금 글씨 크기

                    plt.setp(ax1.get_xticklabels(), ha='right') 

                    label_offset =  y_range * 0.05  # 동적 label_offset 설정 (Y축 범위의 5% 수준)

                    for bar, val in zip(bars, y_values):
                        height = bar.get_height()
                        if val >= 0:
                            pos_y = float(height + label_offset)
                            va = 'bottom'
                        else:
                            pos_y = float(height - label_offset)
                            va = 'top'
                        label_text = f'{val * 100:+.2f}%'
                        ax1.text(bar.get_x() + bar.get_width() / 2, pos_y,
                                label_text, ha='center', va=va, fontsize=16, fontweight='bold', color='black')

                    plt.tight_layout()
                    plt.savefig(chart1_path, dpi=300, bbox_inches='tight')  # DPI 향상 (선명도 개선)
                    plt.close()

                    # Excel 삽입
                    if chart1_path.exists():
                        img1 = ExcelImage(str(chart1_path))
                        img1.width = 800
                        img1.height = 300
                        ws.add_image(img1, f'A{current_row}')

                except Exception as e:
                    ws[f'A{current_row}'] = f"[그래프1 생성 실패: {e}]"
                    ws[f'A{current_row}'].font = Font(size=10, color="FF0000")

                # 두 번째 그래프: 물량비_GAP(%)
                chart2_path = debug_dir / "제품_물량비_GAP.png"
                try:
                    # 데이터 준비
                    products = display_df['제품'].tolist()
                    ref_values = display_df['Ref_물량비'].astype(float).tolist()      # Ref 물량비
                    daily_values = display_df['일 물량비'].astype(float).tolist()     # 일 물량비
                    gap_values = [daily - ref for ref, daily in zip(ref_values, daily_values)]
                    # ──────────────────────────────────────────────────
                    # 전체 Y 축 범위 계산 (모든 제품 통합)
                    # ──────────────────────────────────────────────────
                    all_values = ref_values + daily_values  # 모든 값 통합
                    global_min = min(all_values)
                    global_max = max(all_values)

                    # 여유 포함 (0 반드시 포함)
                    y_min = min(0, global_min * 0.95)
                    y_max = max(global_max * 1.15, 0.01)  # 최소 1% 확보
                    y_range = y_max - y_min

                    n_products = len(products)
                    if n_products == 0:
                        raise ValueError("표시할 제품이 없습니다.")

                    # 3 개 서브플롯 생성 (가로로 나열)
                    fig2, axes = plt.subplots(1, n_products, figsize=(4.5 * n_products, 6))
                    fig2.suptitle('물량변동그래프', fontsize=20, fontweight='bold', y=0.98)

                    # 단일 제품일 경우 리스트로 변환
                    if n_products == 1:
                        axes = [axes]  # 리스트로 감싸서 iterable하게 만듦
                    # 이미 배열이므로 flatten 불필요

                    label_offset = y_range * 0.025  # 막대 값 라벨: 전체 높이의 2.5%
                    gap_offset   = y_range * 0.06   # Gap 라벨: 전체 높이의 6% (더 멀리)

                    for i, ax in enumerate(axes):
                        # 데이터 추출
                        ref_val = ref_values[i]
                        daily_val = daily_values[i]
                        gap_val = gap_values[i]
                                        
                        min_height = min(ref_val, daily_val)
                        max_height = max(ref_val, daily_val)

                        # 막대 그래프 생성 (x 위치: 0=Ref, 1=일)
                        bars = ax.bar([0, 1], [ref_val, daily_val], color=['#0000ff', '#ff0000'])
                        ax.set_ylim(y_min, y_max) #  Y 축 범위 통일 (모든 서브플롯 동일)
                        ax.yaxis.set_major_formatter(PercentFormatter(1.0))
                        
                        # 제목 (제품명, 박스 없이 글만)
                        ax.set_title(f"{products[i]}", fontsize=20, fontweight='bold', pad=10)
                        
                        # X 축 레이블
                        ax.set_xticks([0, 1])
                        ax.set_xticklabels(['Ref.', '일'], fontsize=14)
                        ax.tick_params(axis='x', labelsize=14, rotation=0, pad=8)      # X축 눈금 글씨 크기
                        ax.tick_params(axis='y', labelsize=14, pad=8)                  # Y축 눈금 글씨 크기

                        # 그리드
                        ax.grid(axis='y', linestyle='--', alpha=0.7, zorder=0)
                        
                        # 막대 위 라벨 
                        for bar, val in zip(bars, [ref_val, daily_val]):
                            height = bar.get_height()
                            if val >= 0:
                                pos_y = float(height + label_offset)
                                va = 'bottom'
                            else:
                                pos_y = float(height - label_offset)
                                va = 'top'
                            ax.text(
                                bar.get_x() + bar.get_width() / 2, 
                                pos_y,
                                f'{val*100:.2f}%', 
                                ha='center', 
                                va=va,
                                fontsize=16, 
                                fontweight='bold', 
                                color='black',
                                zorder=4
                            )
                        
                        # Gap 라벨: 항상 max_height 기준 (음수여도 높은 막대 위에 표시)
                        gap_y = max_height + gap_offset
                        va_align = 'bottom'  # 항상 막대 위에 표시

                        # Gap 라벨 (막대 사이 상단)
                        gap_x = 0.5  # 두 막대 중간 위치
                        gap_color = '#0000ff' if gap_val >= 0 else '#ff0000'  # +:파랑, -:빨강
                        gap_text = f'{gap_val*100:+.2f}%'
                        ax.text(gap_x, gap_y, gap_text, ha='center', va='bottom',
                            fontsize=16, fontweight='bold', color=gap_color)
                    
                    plt.tight_layout()
                    plt.savefig(chart2_path, dpi=300, bbox_inches='tight')  # DPI 향상 (선명도 개선)
                    plt.close()

                    # Excel 삽입
                    if chart2_path.exists():
                        img2 = ExcelImage(str(chart2_path))
                        img2.width = 800  
                        img2.height = 300
                        ws.add_image(img2, f'A{current_row + 15}')

                except Exception as e:
                    ws[f'A{current_row + 15}'] = f"[그래프 2 생성 실패: {e}]"
                    ws[f'A{current_row + 15}'].font = Font(size=10, color="FF0000")

                current_row += 30

                # 표 삽입 
                table_start_row = current_row
                headers = ['제품', '제품 Mix비 변동', 'Ref_전체_불량률', 'Ref. 比 물량비 Gap', '일 물량비', 'Ref_물량비', '일 수량', 'Ref(6개월) 수량'] 
                # 헤더 삽입
                for c_idx, header in enumerate(headers, 1):
                    cell = ws.cell(row=table_start_row, column=c_idx, value=header)
                    cell.font = Font(bold=True, size=10)
                    cell.fill = PatternFill("solid", fgColor="D3D3D3")
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    cell.border = Border(left=Side(style='thin'), right=Side(style='thin'),
                                        top=Side(style='thin'), bottom=Side(style='thin'))
                # 데이터 행 삽입 (상위 3개 제품)
                for r_idx, row in display_df.iterrows():
                    for c_idx, col in enumerate(headers, 1):
                        value = ""
                        if col == '제품':
                            value = row['제품']
                        elif col == 'Ref_전체_불량률':
                            value = row['Ref_전체_불량률']
                        elif col == 'Ref. 比 물량비 Gap':
                            value = row['Ref. 比 물량비 Gap']
                        elif col == '일 물량비':
                            value = row['일 물량비']
                        elif col == 'Ref_물량비':
                            value = row['Ref_물량비']
                        elif col == '제품 Mix비 변동':
                            value = row['제품 Mix비 변동']
                        elif col == '일 수량':
                            value = row['Daily_Compile_수량']  # 정수
                        elif col == 'Ref(6개월) 수량':
                            value = row['Ref_Compile_수량']    # 정수

                        cell = ws.cell(row=table_start_row + 1 + r_idx, column=c_idx, value=value)
                        cell.font = Font(size=9)
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                        cell.border = Border(left=Side(style='thin'), right=Side(style='thin'),
                                            top=Side(style='thin'), bottom=Side(style='thin'))

                        # number_format 적용
                        if col in ['Ref_전체_불량률', 'Ref. 比 물량비 Gap', '제품 Mix비 변동']:
                            cell.number_format = '+0.00%;-0.00%;0.00%'
                        elif col in ['일 물량비', 'Ref_물량비']:
                            cell.number_format = '0.00%'
                        elif col in ['일 수량', 'Ref(6개월) 수량']:
                            cell.number_format = '#,##0'

                        # GAP 관련 채색
                        if col in ['Ref. 比 물량비 Gap', '제품 Mix비 변동']:
                            try:
                                gap_val = float(value)
                                if gap_val > 0:
                                    cell.fill = PatternFill("solid", fgColor="FFCCCC")
                                    cell.font = Font(color="FF0000", bold=False, size=9)
                                elif gap_val < 0:
                                    cell.fill = PatternFill("solid", fgColor="CCE5FF")
                                    cell.font = Font(color="0000FF", bold=False, size=9)
                            except:
                                pass

                avg_row = table_start_row + len(total_loss_gap) + 1

                # A 열: '전체 평균'
                cell = ws.cell(row=avg_row, column=1, value='전체 평균')
                cell.font = Font(bold=False, size=9)
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = Border(left=Side(style='thin'), right=Side(style='thin'),
                                    top=Side(style='thin'), bottom=Side(style='thin'))
                
                # B 열: 제품 Mix비 변동
                cell = ws.cell(row=avg_row, column=2, value=total_volume_defect_change / 100)
                cell.font = Font(bold=False, size=9)
                cell.number_format = '+0.00%;-0.00%;0.00%'
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = Border(left=Side(style='thin'), right=Side(style='thin'),
                                    top=Side(style='thin'), bottom=Side(style='thin'))

                # D 열: Ref_전체_불량률 (%)
                cell = ws.cell(row=avg_row, column=3, value=overall_ref_loss_rate / 100.0)
                cell.font = Font(size=9)
                cell.number_format = '+0.00%;-0.00%;0.00%'
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = Border(left=Side(style='thin'), right=Side(style='thin'),
                                    top=Side(style='thin'), bottom=Side(style='thin'))

                for c_idx in [4,5,6]:  
                    cell = ws.cell(row=avg_row, column=c_idx, value="")
                    cell.font = Font(size=9)  
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    cell.border = Border(left=Side(style='thin'), right=Side(style='thin'),
                                        top=Side(style='thin'), bottom=Side(style='thin'))

                # I 열: 일 수량
                cell = ws.cell(row=avg_row, column=7, value=total_daily_qty)
                cell.font = Font(size=9)
                cell.number_format = '#,##0'
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = Border(left=Side(style='thin'), right=Side(style='thin'),
                                    top=Side(style='thin'), bottom=Side(style='thin'))

                # J 열: Ref(6개월) 수량
                cell = ws.cell(row=avg_row, column=8, value=total_ref_qty)
                cell.font = Font(size=9)
                cell.number_format = '#,##0'
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = Border(left=Side(style='thin'), right=Side(style='thin'),
                                    top=Side(style='thin'), bottom=Side(style='thin'))

                # --- 표 아래에 계산식 설명 추가 ---
                explanation_row = avg_row + 1  # 전체 평균 다음 행

                cell = ws.cell(row=explanation_row, column=1, 
                            value='※ 제품 Mix비 변동 = (Ref. 제품별 불량률 - Ref. 평균 불량률) × 물량변동비')
                cell.font = Font(size=9, italic=True, color="555555")  # 작고, 이탤릭, 회색으로 시각적 차별
                cell.alignment = Alignment(horizontal='left', vertical='center')
                ws.row_dimensions[avg_row].height = 18

                current_row = avg_row + 1

            # ──────────────────────────────────────────────────
            # 2. [Prime 불량 목표 比 일실적 변동]
            # ──────────────────────────────────────────────────
            next_start_row = 118
            data_3210_details = report.get('DATA_3210_wafering_300_details', {})
            chart_path = data_3210_details.get('chart_path')

            ws.merge_cells(f'A{next_start_row}:D{next_start_row}')
            ws[f'A{next_start_row}'] = "[ Prime 불량 목표 比 일실적 변동 ]"
            ws[f'A{next_start_row}'].font = Font(size=14, bold=True)
            ws[f'A{next_start_row}'].alignment = Alignment(horizontal='left')

            # ──────────────────────────────────────────────────
            # GAP 기준 상위 3개 불량에 대한 코멘트 생성 
            # ──────────────────────────────────────────────────
            comment_row = next_start_row + 1  # 제목 바로 아래
            if 'summary' in data_3210_details:
                summary_df = data_3210_details['summary'].copy()

                # === Total 행 처리 (맨 앞에 추가) ===
                total_row = summary_df[summary_df['REJ_GROUP'] == 'Total']
                total_comment = ""
                if not total_row.empty:
                    total_loss_pct = total_row['LOSS_RATIO_PCT'].iloc[0]
                    total_gap_pct = total_row['GAP_PCT'].iloc[0]
                    total_status = "미달" if total_gap_pct > 0 else "달성"
                    total_comment = f"Total {total_loss_pct:.2f}%({total_gap_pct:+.2f}%p {total_status})"

                # === Top 3 불량 (Total 제외하고 추출) ===
                non_total_df = summary_df[summary_df['REJ_GROUP'] != 'Total']
                top3 = non_total_df.nlargest(3, 'GAP_PCT')
                
                comment_parts = []
                for _, row in top3.iterrows():
                    rej = row['REJ_GROUP']
                    loss_pct = row['LOSS_RATIO_PCT']   # % 단위
                    gap_pct = row['GAP_PCT']          # %p 단위
                    status = "미달" if gap_pct > 0 else "달성"
                    comment_parts.append(f"{rej} {loss_pct:.2f}%({gap_pct:+.2f}%p {status})")
                
                if comment_parts:
                    all_parts = [total_comment] + comment_parts if total_comment else comment_parts
                    comment_text = "-. Prime 주요 불량 " + ", ".join(all_parts)
                    ws[f'A{comment_row}'] = comment_text
                    ws[f'A{comment_row}'].font = Font(size=10, color="000000", bold=False)
                    ws[f'A{comment_row}'].alignment = Alignment(horizontal='left')
                    comment_row += 1
                else:
                    comment_row += 1
            else:
                comment_row += 1

            # ──────────────────────────────────────────────────
            # 그래프 삽입 (기존 위치 유지)
            # ──────────────────────────────────────────────────
            if not chart_path:
                ws[f'A{comment_row}'] = "[차트 없음]"
                ws[f'A{comment_row}'].font = Font(size=10, color="FF0000")
            else:
                chart_path = Path(chart_path)
                if not chart_path.exists():
                    ws[f'A{comment_row}'] = f"[파일 없음: {chart_path.name}]"
                    ws[f'A{comment_row}'].font = Font(size=10, color="FF0000")
                else:
                    try:
                        img = ExcelImage(str(chart_path))
                        img.width = 600
                        img.height = 300
                        ws.add_image(img, f'A{comment_row}')
                    except Exception as e:
                        ws[f'A{comment_row}'] = f"[삽입 실패: {e}]"
                        ws[f'A{comment_row}'].font = Font(size=10, color="FF0000")

            # 요약 표 삽입 (G열)
            table_df_for_row_height = None
            summary_for_comment = None  
            if 'summary' in data_3210_details:
                table_df = data_3210_details['summary'][['REJ_GROUP', 'GOAL_RATIO_PCT', 'LOSS_RATIO_PCT', 'GAP_PCT']].copy()
                table_df.columns = ['구분', '목표(%)', '실적(%)', 'GAP(%)']
                for col in ['목표(%)', '실적(%)', 'GAP(%)']:
                    table_df[col] = table_df[col] / 100.0

                summary_for_comment = table_df
                table_df_for_row_height = table_df
                start_row = next_start_row + 1
                start_col = 8

                for r_idx, row in enumerate(dataframe_to_rows(table_df, index=False, header=True), start_row):
                    for c_idx, value in enumerate(row, start_col):
                        cell = ws.cell(row=r_idx, column=c_idx, value=value)
                        cell.border = Border(left=Side(style='thin'), right=Side(style='thin'),
                                            top=Side(style='thin'), bottom=Side(style='thin'))
                        cell.font = Font(size=9)
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                        if r_idx == start_row:
                            cell.font = Font(bold=True, size=9)
                            cell.fill = PatternFill("solid", fgColor="D3D3D3")
                        else:
                            cell.number_format = '0.00%'
                            if c_idx == start_col + 3:  # GAP 열
                                try:
                                    gap_val = float(value)
                                    if gap_val > 0:
                                        cell.fill = PatternFill("solid", fgColor="FFCCCC")
                                        cell.font = Font(color="FF0000", bold=False, size=9)
                                    elif gap_val < 0:
                                        cell.fill = PatternFill("solid", fgColor="CCE5FF")
                                        cell.font = Font(color="0000FF", bold=False, size=9)
                                except:
                                    pass

                for row in range(start_row, start_row + len(table_df) + 1):
                    ws.row_dimensions[row].height = 18


            # ──────────────────────────────────────────────────
            # 3. [Prime 주요 열위 불량 세부코드 분석]
            # ──────────────────────────────────────────────────
            row_start = next_start_row + 20
            ws.merge_cells(f'A{row_start-1}:F{row_start-1}')
            ws[f'A{row_start-1}'] = "[ Prime 주요 열위 불량 세부코드 분석 Ref.(3개월) 比 일실적 변동 (상위 3개) ]"
            ws[f'A{row_start-1}'].font = Font(size=12, bold=True)
            ws[f'A{row_start-1}'].alignment = Alignment(horizontal='left')

            mid_analysis = report.get('DATA_3210_wafering_300_3months', {}).get('top3_midgroup_analysis', {})
            plot_paths = mid_analysis.get('plot_paths', {})
            group_tables = mid_analysis.get('tables', {})
            detailed_analysis_dict = self.data.get('DATA_3210_wafering_300_3months', {}).get('top3_midgroup_analysis', {}).get('detailed_analysis', {})  # 딕셔너리

            # ──────────────────────────────────────────────────
            # 각 REJ_GROUP 별 코멘트 생성 함수
            # ──────────────────────────────────────────────────
            def create_midgroup_comment(rej_group, table_df, summary_df):
                """
                세부 코드별 코멘트 생성 (대분류 Total Gap 포함)
                """
                if table_df is None or table_df.empty:
                    return None

                # 1. 대분류 Total Gap 가져오기
                total_gap = None
                if summary_df is not None and '구분' in summary_df.columns:
                    match_row = summary_df[summary_df['구분'] == rej_group]
                    if not match_row.empty:
                        total_gap = match_row.iloc[0]['GAP(%)']  # 소수 형태

                # 2. Total Gap 코멘트 생성
                total_comment = ""
                if total_gap is not None:
                    status = "미달" if total_gap > 0 else "초과" if total_gap < 0 else "동등"
                    total_comment = f"{rej_group} 불량 Total {abs(total_gap * 100):.2f}%p {status}, "

                # 3. 상위 3개 MID_GROUP 코멘트 생성
                top3 = table_df.nlargest(3, 'Gap')
                comment_parts = []
                for _, row in top3.iterrows():
                    mid = row['MID_GROUP']
                    value = row['실적(%)']
                    gap = row['Gap']
                    status = "열위" if gap > 0 else "양호" if gap < 0 else "동등"
                    comment_parts.append(f"{mid} {value:.2f}%(Ref. 比 {gap:+.2f}%p {status})")

                # 4. 최종 코멘트 조합
                if comment_parts:
                    return f"-. {total_comment}" + ", ".join(comment_parts)
                else:
                    return f"-. {rej_group} 불량 Total {abs(total_gap * 100):.2f}%p {status}" if total_gap is not None else None

            # ──────────────────────────────────────────────────
            # 상세 분석 대상 그룹
            # ──────────────────────────────────────────────────
            target_groups = ['PARTICLE', 'FLATNESS', 'WARP&BOW', 'NANO']
            current_row = row_start

            # ──────────────────────────────────────────────────
            # 각 REJ_GROUP 별로 그래프 + 코멘트 + 표 출력
            # ──────────────────────────────────────────────────
            for idx, rej_group in enumerate(plot_paths.keys()):

                # 1. 코멘트 생성 및 출력 (그래프 위)
                comment_text = create_midgroup_comment(rej_group, group_tables.get(rej_group), summary_for_comment)
                if comment_text:
                    ws[f'A{current_row}'] = comment_text
                    ws[f'A{current_row}'].font = Font(size=10, color="000000", bold=False)
                    ws[f'A{current_row}'].alignment = Alignment(horizontal='left')
                    graph_row = current_row + 1
                else:
                    graph_row = current_row

                # 2. 그래프 삽입
                plot_path = plot_paths.get(rej_group)
                if plot_path and Path(plot_path).exists():
                    try:
                        img = ExcelImage(plot_path)
                        img.width = 600
                        img.height = 300
                        ws.add_image(img, f'A{graph_row}')
                    except Exception as e:
                        ws.cell(row=graph_row, column=1, value=f"{rej_group} 이미지 오류").font = Font(size=9)
                else:
                    ws.cell(row=graph_row, column=1, value=f"{rej_group}: 그래프 없음").font = Font(size=9)

                # 3. 표 삽입

                table_start_row = graph_row + 1 
                table_df = group_tables.get(rej_group)
                table_end_row = table_start_row + 1

                if table_df is not None and not table_df.empty:
                    headers = ['MID_GROUP', '실적(%)', 'Ref(3개월)', 'Gap']
                    start_col = 8

                    for c_idx, header in enumerate(headers, start_col):  #F열부터
                        cell = ws.cell(row=table_start_row, column=c_idx, value=header)
                        cell.font = Font(bold=True, size=10)
                        cell.fill = PatternFill("solid", fgColor="D3D3D3")
                        cell.border = Border(left=Side(style='thin'), right=Side(style='thin'),
                                            top=Side(style='thin'), bottom=Side(style='thin'))
                        cell.alignment = Alignment(horizontal='center')

                    table_df_fmt = table_df.copy()
                    for col in ['실적(%)', 'Ref(3개월)', 'Gap']:
                        if col in table_df_fmt.columns:
                            table_df_fmt[col] = pd.to_numeric(table_df_fmt[col], errors='coerce') / 100.0

                    for r_idx, row in enumerate(dataframe_to_rows(table_df_fmt, index=False, header=False), table_start_row + 1):
                        for c_idx, value in enumerate(row, start_col):
                            cell = ws.cell(row=r_idx, column=c_idx, value=value)
                            cell.font = Font(size=9)
                            cell.border = Border(left=Side(style='thin'), right=Side(style='thin'),
                                                top=Side(style='thin'), bottom=Side(style='thin'))
                            cell.alignment = Alignment(horizontal='center')
                            if c_idx in [7,8]:
                                cell.number_format = '#,##0'
                            if c_idx in [9, 10, 11]:  # 퍼센트 컬럼
                                cell.number_format = '0.00%'
                            if c_idx == 11:  # Gap 컬럼
                                try:
                                    gap_val = float(value)
                                    if gap_val > 0:
                                        cell.fill = PatternFill("solid", fgColor="FFCCCC")
                                        cell.font = Font(color="FF0000", size=9)
                                    elif gap_val < 0:
                                        cell.fill = PatternFill("solid", fgColor="CCE5FF")
                                        cell.font = Font(color="0000FF", size=9)
                                except:
                                    pass
                    table_end_row = table_start_row + len(table_df) + 1
                else:
                    ws.cell(row=graph_row, column=8, value="표 없음").font = Font(size=9)
                # 4. 상세분석 텍스트 (M 열) - 표와 동일한 table_start_row 사용
                analysis_lines = detailed_analysis_dict.get(rej_group, [])
                analysis_col = 13  # M 열

                # [1] 제목 셀 (단일 행)
                title_cell = ws.cell(row=table_start_row, column=analysis_col, value=f"[{rej_group} 불량]")
                title_cell.font = Font(size=9, bold=True, color="000000")
                title_cell.alignment = Alignment(horizontal='left', vertical='center')

                merge_start = table_start_row + 1
                merge_end = max(merge_start + 4, table_end_row)  # 최소 5행, 표보다 짧으면 늘림
                ws.merge_cells(f'M{merge_start}:M{merge_end}')

                full_content = "\n".join(analysis_lines) if analysis_lines else "분석 없음"
                content_cell = ws.cell(row=merge_start, column=analysis_col, value=full_content)
                content_cell.font = Font(size=9)
                content_cell.alignment = Alignment(wrap_text=True, horizontal='left', vertical='top')
                ws.column_dimensions['M'].width = 50 # [5] L 열 너비 고정
                # 5. 좌측 블록 끝 행 계산
                left_block_end = max(graph_row + 15, table_end_row)
                # ──────────────────────────────────────────────────
                #  조건부 FS/HG/RESC 상세 분석 (기존 변수 사용)
                # ──────────────────────────────────────────────────
                next_row = left_block_end + 1  # 좌측 블록 아래 시작

                if rej_group.upper() in target_groups:
                    chart_path = group_chart_paths[rej_group]  
                    table_data = loss_rate_table_by_group.get(rej_group)  

                    # 그룹 제목
                    ws[f'A{next_row}'] = f"{rej_group} (FS/HG/RESC 상세)"
                    ws[f'A{next_row}'].font = Font(size=10, bold=True)
                    comment_row = next_row + 1

                    # 코멘트 (RESC, HG)
                    if isinstance(table_data, pd.DataFrame) and not table_data.empty:
                        resc_row = table_data[table_data['구분'] == 'RESC']
                        hg_row = table_data[table_data['구분'] == 'HG']

                        if not resc_row.empty:
                            create_group_comment(ws, comment_row, 'RESC', resc_row)
                            comment_row += 1
                        else:
                            ws[f'A{comment_row}'] = " - [RESC 데이터 없음]"
                            ws[f'A{comment_row}'].font = Font(size=9, color="808080")
                            comment_row += 1

                        if not hg_row.empty:
                            create_group_comment(ws, comment_row, 'HG', hg_row)
                            comment_row += 1
                        else:
                            ws[f'A{comment_row}'] = " - [HG 데이터 없음]"
                            ws[f'A{comment_row}'].font = Font(size=9, color="808080")
                            comment_row += 1
                    else:
                        comment_row += 1

                    # 그래프 삽입
                    detail_graph_row = comment_row + 1
                    if chart_path.exists():
                        try:
                            img = ExcelImage(str(chart_path))
                            img.width = 600
                            img.height = 300
                            ws.add_image(img, f'A{detail_graph_row}')
                        except Exception as e:
                            ws[f'A{detail_graph_row}'] = f"[{rej_group} 그래프 삽입 실패]"
                            ws[f'A{detail_graph_row}'].font = Font(size=9, color="FF0000")
                    else:
                        ws[f'A{detail_graph_row}'] = f"[{rej_group} 그래프 없음]"
                        ws[f'A{detail_graph_row}'].font = Font(size=9, color="FF0000")

                    # 표 삽입 (H열부터)
                    detail_table_start_row = detail_graph_row
                    detail_table_end_row = detail_table_start_row + 1
                    if isinstance(table_data, pd.DataFrame) and not table_data.empty:
                        headers = ['구분', 'Ref.(3개월)', '일', 'Ref.(3개월)%', '일%', 'Gap']
                        start_col = 7
                        for c_idx, header in enumerate(headers, start_col):
                            cell = ws.cell(row=detail_table_start_row, column=c_idx, value=header)
                            cell.font = Font(bold=True, size=10)
                            cell.fill = PatternFill("solid", fgColor="D3D3D3")
                            cell.alignment = Alignment(horizontal='center', vertical='center')
                            cell.border = Border(left=Side(style='thin'), right=Side(style='thin'),
                                                top=Side(style='thin'), bottom=Side(style='thin'))

                        table_group_fmt = table_data.copy()
                        for col in ['Ref.(3개월)%', '일%', 'Gap']:
                            if col in table_group_fmt.columns:
                                table_group_fmt[col] = table_group_fmt[col].apply(safe_pct_to_float)

                        for r_idx, row in enumerate(dataframe_to_rows(table_group_fmt, index=False, header=False), detail_table_start_row + 1):
                            for c_idx, value in enumerate(row, start_col):
                                cell = ws.cell(row=r_idx, column=c_idx, value=value)
                                cell.font = Font(size=9)
                                cell.alignment = Alignment(horizontal='center', vertical='center')
                                cell.border = Border(left=Side(style='thin'), right=Side(style='thin'),
                                                    top=Side(style='thin'), bottom=Side(style='thin'))
                                if c_idx in [8, 9]:
                                    cell.number_format = '#,##0'
                                if c_idx in [10, 11, 12]:
                                    cell.number_format = '0.00%'
                                if c_idx == 12:
                                    try:
                                        gap_val = float(value)
                                        if gap_val > 0:
                                            cell.fill = PatternFill("solid", fgColor="CCE5FF")
                                            cell.font = Font(color="0000FF", bold=False, size=9)
                                        elif gap_val < 0:
                                            cell.fill = PatternFill("solid", fgColor="FFCCCC")
                                            cell.font = Font(color="FF0000", bold=False, size=9)
                                    except:
                                        pass

                        detail_table_end_row = detail_table_start_row + len(table_data) + 1
                    else:
                        ws.cell(row=detail_graph_row, column=8, value=f"[{rej_group} 표 없음]").font = Font(size=9, color="FF0000")

                    # 다음 그룹 시작 위치: 상세 분석 블록 끝 + 1
                    detail_block_end = max(detail_graph_row + 15, detail_table_end_row)
                    current_row = detail_block_end
                else:
                    # 상세 분석 없으면 기존 위치 유지
                    current_row = left_block_end

                # current_row += 1

            # ──────────────────────────────────────────────────
            # 6. [장비별 불량률 GAP 분석]
            # ──────────────────────────────────────────────────
            current_row = current_row + 2
            ws[f'A{current_row}'] = "[장비별 불량률 GAP 분석]"
            ws[f'A{current_row}'].font = Font(size=12, bold=True)
            current_row += 2
            waf_analysis_details = report.get('DATA_WAF_3210_wafering_300_details', {})

            # waf_analysis_ref가 dict가 아니면 빈 dict
            waf_ref_data = waf_analysis_details.get('waf_analysis_ref')
            if not isinstance(waf_ref_data, dict):
                print(f"[ERROR] waf_analysis_ref type 오류: {type(waf_ref_data)} → 빈 dict로 대체")
                waf_ref_data = {}

            waf_daily_data = waf_analysis_details.get('waf_analysis_daily')
            if not isinstance(waf_daily_data, dict):
                print(f"[ERROR] waf_analysis_daily type 오류: {type(waf_daily_data)} → 빈 dict로 대체")
                waf_daily_data = {}

            waf_gap_data = waf_analysis_details.get('waf_analysis_gap', {})
            if not isinstance(waf_gap_data, dict):
                waf_gap_data = {}

            # 열 번호 → 문자 변환 함수 (get_column_letter 대체)
            def col_num_to_letter(n):
                if n <= 0:
                    return 'A'
                result = ''
                while n > 0:
                    n -= 1
                    result = chr(65 + (n % 26)) + result
                    n //= 26
                return result

            top3_rej_groups = report.get('DATA_3210_wafering_300_details', {}).get('top3_rej_groups', [])
            # top3_rej_groups = ['BROKEN', 'EDGE', 'CHIP']  # 원본 그룹 리스트
            expanded_groups = []
            
            for g in top3_rej_groups:
                if g == 'BROKEN':
                    # BROKEN → 4개 공정별 하위 그룹으로 확장
                    expanded_groups.extend(['LAP_BROKEN', 'EP_BROKEN', 'DSP_BROKEN', 'FP_BROKEN'])
                elif g == 'CHIP':
                    # CHIP → 세부 유형으로 확장 (시각화에 사용되는 실제 키)
                    expanded_groups.extend([
                        'CHIP-LAP', 
                        'EDGE-CHIP', 
                        'CHIP_EG1AF', 
                        'CHIP_EG1BF',
                        'EDGE-ECHIP' 
                    ])
                else:
                    # 그 외: SCRATCH, EDGE 등은 그대로 유지
                    expanded_groups.append(g)

            # waf_gap_data에 누락된 CHIP 그룹 추가 (빈 dict)
            for group in expanded_groups:
                if group not in waf_gap_data:
                    if group in waf_ref_data and group in waf_daily_data:
                        waf_gap_data[group] = {}

            # valid_groups: waf_gap_data에 실제로 존재하는 그룹만 필터링
            valid_groups = [g for g in expanded_groups if g in waf_gap_data]

            if not valid_groups:
                ws[f'A{current_row}'] = "[WAF 분석: 상위 3개 그룹 중 대상 없음]"
                ws[f'A{current_row}'].font = Font(size=10, color="FF0000")
                current_row += 10
            else:

                # 대분류 → 하위 그룹 매핑
                MAIN_GROUPS = {
                    'BROKEN': ['LAP_BROKEN', 'EP_BROKEN', 'DSP_BROKEN', 'FP_BROKEN'],
                    'CHIP': ['CHIP-LAP', 'EDGE-CHIP', 'CHIP_EG1AF', 'CHIP_EG1BF', 'EDGE-ECHIP'],
                    'EDGE': ['EDGE'],
                    'SCRATCH': ['SCRATCH'],
                    'PIT': ['PIT'],
                    'VISUAL': ['VISUAL']
                }

                ALLOWED_GROUPS = ['PIT', 'SCRATCH', 'EDGE', 'BROKEN', 'CHIP', 'VISUAL']
                valid_main_groups = [g for g in top3_rej_groups if g in ALLOWED_GROUPS]

                # 대분류 중 유효한 것만 필터링
                processed_main_groups = []
                for main_group in valid_main_groups:
                    if main_group not in MAIN_GROUPS:
                        continue
                    sub_groups = MAIN_GROUPS[main_group]
                    if any(sg in valid_groups for sg in sub_groups):
                        processed_main_groups.append(main_group)

                for main_group in processed_main_groups:
                    # ───────────────────────────────────────────────
                    # 1단계: 해당 대분류의 모든 하위 그룹 gap_data 통합
                    # ───────────────────────────────────────────────
                    all_gap_data = {}
                    all_ref_data = {}
                    all_daily_data = {}

                    if main_group in ['BROKEN', 'CHIP']:
                        sub_groups = MAIN_GROUPS[main_group]
                        for sg in sub_groups:
                            if sg in valid_groups:
                                gap_data_sg = waf_gap_data.get(sg, {})
                                if isinstance(gap_data_sg, dict):
                                    all_gap_data.update(gap_data_sg)
                                all_ref_data.update(waf_ref_data.get(sg, {}))
                                all_daily_data.update(waf_daily_data.get(sg, {}))
                    else:
                        # EDGE 등 단일 그룹
                        all_gap_data = waf_gap_data.get(main_group, {})
                        all_ref_data = waf_ref_data.get(main_group, {})
                        all_daily_data = waf_daily_data.get(main_group, {})

                    if not all_gap_data:
                        print(f"[WARNING] {main_group}: 통합 gap_data 없음")
                        ws[f'A{current_row}'] = f"[{main_group}: 분석 데이터 없음]"
                        ws[f'A{current_row}'].font = Font(size=10, color="FF0000")
                        current_row += 10
                        continue

                    # ───────────────────────────────────────────────
                    # 2단계: Y축 통일을 위한 GAP 값 수집
                    # ───────────────────────────────────────────────
                    all_gap_values = []
                    if isinstance(all_gap_data, dict) and isinstance(next(iter(all_gap_data.values()), {}), dict):
                        for eqp_col, rates in all_gap_data.items():
                            if not rates: continue
                            for val in rates.values():
                                all_gap_values.append(val)

                    if all_gap_values:
                        global_min = min(all_gap_values)
                        global_max = max(all_gap_values)
                        margin = max(0.0001, abs(global_max - global_min) * 0.2)
                        y_min = global_min - margin
                        y_max = global_max + margin
                        if y_min > 0: y_min = -margin
                        if y_max < 0: y_max = margin
                    else:
                        y_min, y_max = -0.0005, 0.0005

                    # ───────────────────────────────────────────────
                    # 3단계: 그래프 생성 (소분류 기준)
                    # ───────────────────────────────────────────────
                    graph_start_row = current_row
                    current_graph_row = graph_start_row
                    graphs_in_row = 0

                    # 소분류별로 그래프 생성
                    for sub_group in sub_groups:
                        if sub_group not in valid_groups:
                            continue
                        
                        gap_data_sg = waf_gap_data.get(sub_group, {})
                        ref_data_sg = waf_ref_data.get(sub_group, {})      
                        daily_data_sg = waf_daily_data.get(sub_group, {})  

                        if isinstance(next(iter(gap_data_sg.values()), {}), dict):
                            eqp_list = []
                            for eqp_col, rates in gap_data_sg.items():
                                proc_match = re.search(r'(\d{4})', eqp_col)
                                proc = int(proc_match.group(1)) if proc_match else 9999
                                if rates:
                                    eqp_list.append({'eqp_col': eqp_col, 'proc': proc, 'rates': rates})

                            sorted_eqp_list = sorted(eqp_list, key=lambda x: x['proc'])
                            top3_eqp_list = sorted_eqp_list[:3]

                            for item in top3_eqp_list:
                                eqp_col = item['eqp_col']
                                proc = str(item['proc'])
                                rates = item['rates']

                                if not rates: continue

                                # 그래프 제목에 소분류 사용
                                safe_rej = "".join(c if c.isalnum() else "_" for c in sub_group)
                                safe_eqp = "".join(c if c.isalnum() else "_" for c in proc)
                                chart_path = debug_dir / f"WAF_{safe_rej}_{safe_eqp}_gap_chart.png"
                                
                                if chart_path.exists():
                                    chart_path.unlink()

                                try:
                                    fig, ax = plt.subplots(figsize=(10, 6))
                                    sorted_rates = sorted(rates.items(), key=lambda x: abs(x[1]), reverse=True)[:3]
                                    equipment_names = [eqp for eqp, _ in sorted_rates]

                                    # Ref와 Daily 값 가져오기 (기존 ref_data_sg, daily_data_sg 사용)
                                    ref_values = []
                                    daily_values = []
                                    gap_values = []
                                    for eqp_name, gap_val in sorted_rates:
                                        # Ref 데이터
                                        level1 = ref_data_sg.get(eqp_col, {})
                                        ref_eqp_data = level1.get(eqp_name, {}) if isinstance(level1, dict) else ref_data_sg.get(eqp_name, {})
                                        ref_rate_val = ref_eqp_data.get('rate', 0.0)

                                        # Daily 데이터
                                        level1_daily = daily_data_sg.get(eqp_col, {})
                                        daily_eqp_data = level1_daily.get(eqp_name, {}) if isinstance(level1_daily, dict) else daily_data_sg.get(eqp_name, {})
                                        daily_rate_val = daily_eqp_data.get('rate', 0.0)

                                        ref_values.append(float(ref_rate_val))
                                        daily_values.append(float(daily_rate_val))
                                        gap_values.append(float(gap_val))

                                    # ✅ 장비 수만큼 서브플롯 생성 (가로로 나열)
                                    n_eqps = len(equipment_names)
                                    fig, axes = plt.subplots(1, n_eqps, figsize=(5 * n_eqps, 4))
                                    
                                    # 단일 장비일 경우 axes 를 리스트로 변환
                                    if n_eqps == 1:
                                        axes = [axes]
                                    
                                    for i, ax in enumerate(axes):
                                        # 데이터 추출
                                        ref_val = ref_values[i]
                                        daily_val = daily_values[i]
                                        gap_val = gap_values[i]
                                        eqp_name = equipment_names[i]
                                        
                                        # ✅ 막대 그래프 생성 (x 위치: 0=Ref, 1=일)
                                        bars = ax.bar([0, 1], [ref_val, daily_val], color=['#0000ff', '#ff0000'])
                                        
                                        # ✅ Y 축 범위 설정 (여유 포함, 0 포함)
                                        all_vals = [ref_val, daily_val, 0]
                                        y_min = min(all_vals) * 0.95 if min(all_vals) <= 0 else 0
                                        y_max = max(all_vals) * 1.15 if max(all_vals) >= 0 else 0
                                        if y_min == y_max:
                                            y_min, y_max = -0.1, 0.1
                                        ax.set_ylim(y_min, y_max)
                                        ax.yaxis.set_major_formatter(PercentFormatter(1.0))
                                        
                                        # ✅ 제목 (장비명, 박스 없이 글만)
                                        ax.set_title(f"{eqp_name}", fontsize=14, fontweight='bold', pad=10)
                                        
                                        # ✅ X 축 레이블
                                        ax.set_xticks([0, 1])
                                        ax.set_xticklabels(['Ref.', '일'], fontsize=14)
                                        
                                        # ✅ 그리드
                                        ax.grid(axis='y', linestyle='--', alpha=0.7, zorder=0)
                                        
                                        # ✅ 막대 위 라벨
                                        for bar, val in zip(bars, [ref_val, daily_val]):
                                            height = bar.get_height()
                                            ax.text(bar.get_x() + bar.get_width() / 2, height,
                                                f'{val*100:.2f}%', ha='center', va='bottom',
                                                fontsize=14, fontweight='bold', color='black')
                                        
                                        # ✅ Gap 라벨 (막대 사이 상단)
                                        gap_x = 0.5  # 두 막대 중간 위치
                                        gap_y = max(ref_val, daily_val) * 1.05  # 높은 막대 위에 표시
                                        gap_color = '#ff0000' if gap_val >= 0 else '#0000ff'  # +:빨강, -:파랑 (표와 동일)
                                        gap_text = f'{gap_val*100:+.2f}%'
                                        ax.text(gap_x, gap_y, gap_text, ha='center', va='bottom',
                                            fontsize=14, fontweight='bold', color=gap_color)
                                    
                                    # ✅ 전체 제목 (소분류 - 공정)
                                    fig.suptitle(f'{sub_group} - 공정 {proc}', fontsize=14, fontweight='bold', y=1.02)
                                    
                                    plt.tight_layout()
                                    plt.savefig(chart_path, dpi=300, bbox_inches='tight')
                                    plt.close()

                                    # ✅ Excel 삽입 (기존 로직 유지)
                                    if chart_path.exists():
                                        img = ExcelImage(str(chart_path))
                                        img.width = 400
                                        img.height = 200

                                        if graphs_in_row >= 3:
                                            current_graph_row += 10
                                            graphs_in_row = 0

                                        col_offset = graphs_in_row * 5
                                        col_letter = col_num_to_letter(1 + col_offset)
                                        ws.add_image(img, f'{col_letter}{current_graph_row}')
                                        graphs_in_row += 1

                                except Exception as e:
                                    print(f"[ERROR] 그래프 생성 실패: {sub_group}-{proc} | {e}")

                    # 그래프 다음 행 계산
                    next_row_after_graph = current_graph_row + (10 if graphs_in_row > 0 else 0)

                    # ───────────────────────────────────────────────
                    # 4단계: 통합 표 생성
                    # ───────────────────────────────────────────────
                    table_start_row = next_row_after_graph + 1

                    headers = ['불량','구분', '장비', 'Ref.(3개월)', '일', 'Ref.(3개월)', '일', 'Gap']
                    for c_idx, header in enumerate(headers, 1):
                        cell = ws.cell(row=table_start_row, column=c_idx, value=header)
                        cell.font = Font(bold=True, size=9)
                        cell.fill = PatternFill("solid", fgColor="D3D3D3")
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                        cell.border = Border(left=Side(style='thin'), right=Side(style='thin'),
                                            top=Side(style='thin'), bottom=Side(style='thin'))

                    # 데이터 수집 (소분류 기준)
                    table_rows = []
                    for sub_group in sub_groups:
                        if sub_group not in valid_groups:
                            continue
                        
                        gap_data_sg = waf_gap_data.get(sub_group, {})
                        ref_data_sg = waf_ref_data.get(sub_group, {})
                        daily_data_sg = waf_daily_data.get(sub_group, {})
                        
                        if not isinstance(gap_data_sg, dict):
                            continue

                        if isinstance(next(iter(gap_data_sg.values()), {}), dict):
                            for eqp_col, rates in gap_data_sg.items():
                                proc = eqp_col[-4:] if eqp_col[-4:].isdigit() else eqp_col
                                if not rates: continue

                                eqp_rows = []
                                for eqp_name in rates.keys():
                                    level1 = ref_data_sg.get(eqp_col)
                                    ref_eqp_data = level1.get(eqp_name, {}) if isinstance(level1, dict) else ref_data_sg.get(eqp_name, {})
                                    ref_count = ref_eqp_data.get('count', 0)
                                    ref_rate_val = ref_eqp_data.get('rate', 0.0)

                                    level1_daily = daily_data_sg.get(eqp_col)
                                    daily_eqp_data = level1_daily.get(eqp_name, {}) if isinstance(level1_daily, dict) else daily_data_sg.get(eqp_name, {})
                                    daily_count = daily_eqp_data.get('count', 0)
                                    daily_rate_val = daily_eqp_data.get('rate', 0.0)

                                    gap_val = rates.get(eqp_name, 0.0)

                                    #  '불량'에 소분류 사용
                                    eqp_rows.append({
                                        '불량': sub_group,  # ← BROKEN → LAP_BROKEN
                                        '구분': proc,
                                        '장비': eqp_name,
                                        'Ref_Count': ref_count,
                                        'Daily_Count': daily_count,
                                        'Ref_rate': ref_rate_val / 100.0,
                                        'Daily_rate': daily_rate_val / 100.0,
                                        'Gap': gap_val
                                    })

                                eqp_rows_sorted = sorted(eqp_rows, key=lambda x: abs(x['Gap']), reverse=True)[:3]
                                table_rows.extend(eqp_rows_sorted)

                    # 표 작성
                    if table_rows:
                        current_defect = None  # 소분류 기준
                        current_process = None
                        defect_merge_start = None
                        process_merge_start = None

                        for r_idx, row in enumerate(table_rows, table_start_row + 1):
                            # ✅ 소분류 기준으로 병합
                            if row['불량'] != current_defect:
                                if defect_merge_start is not None and r_idx - 1 > defect_merge_start:
                                    ws.merge_cells(f'A{defect_merge_start}:A{r_idx - 1}')
                                    ws[f'A{defect_merge_start}'].alignment = Alignment(horizontal='center', vertical='center')
                                current_defect = row['불량']
                                defect_merge_start = r_idx

                            if row['구분'] != current_process:
                                if process_merge_start is not None and r_idx - 1 > process_merge_start:
                                    ws.merge_cells(f'B{process_merge_start}:B{r_idx - 1}')
                                    ws[f'B{process_merge_start}'].alignment = Alignment(horizontal='center', vertical='center')
                                current_process = row['구분']
                                process_merge_start = r_idx

                            for col, key, fmt, align in [
                                (1, '불량', None, 'center'),
                                (2, '구분', None, 'center'),
                                (3, '장비', None, 'center'),
                                (4, 'Ref_Count', '#,##0', 'center'),
                                (5, 'Daily_Count', '#,##0', 'center'),
                                (6, 'Ref_rate', '0.00%', 'center'),
                                (7, 'Daily_rate', '0.00%', 'center'),
                                (8, 'Gap', '+0.00%;-0.00%;0.00%', 'center')
                            ]:
                                cell = ws.cell(row=r_idx, column=col, value=row[key])
                                if fmt: cell.number_format = fmt
                                cell.font = Font(size=9)
                                cell.alignment = Alignment(horizontal=align, vertical='center')
                                cell.border = Border(left=Side(style='thin'), right=Side(style='thin'),
                                                    top=Side(style='thin'), bottom=Side(style='thin'))
                                if col == 8:
                                    gap_num = row['Gap']
                                    if gap_num > 0:
                                        cell.fill = PatternFill("solid", fgColor="FFCCCC")
                                        cell.font = Font(color="FF0000", size=9)
                                    elif gap_num < 0:
                                        cell.fill = PatternFill("solid", fgColor="CCE5FF")
                                        cell.font = Font(color="0000FF", size=9)

                        # 마지막 병합
                        if defect_merge_start:
                            merge_end = table_start_row + len(table_rows)
                            if merge_end > defect_merge_start:
                                ws.merge_cells(f'A{defect_merge_start}:A{merge_end}')
                                ws[f'A{defect_merge_start}'].alignment = Alignment(horizontal='center', vertical='center')
                        if process_merge_start:
                            merge_end = table_start_row + len(table_rows)
                            if merge_end > process_merge_start:
                                ws.merge_cells(f'B{process_merge_start}:B{merge_end}')
                                ws[f'B{process_merge_start}'].alignment = Alignment(horizontal='center', vertical='center')

                        # 모수 행
                        last_row = table_start_row + len(table_rows) + 1
                        ws.merge_cells(f'A{last_row}:B{last_row}')
                        ws.cell(row=last_row, column=1, value="모수").font = Font(size=9, bold=True)
                        ws.cell(row=last_row, column=1).alignment = Alignment(horizontal='center', vertical='center')
                        ws.cell(row=last_row, column=3, value="")
                        cell4 = ws.cell(row=last_row, column=4, value=float(getattr(self, 'avg_in_qty', 0)))
                        cell4.number_format = '#,##0'
                        cell4.font = Font(size=9)
                        cell4.alignment = Alignment(horizontal='center', vertical='center')
                        cell5 = ws.cell(row=last_row, column=5, value=float(getattr(self, 'total_daily_qty', 0)))
                        cell5.number_format = '#,##0'
                        cell5.font = Font(size=9)
                        cell5.alignment = Alignment(horizontal='center', vertical='center')
                        for col in [6, 7, 8]:
                            ws.cell(row=last_row, column=col, value="")
                        for col in range(1, 9):
                            ws.cell(row=last_row, column=col).border = Border(left=Side(style='thin'), right=Side(style='thin'),
                                                                            top=Side(style='thin'), bottom=Side(style='thin'))

                        current_row = last_row + 2
                    else:
                        current_row = next_row_after_graph + 3

               
            # ──────────────────────────────────────────────────
            # 7. [장비별 불량률 추세 분석] - _plot_rej_group_top3_eqp_trend 결과 삽입
            # ──────────────────────────────────────────────────
            current_row += 5
            ws[f'A{current_row}'] = "[장비별 불량률 추세 분석]"
            ws[f'A{current_row}'].font = Font(size=12, bold=True)
            current_row += 2

            # self.data → report 복사 (중복 방지)
            if hasattr(self, 'data') and 'EQP_TREND_GRAPHS' not in report:
                eqp_graphs = self.data.get('EQP_TREND_GRAPHS')
                if eqp_graphs:
                    report['EQP_TREND_GRAPHS'] = eqp_graphs
                    print(f"[INFO] EQP_TREND_GRAPHS report에 복사됨: {list(eqp_graphs.keys())}")

            # top3_rej_groups 가져오기
            top3_rej_groups = self.data.get('DATA_3210_wafering_300', {}).get('top3_rej_groups', [])
            # top3_rej_groups = ['BROKEN', 'EDGE', 'CHIP']
            valid_groups = [g for g in top3_rej_groups if g in ['PIT', 'SCRATCH', 'EDGE', 'BROKEN', 'CHIP', 'VISUAL']]
            eqp_trend_graphs = report.get('EQP_TREND_GRAPHS', {})

            # SCRATCH, BROKEN, CHIP 순서 보장
            display_groups = ['PIT', 'SCRATCH', 'EDGE', 'BROKEN', 'CHIP', 'VISUAL']

            for rej_group in display_groups:
                if rej_group not in valid_groups:
                    current_row += 1
                    continue
                # BROKEN, CHIP은 하위 그룹 합쳐서 그래프 가져오기
                paths = []
                if rej_group == 'BROKEN':
                    sub_groups = ['LAP_BROKEN', 'EP_BROKEN', 'DSP_BROKEN', 'FP_BROKEN']
                    for sg in sub_groups:
                        if sg in eqp_trend_graphs:
                            paths.extend(eqp_trend_graphs[sg])
                elif rej_group == 'CHIP':
                    sub_groups = ['CHIP-LAP', 'EDGE-CHIP', 'CHIP_EG1AF', 'CHIP_EG1BF', 'E_CHIP']
                    for sg in sub_groups:
                        if sg in eqp_trend_graphs:
                            paths.extend(eqp_trend_graphs[sg])
                else:
                    paths = eqp_trend_graphs.get(rej_group, [])
                if not paths:
                    ws.cell(row=current_row, column=1, value=f"{rej_group}: 그래프 없음").font = Font(size=10, color="FF0000")
                    current_row += 1  # 다음 행으로 이동
                    continue

                ws.cell(row=current_row, column=1, value=f"{rej_group} 그룹").font = Font(size=10, bold=True) # 현재 행에 제목 추가 (선택)
                graph_row = current_row + 1

                # 장비별 그래프 삽입 (최대 3개, A, G, M 열)
                for idx, path in enumerate(paths):
                    path_obj = Path(path)
                    if not path_obj.exists():
                        col_letter = ['A', 'F', 'K'][idx % 3]
                        current_col = 1 + (idx % 3) * 5
                        current_row_target = graph_row + (idx // 3) * 11  # 🔹 동적 행 계산
                        ws.cell(row=current_row_target, column=current_col,
                                value=f"{rej_group}: 파일 없음").font = Font(size=9, color="FF0000")
                        continue

                    try:
                        img_path = str(path_obj).replace('\\', '/')
                        img = ExcelImage(img_path)
                        img.width = 400
                        img.height = 200
                        col_letter = ['A', 'F', 'K'][idx % 3]
                        current_row_target = graph_row + (idx // 3) * 11
                        ws.add_image(img, f'{col_letter}{current_row_target}')
                        print(f"[OK] 삽입됨: {img_path} → {col_letter}{current_row_target}")
                    except Exception as e:
                        print(f"[ERROR] 삽입 실패: {e} | idx={idx}, path={path}")
                        current_row_target = graph_row + (idx // 3) * 11
                        ws.cell(row=current_row_target, column=1 + (idx % 3) * 5,
                                value=f"{rej_group} 실패").font = Font(size=9, color="FF0000")

                # 그래프 수에 따라 동적 행 이동
                total_rows_needed = ((len(paths) - 1) // 3 + 1) * 11
                current_row += total_rows_needed

                print(f"[DEBUG] {rej_group}: graph_row={graph_row}, paths={len(paths)}, total_move={total_rows_needed}")

            # 열 너비
            for col, width in zip('ABCDEFGHIJ', [15] + [13] + [13] +[14] +[12]*9):
                ws.column_dimensions[col].width = width

            wb.save(str(excel_path))
            print(f"Excel 저장 성공: {excel_path}")
            return str(excel_path)

        except Exception as e:
            logger.error(f"Excel 생성 실패: {e}")
            raise
