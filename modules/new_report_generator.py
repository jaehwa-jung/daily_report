import matplotlib.pyplot as plt
import pandas as pd
import numpy as np
import os
from datetime import datetime, timedelta
import logging
from logging import FileHandler
import matplotlib
from pathlib import Path
import base64
from analysis.defect_analyzer import analyze_flatness, analyze_warp, analyze_growing, analyze_broken, analyze_nano, analyze_pit, analyze_scratch, analyze_chip, analyze_edge, analyze_HUMAN_ERR, analyze_VISUAL, analyze_NOSALE, analyze_OTHER, analyze_GR, analyze_sample,analyze_particle
from config.mappings import REJ_GROUP_TO_MID_MAPPING

from openpyxl import Workbook
from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
\1
# =========================================================
# Excel Export Utilities (공통화: 이미지/표/시트 빌더)
# - PNG 파일 I/O 최소화: (가능하면) 메모리 이미지로 처리
# - DataFrame 표 생성 공통화
# - 커서 기반 SheetBuilder 제공
# =========================================================
from io import BytesIO

try:
    from PIL import Image as PILImage  # optional (권장)
except Exception:  # pragma: no cover
    PILImage = None


# ---- 공통 스타일 상수 (재사용) ----
_THIN = Side(style="thin")
_BORDER_THIN = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)

HEADER_FONT = Font(bold=True, size=10)
BODY_FONT = Font(size=9)
CENTER_WRAP = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)

HEADER_FILL = PatternFill("solid", fgColor="D3D3D3")


def add_png_image(ws, image_path, anchor, width=None, height=None):
    """PNG 파일 경로를 openpyxl Image로 삽입 (공통)."""
    img = ExcelImage(str(image_path))
    if width is not None:
        img.width = width
    if height is not None:
        img.height = height
    ws.add_image(img, anchor)
    return img


def fig_to_excel_image(fig, width=None, height=None, dpi=150):
    """matplotlib figure를 메모리(PNG)로 렌더링 후 ExcelImage로 변환."""
    bio = BytesIO()
    fig.savefig(bio, format="png", dpi=dpi, bbox_inches="tight")
    bio.seek(0)

    # openpyxl Image는 경로 또는 PIL Image를 받을 수 있음 (버전에 따라 다름)
    if PILImage is not None:
        pil_img = PILImage.open(bio)
        img = ExcelImage(pil_img)
    else:
        # PIL이 없으면 BytesIO를 임시파일로 처리하는 fallback
        import tempfile
        with tempfile.NamedTemporaryFile(suffix=".png", delete=False) as tmp:
            tmp.write(bio.getvalue())
            tmp_path = tmp.name
        img = ExcelImage(tmp_path)

    if width is not None:
        img.width = width
    if height is not None:
        img.height = height
    return img


def write_df_table(
    ws,
    df,
    start_row,
    start_col,
    *,
    header_fill=HEADER_FILL,
    header_font=HEADER_FONT,
    body_font=BODY_FONT,
    alignment=CENTER_WRAP,
    border=_BORDER_THIN,
    number_formats=None,  # 예: {"월 목표": "0.00%", "Gap(월)": "0.00%"}
):
    """DataFrame을 표 형태로 시트에 쓰는 공통 함수."""
    if df is None or getattr(df, "empty", True):
        ws.cell(start_row, start_col, "표 없음").font = Font(size=10, color="FF0000")
        return start_row + 1

    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), start_row):
        for c_offset, value in enumerate(row):
            c_idx = start_col + c_offset
            cell = ws.cell(row=r_idx, column=c_idx, value=value)

            cell.border = border
            cell.alignment = alignment
            if r_idx == start_row:
                cell.font = header_font
                cell.fill = header_fill
            else:
                cell.font = body_font

                if number_formats:
                    col_name = df.columns[c_offset] if c_offset < len(df.columns) else None
                    fmt = number_formats.get(col_name) if col_name else None
                    if fmt:
                        cell.number_format = fmt

    return start_row + len(df) + 1


class SheetBuilder:
    """시트에 순차적으로 내용을 쌓아가는 커서 기반 빌더."""
    def __init__(self, ws, start_row=1, start_col=1):
        self.ws = ws
        self.row = start_row
        self.col = start_col

    def blank(self, n=1):
        self.row += n
        return self

    def title(self, text, merge_from=None, merge_to=None, font=None, alignment=LEFT):
        if font is None:
            font = Font(size=14, bold=True)
        if merge_from and merge_to:
            self.ws.merge_cells(f"{merge_from}:{merge_to}")
            cell = self.ws[merge_from]
        else:
            cell = self.ws.cell(self.row, self.col)
        cell.value = text
        cell.font = font
        cell.alignment = alignment
        if not (merge_from and merge_to):
            self.row += 1
        return self

    def image_from_path(self, image_path, anchor, width=None, height=None):
        add_png_image(self.ws, image_path, anchor, width=width, height=height)
        return self

    def table(self, df, start_row=None, start_col=None, **kwargs):
        if start_row is None:
            start_row = self.row
        if start_col is None:
            start_col = self.col
        self.row = write_df_table(self.ws, df, start_row, start_col, **kwargs)
        return self

import tempfile
from inspect import signature
import re



# 한글 폰트 설정
matplotlib.rcParams['font.family'] = 'Malgun Gothic'  # Windows
matplotlib.rcParams['font.size'] = 10
matplotlib.rcParams['axes.unicode_minus'] = False  # 마이너스 기호 깨짐 방지


# 결과 저장 폴더
REPORT_DIR = "./daily_reports_debug"
os.makedirs(REPORT_DIR, exist_ok=True)

# 기존 로거 설정 대체
logger = logging.getLogger(__name__)
logger.setLevel(logging.INFO)

# 기존 핸들러 제거
if logger.hasHandlers():
    logger.handlers.clear()

# UTF-8로 기록하는 FileHandler 추가
file_handler = FileHandler('daily_report.log', encoding='utf-8')
formatter = logging.Formatter('%(asctime)s - %(name)s - %(levelname)s - %(message)s')
file_handler.setFormatter(formatter)
logger.addHandler(file_handler)

class DailyReportGenerator:
    _ms6_mapping_cache = None

    def __init__(self, data):
        self.data = data

    def _calculate_product_influence(self, df, target_rej_groups):
        """제품 영향성 공통 집계 로직 (Ref/Daily 공통)"""
        if df is None or getattr(df, "empty", True):
            return pd.DataFrame()

        df = df.copy()

        # 숫자 컬럼 타입 보정
        for col in ["IN_QTY", "LOSS_QTY"]:
            if col in df.columns:
                df[col] = pd.to_numeric(df[col], errors="coerce").fillna(0).astype("int64")
            else:
                df[col] = 0

        # 1) 불량개수
        df_defect = df[
            df["REJ_GROUP"].isin(target_rej_groups) & (df["PRODUCT_TYPE"] != "Unknown")
        ].copy()

        if df_defect.empty:
            return pd.DataFrame()

        defect_summary = (
            df_defect.groupby(["REJ_GROUP", "PRODUCT_TYPE"], dropna=False)["LOSS_QTY"]
            .sum()
            .reset_index()
        )
        defect_summary.rename(columns={"LOSS_QTY": "불량개수"}, inplace=True)

        # 2) Compile 수량 (분모)
        df_denom = df[(df["REJ_GROUP"] == "분모") & (df["PRODUCT_TYPE"] != "Unknown")].copy()
        if df_denom.empty:
            return pd.DataFrame()

        compile_summary = (
            df_denom.groupby("PRODUCT_TYPE", dropna=False)["IN_QTY"]
            .sum()
            .reset_index()
        )
        compile_summary.rename(columns={"IN_QTY": "Compile_수량"}, inplace=True)

        total_volume = compile_summary["Compile_수량"].sum()
        if total_volume == 0:
            compile_summary["물량비(%)"] = 0.0
        else:
            compile_summary["물량비(%)"] = (compile_summary["Compile_수량"] / total_volume * 100).round(2)

        # 3) 병합
        result = pd.merge(defect_summary, compile_summary, on="PRODUCT_TYPE", how="left")

        # 4) 불량률 계산
        if total_volume == 0:
            result["불량률(%)"] = 0.0
            result["전체 불량률(%)"] = 0.0
        else:
            result["불량률(%)"] = ((result["불량개수"] / result["Compile_수량"]) * 100).round(2)
            result["전체 불량률(%)"] = ((result["불량개수"] / total_volume) * 100).round(2)

        # 5) 최종 정리
        result = result[
            ["REJ_GROUP", "PRODUCT_TYPE", "불량개수", "Compile_수량", "불량률(%)", "전체 불량률(%)", "물량비(%)"]
        ].sort_values(["REJ_GROUP", "불량률(%)"], ascending=[True, False])

        return result

    
    # ──────────────────────────────────────────────────
    # [신규] MS6.csv 기반 제품 정보 병합 함수
    # ──────────────────────────────────────────────────
    @classmethod
    def _load_ms6_mapping(cls):
        """MS6.csv에서 (MS6 -> 제품1) 매핑을 1회만 로드하여 캐시"""
        if cls._ms6_mapping_cache is not None:
            return cls._ms6_mapping_cache

        project_root = Path(__file__).parent.parent
        ms6_path = project_root / "queries" / "MS6.csv"

        if not ms6_path.exists():
            cls._ms6_mapping_cache = {}
            return cls._ms6_mapping_cache

        try:
            df_ms6 = pd.read_csv(ms6_path, dtype=str, encoding="utf-8")
        except UnicodeDecodeError:
            df_ms6 = pd.read_csv(ms6_path, dtype=str, encoding="cp949")

        if "MS6" not in df_ms6.columns or "제품1" not in df_ms6.columns:
            cls._ms6_mapping_cache = {}
            return cls._ms6_mapping_cache

        df_ms6 = df_ms6.dropna(subset=["MS6", "제품1"]).copy()
        cls._ms6_mapping_cache = dict(
            zip(df_ms6["MS6"].astype(str).str.strip(), df_ms6["제품1"].astype(str).str.strip())
        )
        return cls._ms6_mapping_cache

    def _merge_product_type(self, df):
        """df에 PROD_ID 기준으로 MS6.csv의 '제품1' 컬럼을 병합하여 'PRODUCT_TYPE' 추가 (캐시 적용)"""
        if df is None or getattr(df, "empty", True):
            # empty or None
            try:
                df = df.copy()
            except Exception:
                df = pd.DataFrame()
            df["PRODUCT_TYPE"] = "Unknown"
            return df

        if "PROD_ID" not in df.columns:
            df = df.copy()
            df["PRODUCT_TYPE"] = "Unknown"
            return df

        mapping = self._load_ms6_mapping()
        df = df.copy()
        df["MS6"] = df["PROD_ID"].astype(str).str[:6]
        df["PRODUCT_TYPE"] = df["MS6"].map(mapping).fillna("Unknown")
        return df

    def _get_top3_rej_groups(self):
        """
        안전하게 상위 3개 REJ_GROUP 목록 가져오기
        """
        return self.data.get('DATA_3210_wafering_300', {}).get('top3_rej_groups', [])


    def _create_product_influence_ref(self):
        """전 반기(6개월) 데이터 기반 제품 영향성 Ref 데이터 생성"""
        PROJECT_ROOT = Path(__file__).parent.parent
        cache_dir = PROJECT_ROOT / "data_cache"
        pattern = "DATA_LOT_3210_wafering_300_*.parquet"
        parquet_files = list(cache_dir.glob(pattern))

        # 대상 월 설정: 202506 ~ 202512
        target_months = [f"2025{str(m).zfill(2)}" for m in range(6, 13)]

        target_rej_groups = ["PARTICLE", "FLATNESS", "NANO", "WARP&BOW", "GROWING", "SCRATCH", "VISUAL", "SAMPLE"]

        df_list = []
        for file_path in parquet_files:
            try:
                stem = file_path.stem
                date_part = stem.split("_")[-1]
                if len(date_part) == 6 and date_part.isdigit() and date_part in target_months:
                    df_part = pd.read_parquet(file_path)
                    df_part = self._merge_product_type(df_part)

                    for col in ["IN_QTY", "LOSS_QTY"]:
                        if col in df_part.columns:
                            df_part[col] = pd.to_numeric(df_part[col], errors="coerce").fillna(0).astype("int64")
                        else:
                            df_part[col] = 0

                    if "PRODUCT_TYPE" not in df_part.columns:
                        continue
                    df_list.append(df_part)
            except Exception:
                continue

        if not df_list:
            return pd.DataFrame()

        df_full = pd.concat(df_list, ignore_index=True)

        if "PRODUCT_TYPE" not in df_full.columns:
            return pd.DataFrame()

        return self._calculate_product_influence(df_full, target_rej_groups)

    def _create_product_influence_daily(self):
        """금일 DATA_LOT_3210_wafering_300 데이터 기반 제품 영향성 분석"""
        key = "DATA_LOT_3210_wafering_300"
        if key not in self.data or self.data[key].empty:
            return pd.DataFrame()

        df = self.data[key].copy()

        if "PRODUCT_TYPE" not in df.columns:
            df = self._merge_product_type(df)
            if "PRODUCT_TYPE" not in df.columns:
                return pd.DataFrame()

        target_rej_groups = ["PARTICLE", "FLATNESS", "NANO", "WARP&BOW", "GROWING", "SCRATCH", "VISUAL", "SAMPLE"]
        return self._calculate_product_influence(df, target_rej_groups)

    def _analyze_product_influence_gap(self):
        """
        제품별 불량률 GAP 분석: 6개월 기준(Ref) vs 금일(Daily)
        - 기준: REJ_GROUP + PRODUCT_TYPE
        - 출력: 불량률(%) GAP, 전체 불량률(%) GAP
        - 필터: _get_top3_rej_groups() 기반
        """

        # 1. Ref 데이터 확인
        if 'product_influence_ref' not in self.data:
            print("product_influence_ref 데이터 없음")
            return pd.DataFrame()
        
        ref_df = self.data['product_influence_ref']
        if ref_df.empty:
            print("product_influence_ref 데이터가 비어 있음")
            return pd.DataFrame()

        # 2. Daily 데이터 확인
        if 'product_influence_daily' not in self.data:
            print("product_influence_daily 데이터 없음")
            return pd.DataFrame()
        
        daily_df = self.data['product_influence_daily']
        if daily_df.empty:
            print("product_influence_daily 데이터가 비어 있음")
            return pd.DataFrame()

        # 3. 컬럼 선택 및 이름 변경
        key_cols = ['REJ_GROUP', 'PRODUCT_TYPE']
        ref = ref_df[key_cols + ['불량개수', 'Compile_수량','불량률(%)', '전체 불량률(%)', '물량비(%)']].copy()
        ref.rename(columns={
            '불량개수' : 'Ref_불량개수',
            'Compile_수량' : 'Ref_Compile_수량',
            '불량률(%)': 'Ref_불량률(%)',
            '전체 불량률(%)': 'Ref_전체_불량률(%)',
            '물량비(%)' : 'Ref_물량비(%)'
        }, inplace=True)

        daily = daily_df[key_cols + ['불량개수', 'Compile_수량','불량률(%)', '전체 불량률(%)', '물량비(%)']].copy()
        daily.rename(columns={
            '불량개수' : 'Daily_불량개수',
            'Compile_수량' : 'Daily_Compile_수량',
            '불량률(%)': 'Daily_불량률(%)',
            '전체 불량률(%)': 'Daily_전체_불량률(%)',
            '물량비(%)' : 'Daily_물량비(%)'
        }, inplace=True)

        # 4. 병합 (외부 조인 → 누락 데이터 보존)
        gap = pd.merge(daily, ref, on=key_cols, how='outer').fillna(0.0)

        # 5. GAP 계산
        gap['불량률_GAP(%)'] = (gap['Daily_불량률(%)'] - gap['Ref_불량률(%)']).round(2)
        gap['전체_불량률_GAP(%)'] = (gap['Daily_전체_불량률(%)'] - gap['Ref_전체_불량률(%)']).round(2)
        gap['물량비_GAP(%)'] = (gap['Daily_물량비(%)'] - gap['Ref_물량비(%)']).round(2)
        gap['물량비_불량GAP'] = ((gap['Ref_불량률(%)'] - gap['Ref_전체_불량률(%)']) * gap['물량비_GAP(%)']).round(2)

        # 5. 상위 3개 REJ_GROUP 필터링
        top3_rej_groups = self._get_top3_rej_groups()
        if not top3_rej_groups:
            print("상위 3개 REJ_GROUP 없음 → 전체 데이터 사용")
            filtered_gap = gap
        else:
            print(f"필터링 기준: {top3_rej_groups}")
            filtered_gap = gap[gap['REJ_GROUP'].isin(top3_rej_groups)]

        if filtered_gap.empty:
            print("필터링 후 데이터 없음")
            return pd.DataFrame()

        # 5. 각 REJ_GROUP별로 불량률_GAP(%) 기준 상위 3개씩 추출
        top3_per_group_list = []

        for rej_group in top3_rej_groups:
            group_data = filtered_gap[filtered_gap['REJ_GROUP'] == rej_group]
            if group_data.empty:
                continue
            # GAP 기준 상위 3개
            top3_in_group = group_data.nlargest(3, '물량비_불량GAP')
            top3_per_group_list.append(top3_in_group)

        # 6. 병합
        if not top3_per_group_list:
            print("각 그룹별 상위 3개 추출 실패")
            return pd.DataFrame()

        final_result = pd.concat(top3_per_group_list, ignore_index=True)

        # 7. 정렬: REJ_GROUP → 전체_불량률_영향성 내림차순
        final_result = final_result.sort_values(
            ['REJ_GROUP', '물량비_불량GAP'],
            ascending=[True, False]
        ).reset_index(drop=True)

        print(f"최종 출력: 각 REJ_GROUP별 GAP 상위 3개 제품")
        print(f"결과 (총 {len(final_result)} 건):\n{final_result}")

        return final_result  

    def generate(self):
        """데일리 리포트 생성"""
        try:
            logger.info("리포트 생성 시작")
            # ===================================================================
            # 모든 데이터에 PRODUCT_TYPE 일괄 병합 (가장 먼저 실행)
            # ===================================================================

            for key in ['DATA_LOT_3210_wafering_300', 'DATA_WAF_3210_wafering_300']:
                if key in self.data and not self.data[key].empty:
                    self.data[key] = self._merge_product_type(self.data[key])
                    if 'PRODUCT_TYPE' in self.data[key].columns:
                        sample = self.data[key].sample(1)[['PROD_ID', 'PRODUCT_TYPE']].to_dict('records')
                else:
                    print(f"{key} 없거나 빈 데이터")


            product_influence_ref = self._create_product_influence_ref() #[신규] 제품 영향성 Ref 데이터 생성

            # 3010 보고서 생성
            data_3010_details = self._create_3010_wafering_300()

            # 1. DATA_3210_wafering_300 생성 + 저장
            data_3210_details = self._create_DATA_3210_wafering_300()
            self.data['DATA_3210_wafering_300'] = data_3210_details

            # 2. 제품 영향성 분석
            product_influence_ref = self._create_product_influence_ref()
            product_influence_daily = self._create_product_influence_daily()

            self.data['product_influence_ref'] = product_influence_ref
            self.data['product_influence_daily'] = product_influence_daily

            # 3. GAP 분석 실행 
            product_influence_gap = self._analyze_product_influence_gap()

            # 2. DATA_3210_wafering_300_3months 생성 + 저장 (핵심!)
            data_3210_3months = self._create_DATA_3210_wafering_300_3months()
            self.data['DATA_3210_wafering_300_3months'] = data_3210_3months  

            data_lot_details = self._create_DATA_LOT_3210_wafering_300()
            data_waf_details = self._create_DATA_WAF_3210_wafering_300()

            DATA_1511_SMAX_wafering_300 = self.data.get('DATA_1511_SMAX_wafering_300')

            report = {
                'DATA_3010_wafering_300' : data_3010_details,
                'DATA_3210_wafering_300_details': data_3210_details,
                'DATA_3210_wafering_300_3months': data_3210_3months,
                'DATA_LOT_3210_wafering_300_details': data_lot_details,
                'DATA_WAF_3210_wafering_300_details': data_waf_details,
                'product_influence_gap' : product_influence_gap,
                'DATA_1511_SMAX_wafering_300' : DATA_1511_SMAX_wafering_300,
                'raw_data': self.data
            }
            
            # Excel 생성 시 report 전체 전달
            try:
                excel_path = self._export_to_excel(report, output_dir="./daily_reports_debug")
                report['excel_report'] = str(excel_path)
                print(f"Excel 보고서도 생성됨: {excel_path}")
            except Exception as e:
                print(f"Excel 생성 실패: {e}")
                report['excel_report'] = None

            logger.info("리포트 생성 완료")
            return report
        except Exception as e:
            logger.error(f"리포트 생성 실패: {e}")
            raise
    

    def _create_3010_wafering_300(self):
        """3010 수율 데이터 분석 및 그래프 생성 (WF RTY만, 최신 일실적 기준)"""
        details = {}

        if 'DATA_3010_wafering_300' not in self.data or self.data['DATA_3010_wafering_300'].empty:
            print("DATA_3010_wafering_300 데이터 없음 또는 비어 있음")
            return details

        df = self.data['DATA_3010_wafering_300'].copy()

         # --- 전처리 ---
        df['rate'] = pd.to_numeric(df['rate'], errors='coerce')
        df['item_type'] = df['item_type'].astype(str).str.strip()

        # dt_range_raw: 문자열 정리
        df['dt_range_raw'] = df['dt_range'].astype(str).str.strip()

        # item_type에 따라 파싱 전략 분기
        def parse_date(row):
            raw = row['dt_range_raw']
            item_type = row['item_type']
            
            if item_type in ['월실적', '월사업계획']:
                return pd.to_datetime(raw, format='%Y-%m', errors='coerce')
            else:
                return pd.to_datetime(raw, format='%Y-%m-%d', errors='coerce')

        df['dt_range'] = df.apply(parse_date, axis=1)

        # month_str 생성
        df['month_str'] = df['dt_range'].dt.strftime('%Y-%m')
        current_month = (datetime.now() - timedelta(days=1)).strftime('%Y-%m')

        # ──────────────────────────────────────────────────
        # 1. 월 목표/실적
        # ──────────────────────────────────────────────────
        monthly_plan = df[
            (df['item_type'] == '월사업계획') &
            (df['month_str'] == current_month)
        ].copy()
        monthly_plan_val = float(monthly_plan['rate'].iloc[0]) if not monthly_plan.empty else 0.0

        monthly_actual = df[
            (df['item_type'] == '월실적') &
            (df['month_str'] == current_month)
        ].copy()
        monthly_actual_val = float(monthly_actual['rate'].iloc[0]) if not monthly_actual.empty else 0.0

        # ──────────────────────────────────────────────────
        # 2. 기준일: 어제
        # ──────────────────────────────────────────────────
        target_date = (datetime.now().date() - timedelta(days=1))  # 2026-02-03
        print(f"기준일: {target_date}")

        # ──────────────────────────────────────────────────
        # 3. 재사용 함수 정의
        # ──────────────────────────────────────────────────
        def get_latest_or_target(df, item_type, target_date):
            # 동일 날짜 찾기
            same_day = df[
                (df['item_type'] == item_type) &
                (df['dt_range'].notna()) &
                (df['dt_range'].dt.date == target_date)
            ]
            if not same_day.empty:
                return same_day.iloc[0]

            # 없으면 최신 날짜 사용
            latest = df[
                (df['item_type'] == item_type) &
                (df['dt_range'].notna())
            ]
            if not latest.empty:
                return latest.sort_values('dt_range', ascending=False).iloc[0]
            return None

        # ──────────────────────────────────────────────────
        # 4. 일 실적: 어제 기준 → 없으면 최신
        # ──────────────────────────────────────────────────
        daily_actual_row = get_latest_or_target(df, '일실적', target_date)
        if daily_actual_row is not None:
            daily_actual_val = float(daily_actual_row['rate'])
            daily_actual_date = daily_actual_row['dt_range'].strftime('%Y-%m-%d')
        else:
            daily_actual_val = 0.0
            daily_actual_date = "N/A"
            print("일 실적: 데이터 없음")

        # ──────────────────────────────────────────────────
        # 5. 일 목표: 어제 기준 → 없으면 최신
        # ──────────────────────────────────────────────────
        daily_plan_row = get_latest_or_target(df, '일사업계획', target_date)
        if daily_plan_row is not None:
            daily_plan_val = float(daily_plan_row['rate'])
            daily_plan_date = daily_plan_row['dt_range'].strftime('%Y-%m-%d')
        else:
            daily_plan_val = 0.0
            print("일 목표: 데이터 없음")

        # ──────────────────────────────────────────────────
        # 4. 그래프 생성
        # ──────────────────────────────────────────────────
        # PROJECT_ROOT 및 날짜 폴더
        PROJECT_ROOT = Path(__file__).parent.parent
        base_date = (datetime.now().date() - timedelta(days=1))
        date_folder_name = base_date.strftime("%Y%m%d")
        debug_dir = PROJECT_ROOT / "daily_reports_debug" / date_folder_name
        debug_dir.mkdir(exist_ok=True, parents=True)

        chart_path = debug_dir / "3010_yield_chart.png"

        if chart_path.exists():
            chart_path.unlink() #파일 삭제
            print(f"기존 그래프 파일 삭제됨 : {chart_path}")

        fig, ax = plt.subplots(figsize=(12, 6))
        # X축 레이블: [월, 일] → 각각 2개의 카테고리 (WF RTY, WF OAY)
        # 현재는 WF RTY만 사용 중이므로, WF RTY만 표시
        categories = ['WF RTY']
        x_labels = ['월', '일']
        x = np.arange(len(x_labels))  # 월, 일 위치

        # 막대 너비
        bar_width = 0.35

        # 목표/실적 값
        monthly_values = [monthly_plan_val, monthly_actual_val]
        daily_values = [daily_plan_val, daily_actual_val]

        # 색상
        goal_color = 'steelblue'   # 목표
        actual_color = 'orange'     # 실적

        # 월 그룹
        bar1 = ax.bar(x[0] - bar_width/2, monthly_values[0], bar_width, label='목표', color=goal_color)
        bar2 = ax.bar(x[0] + bar_width/2, monthly_values[1], bar_width, label='실적', color=actual_color)

        # 일 그룹
        bar3 = ax.bar(x[1] - bar_width/2, daily_values[0], bar_width, color=goal_color)
        bar4 = ax.bar(x[1] + bar_width/2, daily_values[1], bar_width, color=actual_color)

        # X축 레이블 설정
        ax.set_xticks(x)
        ax.set_xticklabels(x_labels, fontsize=12, fontweight='bold')
        ax.set_xlabel('기간', fontsize=12)

        # Y축 범위
        all_vals = monthly_values + daily_values
        min_ylim = min(88.0, min(all_vals) - 0.3)
        max_ylim = max(98.0, max(all_vals) + 0.3)

        ax.set_ylim(min_ylim, max_ylim)        
        ax.set_ybound(min_ylim, max_ylim)      

        # 제목
        ax.set_title(f'WF RTY 수율 비교 (월/일 목표 vs 실적) - 기준일: {daily_actual_date}', fontsize=14, fontweight='bold')
        ax.set_ylabel('수율 (%)', fontsize=12)
        ax.set_xlabel('기간', fontsize=12)

        # 범례 (목표, 실적)
        ax.legend(loc='upper right', fontsize=10)

        # ──────────────────────────────────────────────────
        # 값 표시: 막대 바로 위
        # ──────────────────────────────────────────────────
        def autolabel(rects, values, color='white'):
            for i, rect in enumerate(rects):
                height = rect.get_height()
                ax.text(
                    rect.get_x() + rect.get_width() / 2.,  # 막대 중앙
                    height + 0.05,                         # 막대 바로 위 (약간 높이)
                    f'{values[i]:.2f}%',                   # 값 표시
                    ha='center', va='bottom',               # 수평 중앙, 수직 아래
                    fontsize=9, fontweight='bold', color=color
                )

        autolabel([bar1[0], bar2[0]], monthly_values, 'black')
        autolabel([bar3[0], bar4[0]], daily_values, 'black')

        # ──────────────────────────────────────────────────
        # Gap 표시: 막대 중간에 수직 정렬
        # ──────────────────────────────────────────────────
        monthly_gap = monthly_actual_val - monthly_plan_val
        daily_gap = daily_actual_val - daily_plan_val

        gap_x = [x[0], x[1]]
        gap_y = [(monthly_plan_val + monthly_actual_val) / 2, (daily_plan_val + daily_actual_val) / 2]

        monthly_gap_color = 'orange' if monthly_gap < 0 else 'steelblue'
        daily_gap_color = 'orange' if daily_gap < 0 else 'steelblue'

        ax.text(
            gap_x[0], gap_y[0],
            f'{monthly_gap:+.2f}%',
            ha='center', va='bottom',  # 수평/수직 중앙
            fontsize=9, fontweight='bold', color=monthly_gap_color
        )
        ax.text(
            gap_x[1], gap_y[1],
            f'{daily_gap:+.2f}%',
            ha='center', va='bottom',  # 수평/수직 중앙
            fontsize=9, fontweight='bold', color=daily_gap_color
        )

        # 그리드
        ax.grid(axis='y', linestyle='--', alpha=0.7)

        # 여백 조정
        plt.tight_layout()
        plt.savefig(chart_path, dpi=300, bbox_inches='tight')
        plt.close()

        # Base64 인코딩
        with open(chart_path, "rb") as img_file:
            img_base64 = base64.b64encode(img_file.read()).decode()

        # ──────────────────────────────────────────────────
        # 5. 표 생성 (DataFrame)
        # ──────────────────────────────────────────────────
        table_data = {
            '항목': ['WF RTY'],
            '월 목표': [monthly_plan_val],
            '월 실적': [monthly_actual_val],
            '일 목표': [daily_plan_val],
            '일 실적': [daily_actual_val],
            'Gap(월)': [monthly_actual_val - monthly_plan_val],
            'Gap(일)': [daily_actual_val - daily_plan_val],
            '기준일': [daily_actual_date]
        }
        table_df = pd.DataFrame(table_data)

        # details 업데이트
        details.update({
            'chart_path': str(chart_path),
            'img_base64': img_base64,
            'table_df': table_df,
            'summary': table_df,
            'daily_actual_date': daily_actual_date  # Excel에 표시용
        })

        return details

    def _create_DATA_3210_wafering_300(self):
        """3210 불량률 상세 분석 """
        details = {}
        
        if 'DATA_3210_wafering_300' not in self.data or self.data['DATA_3210_wafering_300'].empty:
            print("DATA_3210_wafering_300 데이터 없음 또는 비어 있음")
            return details

        df = self.data['DATA_3210_wafering_300'].copy()

        # 컬럼 타입 변환
        numeric_cols = ['LOSS_RATIO', 'GOAL_RATIO', 'GOAL_RATIO_SUM', 'GAP_RATIO', 'LOSS_QTY', 'MGR_QTY']
        for col in numeric_cols:
            if col in df.columns:
                df[col] = pd.to_numeric(df[col], errors='coerce')

        # 그룹별 집계
        summary = df.groupby(['BASE_DT_NM', 'REJ_GROUP'], dropna=False).agg(
            AVG_LOSS_RATIO=('LOSS_RATIO', 'sum'),
            AVG_GOAL_RATIO=('GOAL_RATIO', 'mean'),
            TOTAL_MGR_QTY=('MGR_QTY', 'mean')
        ).reset_index()

        # 백분율 계산
        summary['LOSS_RATIO_PCT'] = (summary['AVG_LOSS_RATIO'] * 100).round(2)
        summary['GOAL_RATIO_PCT'] = (summary['AVG_GOAL_RATIO'] * 100).round(2)
        summary['GAP_PCT'] = (summary['LOSS_RATIO_PCT'] - summary['GOAL_RATIO_PCT']).round(2)

        # 정렬: GAP 큰 순서대로
        summary = summary.sort_values('GAP_PCT', ascending=False).reset_index(drop=True)

        base_date = summary['BASE_DT_NM'].iloc[0] if len(summary) > 0 else "Unknown"
        print(f"분석 대상일: {base_date}")

        # 출력 디렉터리
        PROJECT_ROOT = Path(__file__).parent.parent
        base_date = (datetime.now().date() - timedelta(days=1))
        date_folder_name = base_date.strftime("%Y%m%d")
        debug_dir = PROJECT_ROOT / "daily_reports_debug" / date_folder_name
        debug_dir.mkdir(exist_ok=True, parents=True)

        # ──────────────────────────────────────────────────
        # 1. 그래프 저장 → Base64 인코딩
        # ──────────────────────────────────────────────────
        chart_path = debug_dir / "prime_gap_chart.png"

        if chart_path.exists():
            chart_path.unlink() #파일 삭제
            print(f"기존 그래프 파일 삭제됨 : {chart_path}")


        plt.figure(figsize=(10, 6))
        x = np.arange(len(summary))
        bars = plt.bar(x, summary['GAP_PCT'],
                    color=summary['GAP_PCT'].apply(lambda x: 'orange' if x > 0 else 'steelblue'), linewidth=1)

        # for i, bar in enumerate(bars):
        #     if summary['GAP_PCT'].iloc[i] > 0:
        #         bar.set_edgecolor('red')
        #         bar.set_linewidth(2)

        plt.title(f"Gap 분석 - {base_date}", fontsize=14, fontweight='bold')
        plt.xlabel('REJ_GROUP', fontsize=12)
        plt.ylabel('GAP (%)', fontsize=12)
        plt.xticks(x, summary['REJ_GROUP'], rotation=45, ha='right')

        for i, bar in enumerate(bars):
            height = bar.get_height()
            plt.text(bar.get_x() + bar.get_width() / 2, height + 0.01 * (1 if height >= 0 else -1),
                    f"{height:.2f}%", ha='center', va='bottom' if height >= 0 else 'top',
                    fontsize=9, fontweight='bold')

        plt.ylim(min(-0.15, summary['GAP_PCT'].min() - 0.05), max(1.3, summary['GAP_PCT'].max() + 0.05))
        plt.grid(axis='y', linestyle='--', alpha=0.7)
        plt.tight_layout()
        plt.savefig(chart_path, dpi=300, bbox_inches='tight')
        plt.close()

        # Base64 인코딩
        with open(chart_path, "rb") as img_file:
            img_base64 = base64.b64encode(img_file.read()).decode()

        # ──────────────────────────────────────────────────
        # 2. 상위 3개 불량 상세분석
        # ──────────────────────────────────────────────────
        top3_rej_groups = summary.nlargest(3, 'GAP_PCT')['REJ_GROUP'].tolist()
        print(f"상위 3개 불량: {top3_rej_groups}")

        yesterday_mid_list = []
        for rej_group in top3_rej_groups:
            group_df = df[df['REJ_GROUP'] == rej_group].copy()

            # MID_GROUP 매핑 적용
            mid_mapping = REJ_GROUP_TO_MID_MAPPING.get(rej_group, {})
            group_df['MID_GROUP'] = group_df['AFT_BAD_RSN_CD'].map(mid_mapping)
            group_df['MID_GROUP'] = group_df['MID_GROUP'].fillna(group_df['AFT_BAD_RSN_CD'])

            # MID_GROUP별 평균 LOSS_RATIO 계산
            mid_agg = group_df.groupby('MID_GROUP', dropna=False).agg(
                YESTERDAY_LOSS_RATIO=('LOSS_RATIO', 'mean')
            ).reset_index()

            mid_agg['REJ_GROUP'] = rej_group
            mid_agg['YESTERDAY_LOSS_PCT'] = (mid_agg['YESTERDAY_LOSS_RATIO'] * 100).round(2)
            yesterday_mid_list.append(mid_agg[['REJ_GROUP', 'MID_GROUP', 'YESTERDAY_LOSS_RATIO', 'YESTERDAY_LOSS_PCT']])

        # 전체 yesterday MID_GROUP 실적
        yesterday_mid_summary = pd.concat(yesterday_mid_list, ignore_index=True) if yesterday_mid_list else pd.DataFrame()

        # ──────────────────────────────────────────────────
        # 3. 세부분석: 상위 3개 REJ_GROUP에 해당하는 함수만 실행
        # ──────────────────────────────────────────────────
        detailed_analysis = []

        if not top3_rej_groups:
            detailed_analysis.append("[세부분석] 상위 3개 불량 그룹 없음")
        else:
            print(f"분석 대상 REJ_GROUP: {top3_rej_groups}")
            df_wafer = self.data.get('DATA_WAF_3210_wafering_300')
            df_lot = self.data.get('DATA_LOT_3210_wafering_300')

            if df_wafer is None:
                detailed_analysis.append("[세부분석] DATA_WAF_3210_wafering_300 데이터 없음")
            else:

                # REJ_GROUP → 분석 함수 매핑
                REJ_GROUP_TO_ANALYZER = {
                    'FLATNESS': analyze_flatness,
                    'WARP&BOW': analyze_warp,
                    'GROWING': analyze_growing,
                    'BROKEN': analyze_broken,
                    'NANO': analyze_nano,
                    'PIT': analyze_pit,
                    'SCRATCH': analyze_scratch,
                    'CHIP': analyze_chip,
                    'EDGE': analyze_edge,
                    'HUMAN_ERR': analyze_HUMAN_ERR,
                    'VISUAL': analyze_VISUAL,
                    'NOSALE': analyze_NOSALE,
                    'OTHER': analyze_OTHER,
                    'GR_보증': analyze_GR,
                    'SAMPLE' : analyze_sample,
                    'PARTICLE': analyze_particle
                }

                for rej in top3_rej_groups:
                    rej = rej.strip()
                    if rej not in REJ_GROUP_TO_ANALYZER:
                        detailed_analysis.append(f"[{rej} 분석] 매핑된 분석 함수 없음")
                        continue

                    print(f"  → {rej} 분석 시작")
                    analyzer_func = REJ_GROUP_TO_ANALYZER[rej]

                    # 함수 시그니처 기반 자동 인자 바인딩
                    sig = signature(analyzer_func)
                    params = sig.parameters

                    args = []

                    for param_name in params.keys():
                        if param_name.endswith('wafer'):
                            if df_wafer is not None:
                                df_target = df_wafer[df_wafer['REJ_GROUP'] == rej].copy()
                                args.append(df_target)
                        elif param_name.endswith('lot'):
                            if df_lot is not None:
                                args.append(df_lot)  
                            else:
                                result = [f"[{rej} 분석] DATA_LOT_3210_wafering_300 없음"]
                                break
                    else:
                        # 모든 인자 준비 완료 → 함수 호출
                        result = analyzer_func(*args)

                    detailed_analysis.extend(result)

        # ──────────────────────────────────────────────────
        #  5. details에 top3 + yesterday_mid_summary 저장
        # ──────────────────────────────────────────────────
        details.update({
            'summary': summary,
            'top3_rej_groups': top3_rej_groups,
            'yesterday_mid_summary': yesterday_mid_summary,  # 핵심: MID_GROUP 실적 저장
            'chart_path': str(chart_path),
            'img_base64': img_base64,
            'detailed_analysis': detailed_analysis
        })

        self.top3_rej_groups = top3_rej_groups


        return details


    def _create_DATA_3210_wafering_300_3months(self):
        """3210 불량률 상세 분석(3개월) """
        details = {}
        
        if 'DATA_3210_wafering_300_3months' not in self.data or self.data['DATA_3210_wafering_300_3months'].empty:
            print("DATA_3210_wafering_300_3months 데이터 없음 또는 비어 있음")
            return details

        df = self.data['DATA_3210_wafering_300_3months'].copy()

        # 컬럼 타입 변환
        numeric_cols = ['LOSS_RATIO', 'LOSS_QTY', 'MGR_QTY']
        for col in numeric_cols:
            if col in df.columns:
                df[col] = pd.to_numeric(df[col], errors='coerce')

        # REJ_GROUP별로 중분류(MID_GROUP) 적용
        summary_list = []
        for rej_group, group_df in df.groupby('REJ_GROUP', dropna=False):
            # 해당 REJ_GROUP의 매핑 가져오기
            mid_mapping = REJ_GROUP_TO_MID_MAPPING.get(rej_group, {})
            
            # AFT_BAD_RSN_CD 기준으로 MID_GROUP 생성
            group_df = group_df.copy()
            group_df['MID_GROUP'] = group_df['AFT_BAD_RSN_CD'].map(mid_mapping)
            
            #  매핑되지 않은 경우: 원래 AFT_BAD_RSN_CD 값 유지
            group_df['MID_GROUP'] = group_df['MID_GROUP'].fillna(group_df['AFT_BAD_RSN_CD'])

            # 그룹 집계: REJ_GROUP + MID_GROUP + AFT_BAD_RSN_CD
            agg_df = group_df.groupby(['REJ_GROUP', 'MID_GROUP', 'AFT_BAD_RSN_CD'], dropna=False).agg(
                AVG_LOSS_RATIO=('LOSS_RATIO', 'mean'),
                TOTAL_MGR_QTY=('MGR_QTY', 'mean'),
                COUNT_DAYS=('LOSS_RATIO', 'count')
            ).reset_index()

            summary_list.append(agg_df)

        # 전체 요약 병합
        summary_3months = pd.concat(summary_list, ignore_index=True)
        summary_3months['LOSS_RATIO_PCT'] = (summary_3months['AVG_LOSS_RATIO'] * 100).round(2)

    # yesterday_mid_summary 가져오기
        yesterday_mid = self.data.get('DATA_3210_wafering_300', {}).get('yesterday_mid_summary', pd.DataFrame())
        if yesterday_mid.empty:
            details['summary'] = summary_3months
            return details

        # 상위 3개 REJ_GROUP 가져오기 (Gap 기준)
        top3_rej_groups = self.data.get('DATA_3210_wafering_300', {}).get('top3_rej_groups', [])

        # 3개월 평균 (Ref) 준비
        ref_3months = summary_3months[summary_3months['REJ_GROUP'].isin(yesterday_mid['REJ_GROUP'])].copy()
        ref_3months = ref_3months.groupby(['REJ_GROUP', 'MID_GROUP'], dropna=False).agg(
            REF_AVG_LOSS_RATIO=('AVG_LOSS_RATIO', 'mean')
        ).reset_index()

        # 병합 → Gap 계산 (전체 사용)
        merged = pd.merge(
            yesterday_mid,
            ref_3months,
            on=['REJ_GROUP', 'MID_GROUP'],
            how='inner'
        )

        merged['GAP'] = merged['YESTERDAY_LOSS_PCT'] - merged['REF_AVG_LOSS_RATIO']
        merged['Gap'] = merged['GAP'].round(2)
        merged['실적(%)'] = merged['YESTERDAY_LOSS_PCT']
        merged['Ref(3개월)'] = merged['REF_AVG_LOSS_RATIO'].round(2)
        merged['범례'] = merged['MID_GROUP']

        # 개별 플롯 생성
        plot_paths = self._create_top3_midgroup_plot_per_group(merged, top3_rej_groups)
        # 각 그룹별 표도 상위 3개만
        group_tables = {}
        analysis_text = "[ Prime 주요 열위 불량 세부코드 분석 Ref.(3개월) 比 일실적 변동 (상위 3개) ]\n"
        for rej in top3_rej_groups:
            df_group = merged[merged['REJ_GROUP'] == rej].copy()
            if df_group.empty:
                continue
            top3 = df_group.nlargest(3, 'Gap')[['MID_GROUP', '실적(%)', 'Ref(3개월)', 'Gap']].copy()
            group_tables[rej] = top3

            if len(top3) > 0:
                top_row = top3.iloc[0]
                analysis_text += f"\n {rej} 최대 Gap: {top_row['MID_GROUP']} ({top_row['Gap']:.2f}%)"

        # 최종 details 업데이트
        details.update({
            'summary': summary_3months,
            'top3_midgroup_analysis': {
                'tables': group_tables,
                'plot_paths': plot_paths,
                'analysis': analysis_text.strip()
            }
        })

        return details

    def _create_top3_midgroup_plot_per_group(self, merged_df, top3_rej_groups):
        """
        각 REJ_GROUP별로 Gap 상위 3개 MID_GROUP만 추출하여 개별 막대그래프 생성
        → 결과: {'GR_보증': 'path1.png', 'SAMPLE': 'path2.png', ...}
        """
        # PROJECT_ROOT 및 날짜 폴더
        PROJECT_ROOT = Path(__file__).parent.parent
        base_date = (datetime.now().date() - timedelta(days=1))
        date_folder_name = base_date.strftime("%Y%m%d")
        debug_dir = PROJECT_ROOT / "daily_reports_debug" / date_folder_name
        debug_dir.mkdir(exist_ok=True, parents=True)

        plot_paths = {}

        for rej_group in top3_rej_groups:
            try:
                # 해당 REJ_GROUP 데이터 필터링
                group_df = merged_df[merged_df['REJ_GROUP'] == rej_group].copy()
                if group_df.empty:
                    print(f"{rej_group}: 분석 데이터 없음")
                    continue

                # Gap 기준 상위 3개만 추출
                top3_mids = group_df.nlargest(3, 'Gap')

                # 파일명
                safe_rej = "".join(c if c.isalnum() else "_" for c in rej_group)
                plot_path = debug_dir  / f"prime_midgroup_top3_gap_{safe_rej}.png"

                # 기존 파일 삭제
                if plot_path.exists():
                    plot_path.unlink()

                plt.figure(figsize=(8, 5))
                x = np.arange(len(top3_mids))
                bars = plt.bar(x, top3_mids['Gap'],
                            color=top3_mids['Gap'].apply(lambda x: 'orange' if x > 0 else 'steelblue'), linewidth=1)

                # # Gap > 0인 경우 빨간 테두리 강조
                # for i, bar in enumerate(bars):
                #     if top3_mids['Gap'].iloc[i] > 0:
                #         bar.set_edgecolor('red')
                #         bar.set_linewidth(2)

                plt.title(f"[ {rej_group} 상위 3개 MID_GROUP Gap 분석 ]", fontsize=12, fontweight='bold')
                plt.xlabel('MID_GROUP', fontsize=11)
                plt.ylabel('Gap (%)', fontsize=11)
                plt.xticks(x, top3_mids['MID_GROUP'], rotation=0, ha='center')  #  여기서 rotation=0 → 수평

                # 값 표시
                for i, bar in enumerate(bars):
                    height = bar.get_height()
                    plt.text(bar.get_x() + bar.get_width() / 2, height + 0.01 * (1 if height >= 0 else -1),
                            f"{height:.2f}%", ha='center', va='bottom' if height >= 0 else 'top',
                            fontsize=12, fontweight='bold')

                # y축 범위
                plt.ylim(min(-0.15, top3_mids['Gap'].min() - 0.05), max(1.3, top3_mids['Gap'].max() + 0.05))
                plt.grid(axis='y', linestyle='--', alpha=0.7)
                plt.tight_layout()

                # 저장
                plt.savefig(str(plot_path), dpi=300, bbox_inches='tight')
                plt.close()

                if plot_path.exists():
                    plot_paths[rej_group] = str(plot_path)
                else:
                    raise RuntimeError(f"파일 생성 실패: {plot_path}")

            except Exception as e:
                print(f"{rej_group} 플롯 생성 실패: {e}")
                continue

        return plot_paths

    
    def _create_DATA_LOT_3210_wafering_300(self):
        """3210 LOT 상세 분석 - 캐시된 3개월 데이터 + self.data의 당일 데이터 모두 활용"""

        details = {}

        # ===================================================================
        # 1. [신규] data_cache에서 3개월 데이터 직접 로드 (장기 분석용)
        # ===================================================================
        PROJECT_ROOT = Path(__file__).parent.parent  

        # 어제 날짜 폴더 생성
        base_date = (datetime.now().date() - timedelta(days=1))
        date_folder_name = base_date.strftime("%Y%m%d")  # 예: 20260204

        # 출력 폴더: daily_reports_debug/YYYYMMDD
        debug_dir = PROJECT_ROOT / "daily_reports_debug" / date_folder_name
        debug_dir.mkdir(exist_ok=True, parents=True)  # 폴더 생성

        target_months = []
        current = base_date.replace(day=1)
        for _ in range(3):
            # 전월로 이동
            current = (current - timedelta(days=1)).replace(day=1)
            month_str = current.strftime("%Y%m")
            target_months.append(month_str)

        # 역순 정렬 (과거 → 최근)
        target_months = sorted(target_months)

        print(f"[캐시 필터링] 최근 3개월 대상 월: {target_months}")

        cache_dir = PROJECT_ROOT / "data_cache"
        pattern = "DATA_LOT_3210_wafering_300_*.parquet"
        parquet_files = list(cache_dir.glob(pattern))

        df_cached_3months = pd.DataFrame()

        if parquet_files:
            valid_files = []
            for file_path in parquet_files:
                try:
                    stem = file_path.stem  # 전체 이름 (확장자 제외)
                    date_part = stem.split('_')[-1]  # '202506'

                    if len(date_part) != 6 or not date_part.isdigit():
                        continue  # 형식 맞지 않으면 건너뜀

                    file_ym = date_part  # '202506' 형식
                except Exception as e:
                    print(f"[캐시] {file_path.name}에서 월 정보 추출 실패 → 건너뜀: {e}")
                    continue

                if file_ym in target_months:
                    valid_files.append(file_path)

            print(f"[캐시 필터링] 전체 {len(parquet_files)}개 중 대상 {len(valid_files)}개 파일 선정: {[f.name for f in valid_files]}")

            dfs = []
            for file_path in valid_files:
                try:
                    df_part = pd.read_parquet(file_path)
                    print(f"[캐시] {file_path.name} 로드 완료: {len(df_part):,} 건")
                    dfs.append(df_part)
                except Exception as e:
                    print(f"[캐시] {file_path.name} 읽기 실패: {e}")

            if dfs:
                df_cached_3months = pd.concat(dfs, ignore_index=True)
                print(f"[캐시] 총 {len(df_cached_3months):,} 건 데이터 병합 완료")
            else:
                print("[캐시] 모든 파일 로드 실패 → 3개월 데이터 없음")
        else:
            print("[캐시] data_cache에 DATA_LOT_3210_wafering_300_*.parquet 파일 없음")

        # ===================================================================
        # 2. [기존] self.data에서 당일 데이터 사용 (실시간 리포트용)
        # ===================================================================
        df_self_data = pd.DataFrame()
        if 'DATA_LOT_3210_wafering_300' in self.data and not self.data['DATA_LOT_3210_wafering_300'].empty:
            df_self_data = self.data['DATA_LOT_3210_wafering_300']
            print(f"[self.data] DATA_LOT_3210_wafering_300 데이터 건수: {len(df_self_data):,} 건")
        else:
            print("[self.data] DATA_LOT_3210_wafering_300 없거나 빈 데이터")


        # ===================================================================
        # [핵심] MS6 기반 PRODUCT_TYPE 병합
        # ===================================================================
        if not df_cached_3months.empty:
            df_cached_3months = self._merge_product_type(df_cached_3months)

        if not df_self_data.empty:
            df_self_data = self._merge_product_type(df_self_data)

        print(f"PRODUCT_TYPE 병합 완료: 3개월 {df_cached_3months['PRODUCT_TYPE'].notna().sum()}건, 당일 {df_self_data['PRODUCT_TYPE'].notna().sum()}건")

        # ===================================================================
        # 3. [핵심] 3개월 데이터 기반 Loss Rate 분석
        # ===================================================================
        if not df_cached_3months.empty:
            # 3개월 수량 합계 → 평균으로 변환 (3으로 나눔)
            total_months = 3

            # 분모: REJ_GROUP == "분모" 인 IN_QTY 합계
            denominator_data = df_cached_3months[df_cached_3months['REJ_GROUP'] == '분모']
            total_in_qty = denominator_data['IN_QTY'].sum() 
            avg_in_qty = total_in_qty / total_months  # 3개월 평균 전체 분모

            if avg_in_qty == 0:
                print(" 분모(IN_QTY)가 0입니다. Loss Rate 계산 불가")
                self.avg_in_qty = 0 # 인스턴스 변수에 0 저장
                self.total_daily_qty = 0 # 인스턴스 변수에 0 저장
                return details

            self.avg_in_qty = avg_in_qty # 인스턴스 변수에 저장 → WAF 분석에서 사용

            # ===================================================================
            #  1. 전체 (Total) CRET_CD별 Loss Rate
            # ===================================================================

            valid_cached = df_cached_3months[df_cached_3months['REJ_GROUP'].notna()]
            total_loss_by_cret = valid_cached.groupby('CRET_CD')['LOSS_QTY'].sum() / total_months #FS/HG/RESC 별 loss_qty 3개월 평균

            # ===================================================================
            #  2. 당일 CRET_CD별 LOSS_QTY
            # ===================================================================
            daily_loss_by_cret = pd.Series(dtype='int64')
            total_daily_qty = 0

            if not df_self_data.empty:
                valid_daily = df_self_data[df_self_data['REJ_GROUP'].notna()]
                daily_loss_by_cret = valid_daily.groupby('CRET_CD')['LOSS_QTY'].sum()

                denominator_daily = df_self_data[df_self_data['REJ_GROUP'] == '분모']
                total_daily_qty = denominator_daily['IN_QTY'].sum()  #  정의 추가
            else:
                print("[self.data] DATA_LOT_3210_wafering_300 없거나 빈 데이터")

            self.total_daily_qty = total_daily_qty # 인스턴스 변수에 저장

            # ===================================================================
            #  3.  전체 비교 표 생성 (모수 포함)
            # ===================================================================
            cret_list = ['FS', 'HG', 'RESC']
            report_table_total = []

            #  원시 데이터 저장용
            ref_qty_dict = {}
            daily_qty_dict = {}

            for cret_cd in cret_list:
                ref_qty = total_loss_by_cret.get(cret_cd, 0)
                daily_qty = daily_loss_by_cret.get(cret_cd, 0)

                ref_rate = (ref_qty / avg_in_qty) * 100 if avg_in_qty != 0 else 0
                daily_rate = (daily_qty / total_daily_qty) * 100 if avg_in_qty != 0 else 0
                gap = daily_rate - ref_rate

                report_table_total.append({
                    '구분': cret_cd,
                    'Ref.(3개월)': int(ref_qty),
                    '일': int(daily_qty),
                    'Ref.(3개월)%': f"{ref_rate:.2f}%",
                    '일%': f"{daily_rate:.2f}%",
                    'Gap': f"{gap:+.2f}%"
                })

                #  원시 데이터 저장
                ref_qty_dict[cret_cd] = int(ref_qty)
                daily_qty_dict[cret_cd] = int(daily_qty)

            #  모수 저장
            ref_qty_dict['모수'] = int(avg_in_qty) #3개월 평균 분모 -> ref 분모
            daily_qty_dict['모수'] = int(total_daily_qty) #일 분모


            report_table_total.append({
                '구분': '모수',
                'Ref.(3개월)': ref_qty_dict['모수'],
                '일': daily_qty_dict['모수'],
                'Ref.(3개월)%': "",
                '일%': "",
                'Gap': ""
            })

            #  details에 저장 (표 X, 값 O)
            details['rc_hg_ref_qty_total'] = ref_qty_dict
            details['rc_hg_daily_qty_total'] = daily_qty_dict
            details['rc_hg_avg_in_qty'] = avg_in_qty

            report_table_total_df = pd.DataFrame(report_table_total)
            details['summary'] = report_table_total_df

            # ===================================================================
            #  4. 그룹별 비교 표 생성 + 그래프 생성 (모수 제외)
            # ===================================================================
            rej_groups = ['PARTICLE', 'FLATNESS', 'WARP&BOW', 'NANO']
            details['rc_hg_ref_qty_by_group'] = {}
            details['rc_hg_daily_qty_by_group'] = {}
            details['rc_hg_gap_data_by_group'] = {}
            details['loss_rate_table_by_group'] = {}
            details['rc_hg_gap_chart_path_by_group'] = {}

            for group in rej_groups:
                # 각 그룹별 3개월 데이터 필터링
                group_data = df_cached_3months[df_cached_3months['REJ_GROUP'] == group]
                group_loss_by_cret = group_data.groupby('CRET_CD')['LOSS_QTY'].sum() / total_months
                # 각 그룹별 당일 데이터 필터링
                group_daily_loss_by_cret = pd.Series(dtype='int64')
                if not df_self_data.empty:
                    group_self_data = df_self_data[df_self_data['REJ_GROUP'] == group]
                    group_daily_loss_by_cret = group_self_data.groupby('CRET_CD')['LOSS_QTY'].sum()

                group_table = []
                gap_data = {}
                ref_qty_dict_group = {}
                daily_qty_dict_group = {}

                for cret_cd in cret_list:
                    ref_qty = group_loss_by_cret.get(cret_cd, 0)
                    daily_qty = group_daily_loss_by_cret.get(cret_cd, 0)

                    ref_rate = (ref_qty / avg_in_qty) * 100 if avg_in_qty != 0 else 0
                    daily_rate = (daily_qty / avg_in_qty) * 100 if avg_in_qty != 0 else 0
                    gap = daily_rate - ref_rate

                    group_table.append({
                        '구분': cret_cd,
                        'Ref.(3개월)': int(ref_qty),
                        '일': int(daily_qty),
                        'Ref.(3개월)%': f"{ref_rate:.2f}%",
                        '일%': f"{daily_rate:.2f}%",
                        'Gap': f"{gap:+.2f}%"
                    })

                    gap_data[cret_cd] = gap
                    ref_qty_dict_group[cret_cd] = int(ref_qty)
                    daily_qty_dict_group[cret_cd] = int(daily_qty)

                # 기존 방식과 동일하게 DataFrame으로 저장
                group_table_df = pd.DataFrame(group_table)
                if group_table_df.empty:
                    group_table_df = pd.DataFrame(columns=['구분', 'Ref.(3개월)', '일', 'Ref.(3개월)%', '일%', 'Gap'])
                details['loss_rate_table_by_group'][group] = group_table_df

                #  저장
                details['rc_hg_ref_qty_by_group'][group] = ref_qty_dict_group
                details['rc_hg_daily_qty_by_group'][group] = daily_qty_dict_group
                details['rc_hg_gap_data_by_group'][group] = gap_data  # 그래프용

                fig, ax = plt.subplots(figsize=(8, 4))

                categories = ['FS', 'HG', 'RESC']
                values = [float(gap_data.get(c, 0.0)) for c in categories]

                # 색상 설정: 양수=주황, 음수=파랑, 0=회색
                
                colors = ['orange' if v > 0 else 'steelblue' if v < 0 else 'gray' for v in values]

                # bar (수직 막대)
                bars = ax.bar(categories, values, color=colors, width=0.6)

                # 제목 및 라벨
                ax.set_title(f'RC/HG 보상({group})', fontsize=12, fontweight='bold')
                ax.set_ylabel('Gap (%)', fontsize=10)  # Y축이 Gap
                ax.set_xlabel('구분', fontsize=10)     # X축이 구분

                min_ylim = min(0, min(values) - 0.3)
                max_ylim = max(0, max(values) + 0.3)

                ax.set_ylim(min_ylim, max_ylim)        
                ax.set_ybound(min_ylim, max_ylim)      

                ax.grid(True, axis='y', linestyle='--', alpha=0.7)  # Y축 기준 그리드

                # 막대 위에 값 표시
                for bar, val in zip(bars, values):
                    height = bar.get_height()
                    if height >= 0:
                        y_pos = height + 0.005
                        va = 'bottom'
                    else:
                        y_pos = height - 0.005
                        va = 'top'
                    ax.text(bar.get_x() + bar.get_width() / 2, y_pos, f"{val:+.2f}%", ha='center', va=va, fontsize=9, fontweight='bold', color='black')

                plt.tight_layout()

                graph_path = debug_dir / f"RC_HG_보상_{group}.png"
                if graph_path.exists():
                    graph_path.unlink()
                    print(f"기존 그래프 파일 삭제됨: {graph_path}")

                plt.savefig(graph_path, dpi=150, bbox_inches='tight')
                plt.close()

                details['rc_hg_gap_chart_path_by_group'][group] = str(graph_path)

            # ===================================================================
            #  7. 전체 RC/HG 보상 그래프 생성
            # ===================================================================
            total_gap_data = {}
            for row in report_table_total:
                if row['구분'] in ['FS', 'HG', 'RESC']:
                    gap_str = row['Gap'].replace('%', '').replace('+', '')
                    total_gap_data[row['구분']] = float(gap_str)

            categories = ['FS', 'HG', 'RESC']
            values = [total_gap_data.get(c, 0.0) for c in categories]

            colors = ['orange' if total_gap_data.get(c, 0) > 0 else 
                    'steelblue' if total_gap_data.get(c, 0) < 0 else 'gray' for c in categories]

            fig, ax = plt.subplots(figsize=(8, 4))
            bars = ax.bar(categories, values, color=colors,  width=0.6)

            ax.set_title('RC/HG 보상(Ref.비 수준)', fontsize=12, fontweight='bold')
            ax.set_ylabel('Gap (%)', fontsize=10)
            ax.set_xlabel('구분', fontsize=10)
            
            min_ylim = min(0, min(values) - 0.3)
            max_ylim = max(0, max(values) + 0.3)

            ax.set_ylim(min_ylim, max_ylim)        
            ax.set_ybound(min_ylim, max_ylim)    

            ax.grid(True, axis='y', linestyle='--', alpha=0.7)

            for bar, val in zip(bars, values):
                height = bar.get_height()
                if height >= 0:
                    y_pos = height + 0.005
                    va = 'bottom'
                else:
                    y_pos = height - 0.005
                    va = 'top'
                ax.text(bar.get_x() + bar.get_width() / 2,  y_pos, f"{val:+.2f}%",  ha='center',  va=va, fontsize=9,  fontweight='bold', color='black')

            plt.tight_layout()
            total_graph_path = debug_dir / "RC_HG_보상_전체.png"
            if total_graph_path.exists():
                total_graph_path.unlink()
                print(f"기존 전체 그래프 파일 삭제됨: {total_graph_path}")

            plt.savefig(total_graph_path, dpi=150, bbox_inches='tight')
            plt.close()

            details['rc_hg_gap_chart_path_total'] = str(total_graph_path)

            # ===================================================================
            # 8. 기본 정보 추가
            # ===================================================================
            details['cache_data_available'] = not df_cached_3months.empty
            details['self_data_available'] = not df_self_data.empty
            details['cache_total_count'] = len(df_cached_3months) if not df_cached_3months.empty else 0
            details['self_data_count'] = len(df_self_data) if not df_self_data.empty else 0
            details['avg_in_qty'] = avg_in_qty

        else:
            # 빈 값 저장
            details['rc_hg_ref_qty_total'] = {}
            details['rc_hg_daily_qty_total'] = {}
            details['rc_hg_ref_qty_by_group'] = {}
            details['rc_hg_daily_qty_by_group'] = {}
            details['rc_hg_avg_in_qty'] = 0
            details['rc_hg_gap_chart_path_by_group'] = {}
            details['rc_hg_gap_chart_path_total'] = ""
            self.avg_in_qty = 0
            self.total_daily_qty = 0
        print(f"[LOT] self.avg_in_qty 저장 완료: {self.avg_in_qty}")
        print(f"[LOT] self.total_daily_qty 저장 완료: {self.total_daily_qty}")

        return details



    def _create_DATA_WAF_3210_wafering_300(self):
        """
        3210 WAF 상세 분석 - 캐시된 3개월 데이터 + self.data의 당일 데이터 모두 활용
        """
        details = {}
        PROJECT_ROOT = Path(__file__).parent.parent

        # ===================================================================
        # 1. [신규] data_cache에서 3개월 데이터 직접 로드 (장기 분석용)
        # ===================================================================
        base_date = (datetime.now().date() - timedelta(days=1))
        target_months = []
        current = base_date.replace(day=1)
        for _ in range(3):
            current = (current - timedelta(days=1)).replace(day=1)
            month_str = current.strftime("%Y%m")
            target_months.append(month_str)
        target_months = sorted(target_months)  # 과거 → 최근

        print(f"[WAF 캐시 필터링] 최근 3개월 대상 월: {target_months}")

        cache_dir = PROJECT_ROOT / "data_cache"
        pattern = "DATA_WAF_3210_wafering_300_*.parquet"
        parquet_files = list(cache_dir.glob(pattern))

        df_cached_3months = pd.DataFrame()

        if parquet_files:
            valid_files = []
            for file_path in parquet_files:
                try:
                    stem = file_path.stem
                    date_part = stem.split('_')[-1]  # '202506'

                    if len(date_part) != 6 or not date_part.isdigit():
                        continue

                    file_ym = date_part
                except Exception as e:
                    print(f"[WAF 캐시] {file_path.name}에서 월 정보 추출 실패 → 건너뜀: {e}")
                    continue

                if file_ym in target_months:
                    valid_files.append(file_path)

            print(f"[WAF 캐시 필터링] 전체 {len(parquet_files)}개 중 대상 {len(valid_files)}개 파일 선정: {[f.name for f in valid_files]}")

            dfs = []
            for file_path in valid_files:
                try:
                    df_part = pd.read_parquet(file_path)
                    print(f"[WAF 캐시] {file_path.name} 로드 완료: {len(df_part):,} 건")
                    dfs.append(df_part)
                except Exception as e:
                    print(f"[WAF 캐시] {file_path.name} 읽기 실패: {e}")

            if dfs:
                df_cached_3months = pd.concat(dfs, ignore_index=True)
                print(f"[WAF 캐시] 총 {len(df_cached_3months):,} 건 데이터 병합 완료")
            else:
                print("[WAF 캐시] 모든 파일 로드 실패 → 3개월 데이터 없음")
        else:
            print("[WAF 캐시] data_cache에 DATA_WAF_3210_wafering_300_*.parquet 파일 없음")

        # ===================================================================
        # 2. [기존] self.data에서 당일 데이터 사용
        # ===================================================================
        key = 'DATA_WAF_3210_wafering_300'
        df_self_data = pd.DataFrame()
        if key in self.data and not self.data[key].empty:
            df_self_data = self.data[key].copy()
            print(f"[self.data] {key} 데이터 건수: {len(df_self_data):,} 건")
        else:
            print(f"[self.data] {key} 없거나 빈 데이터")

        # ===================================================================
        # 3. [핵심] PRODUCT_TYPE 병합
        # ===================================================================
        if not df_cached_3months.empty:
            df_cached_3months = self._merge_product_type(df_cached_3months)

        if not df_self_data.empty:
            df_self_data = self._merge_product_type(df_self_data)

        print(f"[WAF] PRODUCT_TYPE 병합 완료: 3개월 {df_cached_3months['PRODUCT_TYPE'].notna().sum()}건, 당일 {df_self_data['PRODUCT_TYPE'].notna().sum()}건")


        # ===================================================================
        # 2. [기존] self.data에서 당일 데이터 사용 (실시간 리포트용)
        # ===================================================================
        df_self_data = pd.DataFrame()
        if 'DATA_WAF_3210_wafering_300' in self.data and not self.data['DATA_WAF_3210_wafering_300'].empty:
            df_self_data = self.data['DATA_WAF_3210_wafering_300']
            print(f"[self.data] DATA_WAF_3210_wafering_300 데이터 건수: {len(df_self_data):,} 건")
        else:
            print("[self.data] DATA_WAF_3210_wafering_300 없거나 빈 데이터")

        # ===================================================================
        # [핵심] MS6 기반 PRODUCT_TYPE 병합
        # ===================================================================
        if not df_cached_3months.empty:
            df_cached_3months = self._merge_product_type(df_cached_3months)

        if not df_self_data.empty:
            df_self_data = self._merge_product_type(df_self_data)

        print(f"PRODUCT_TYPE 병합 완료: 3개월 {df_cached_3months['PRODUCT_TYPE'].notna().sum()}건, 당일 {df_self_data['PRODUCT_TYPE'].notna().sum()}건")


        # ===================================================================
        # 3. [핵심] 3개월 데이터 기반 Loss Rate 분석
        # ===================================================================

        avg_in_qty = getattr(self, 'avg_in_qty', 0)
        total_daily_qty = getattr(self, 'total_daily_qty', 0)

        if avg_in_qty == 0:
            print(" 분모(IN_QTY)가 0입니다. Loss Rate 계산 불가")
            return details
        
        if total_daily_qty == 0:
            print("total_daily_qty = 0 → Daily 분석 생략 (Ref만 분석)")

        # 공통 함수: 장비별 LOSS_QTY 합계 → 불량률 계산
        def calculate_loss_rate(group_data, eqp_col, denominator):
            if eqp_col not in group_data.columns or denominator == 0:
                return {}
            valid = group_data.dropna(subset=[eqp_col])
            if valid.empty:
                return {}
            valid['LOSS_QTY'] = pd.to_numeric(valid['LOSS_QTY'], errors='coerce').fillna(0.0).astype(float)
            loss_sum = valid.groupby(eqp_col)['LOSS_QTY'].sum()
            return {eqp: round(qty / denominator * 100, 4) for eqp, qty in loss_sum.items()}

        # ===================================================================
        # [분석] Ref(3개월) 장비별 불량률 계산
        # ===================================================================
        ref_results = {}


        # 1) PIT
        df_pit = df_cached_3months[df_cached_3months['REJ_GROUP'] == 'PIT']
        if not df_pit.empty:
            rates = calculate_loss_rate(df_pit, 'EQP_NM_300_WF_3670', avg_in_qty)
            ref_results['PIT'] = rates

        # 2) SCRATCH
        df_scratch = df_cached_3months[df_cached_3months['REJ_GROUP'] == 'SCRATCH']
        if not df_scratch.empty:
            eqps_scratch = ['EQP_NM_300_WF_3670', 'EQP_NM_300_WF_6100']
            scratch_rates = {}
            for eqp in eqps_scratch:
                rates = calculate_loss_rate(df_scratch, eqp, avg_in_qty)
                if rates:
                    scratch_rates[eqp] = rates
            ref_results['SCRATCH'] = scratch_rates

        # 3) EDGE
        df_edge = df_cached_3months[df_cached_3months['REJ_GROUP'] == 'EDGE']
        if not df_edge.empty:
            eqps = ['EQP_NM_300_WF_3335', 'EQP_NM_300_WF_3696', 'EQP_NM_300_WF_7000']
            edge_rates = {}
            for eqp in eqps:
                rates = calculate_loss_rate(df_edge, eqp, avg_in_qty)
                if rates:
                    edge_rates[eqp] = rates
            ref_results['EDGE'] = edge_rates

        # 4) BROKEN
        df_broken = df_cached_3months[df_cached_3months['REJ_GROUP'] == 'BROKEN']
        if not df_broken.empty:
            eqps = ['EQP_NM_300_WF_3670', 'EQP_NM_300_WF_6100', 'EQP_NM_300_WF_6500']
            broken_rates = {}
            for eqp in eqps:
                rates = calculate_loss_rate(df_broken, eqp, avg_in_qty)
                if rates:
                    broken_rates[eqp] = rates
            ref_results['BROKEN'] = broken_rates

        # 5) CHIP
        df_chip = df_cached_3months[df_cached_3months['REJ_GROUP'] == 'CHIP']
        if not df_chip.empty:
            chip_rates = {}
            cond_edge = df_chip['AFT_BAD_RSN_CD'] == 'EDGE-CHIP'
            cond_lap = df_chip['AFT_BAD_RSN_CD'] == 'CHIP-LAP'
            cond_eg1af = df_chip['AFT_BAD_RSN_CD'] == 'CHIP_EG1AF'
            cond_eg1bf = df_chip['AFT_BAD_RSN_CD'] == 'CHIP_EG1BF'

            if not df_chip[cond_edge].empty:
                for eqp in ['EQP_NM_300_WF_3335', 'EQP_NM_300_WF_3696']:
                    rates = calculate_loss_rate(df_chip[cond_edge], eqp, avg_in_qty)
                    if rates:
                        chip_rates[f'EDGE-CHIP_{eqp}'] = rates
            if not df_chip[cond_lap].empty:
                rates = calculate_loss_rate(df_chip[cond_lap], 'EQP_NM_300_WF_3670', avg_in_qty)
                if rates:
                    chip_rates['CHIP-LAP_EQP_NM_300_WF_3670'] = rates
            if not df_chip[cond_eg1af].empty:
                for eqp in ['EQP_NM_300_WF_3335', 'EQP_NM_300_WF_3696']:
                    rates = calculate_loss_rate(df_chip[cond_eg1af], eqp, avg_in_qty)
                    if rates:
                        chip_rates[f'CHIP_EG1AF_{eqp}'] = rates
            if not df_chip[cond_eg1bf].empty:
                rates = calculate_loss_rate(df_chip[cond_eg1bf], 'EQP_NM_300_WF_3300', avg_in_qty)
                if rates:
                    chip_rates['CHIP_EG1BF_EQP_NM_300_WF_3300'] = rates
            ref_results['CHIP'] = chip_rates

        # 6) VISUAL
        df_visual = df_cached_3months[df_cached_3months['REJ_GROUP'] == 'VISUAL']
        if not df_visual.empty:
            cond = df_visual['AFT_BAD_RSN_CD'].isin(['B_PARTICLE', 'B_PAR2'])
            visual_filtered = df_visual[cond]
            if not visual_filtered.empty:
                rates = calculate_loss_rate(visual_filtered, 'EQP_NM_300_WF_6100', avg_in_qty)
                ref_results['VISUAL'] = rates

        # ===================================================================
        # [분석] Daily 장비별 불량률 계산
        # ===================================================================
        daily_results = {}

        if df_self_data.empty:
            print("당일 데이터 없음 → Daily 분석 건너뜀")
        else:
            if total_daily_qty == 0:
                print("당일 분모가 0 → Daily 분석 불가")
            else:
                df_pit_d = df_self_data[df_self_data['REJ_GROUP'] == 'PIT']
                if not df_pit_d.empty:
                    rates = calculate_loss_rate(df_pit_d, 'EQP_NM_300_WF_3670', total_daily_qty)
                    daily_results['PIT'] = rates

                # 2) SCRATCH
                df_scratch_d = df_self_data[df_self_data['REJ_GROUP'] == 'SCRATCH']
                if not df_scratch_d.empty:
                    eqps_scratch = ['EQP_NM_300_WF_3670', 'EQP_NM_300_WF_6100']
                    scratch_rates = {}
                    for eqp in eqps_scratch:
                        rates = calculate_loss_rate(df_scratch_d, eqp, avg_in_qty)
                        if rates:
                            scratch_rates[eqp] = rates
                    daily_results['SCRATCH'] = scratch_rates

                # 3) EDGE
                df_edge_d = df_self_data[df_self_data['REJ_GROUP'] == 'EDGE']
                if not df_edge_d.empty:
                    eqps = ['EQP_NM_300_WF_3335', 'EQP_NM_300_WF_3696', 'EQP_NM_300_WF_7000']
                    edge_rates = {}
                    for eqp in eqps:
                        rates = calculate_loss_rate(df_edge_d, eqp, total_daily_qty)
                        if rates:
                            edge_rates[eqp] = rates
                    daily_results['EDGE'] = edge_rates

                # 4) BROKEN
                df_broken_d = df_self_data[df_self_data['REJ_GROUP'] == 'BROKEN']
                if not df_broken_d.empty:
                    eqps = ['EQP_NM_300_WF_3670', 'EQP_NM_300_WF_6100', 'EQP_NM_300_WF_6500']
                    broken_rates = {}
                    for eqp in eqps:
                        rates = calculate_loss_rate(df_broken_d, eqp, total_daily_qty)
                        if rates:
                            broken_rates[eqp] = rates
                    daily_results['BROKEN'] = broken_rates

                # 5) CHIP
                df_chip_d = df_self_data[df_self_data['REJ_GROUP'] == 'CHIP']
                if not df_chip_d.empty:
                    chip_rates = {}
                    cond_edge = df_chip_d['AFT_BAD_RSN_CD'] == 'EDGE-CHIP'
                    cond_lap = df_chip_d['AFT_BAD_RSN_CD'] == 'CHIP-LAP'
                    cond_eg1af = df_chip_d['AFT_BAD_RSN_CD'] == 'CHIP_EG1AF'
                    cond_eg1bf = df_chip_d['AFT_BAD_RSN_CD'] == 'CHIP_EG1BF'

                    if not df_chip_d[cond_edge].empty:
                        for eqp in ['EQP_NM_300_WF_3335', 'EQP_NM_300_WF_3696']:
                            rates = calculate_loss_rate(df_chip_d[cond_edge], eqp, total_daily_qty)
                            if rates:
                                chip_rates[f'EDGE-CHIP_{eqp}'] = rates
                    if not df_chip_d[cond_lap].empty:
                        rates = calculate_loss_rate(df_chip_d[cond_lap], 'EQP_NM_300_WF_3670', total_daily_qty)
                        if rates:
                            chip_rates['CHIP-LAP_EQP_NM_300_WF_3670'] = rates
                    if not df_chip_d[cond_eg1af].empty:
                        for eqp in ['EQP_NM_300_WF_3335', 'EQP_NM_300_WF_3696']:
                            rates = calculate_loss_rate(df_chip_d[cond_eg1af], eqp, total_daily_qty)
                            if rates:
                                chip_rates[f'CHIP_EG1AF_{eqp}'] = rates
                    if not df_chip_d[cond_eg1bf].empty:
                        rates = calculate_loss_rate(df_chip_d[cond_eg1bf], 'EQP_NM_300_WF_3300', total_daily_qty)
                        if rates:
                            chip_rates['CHIP_EG1BF_EQP_NM_300_WF_3300'] = rates
                    daily_results['CHIP'] = chip_rates

                # 6) VISUAL
                df_visual_d = df_self_data[df_self_data['REJ_GROUP'] == 'VISUAL']
                if not df_visual_d.empty:
                    cond = df_visual_d['AFT_BAD_RSN_CD'].isin(['B_PARTICLE', 'B_PAR2'])
                    visual_filtered = df_visual_d[cond]
                    if not visual_filtered.empty:
                        rates = calculate_loss_rate(visual_filtered, 'EQP_NM_300_WF_6100', avg_in_qty)
                        daily_results['VISUAL'] = rates

        # ===================================================================
        # 8. Gap 계산 (각 공정별 상위 3개 장비만)
        # ===================================================================
        gap_results = {}

        # 공정명 추출 함수 (예: 'EQP_NM_300_WF_3670' → '3670')
        def extract_process(eqp_col):
            import re
            match = re.search(r'(\d{4})$', eqp_col)
            return match.group(1) if match else eqp_col

        for group, ref_dict in ref_results.items():
            if group not in daily_results:
                continue

            daily_dict = daily_results[group]

            # 단일 dict (PIT, VISUAL 등)
            if not isinstance(next(iter(ref_dict.values()), {}), dict):
                # 장비-불량률 리스트 생성
                all_items = [(k, ref_dict.get(k, 0)) for k in set(ref_dict.keys()) | set(daily_dict.keys())]
                # 불량률 기준 상위 3개
                top3_keys = sorted(all_items, key=lambda x: x[1], reverse=True)[:3]
                top3_keys = [k for k, v in top3_keys]

                gap_dict = {k: round(float(daily_dict.get(k, 0)) - float(ref_dict.get(k, 0)), 4)
                    for k in top3_keys}
                gap_results[group] = gap_dict

            else:
                # 중첩 dict (SCRATCH, EDGE 등)
                gap_sub = {}
                for eqp_col, rates in ref_dict.items():
                    if eqp_col not in daily_dict:
                        continue
                    daily_rates = daily_dict[eqp_col]

                    # 장비-불량률 리스트 생성
                    all_items = [(k, rates.get(k, 0)) for k in set(rates.keys()) | set(daily_rates.keys())]
                    # 불량률 기준 상위 3개
                    top3_keys = sorted(all_items, key=lambda x: x[1], reverse=True)[:3]
                    top3_keys = [k for k, v in top3_keys]

                gap_sub[eqp_col] = {
                    k: round(
                        float(daily_rates.get(k, 0)) - float(rates.get(k, 0)),
                        4
                    )
                    for k in top3_keys
                }
                gap_results[group] = gap_sub

        # ===================================================================
        # 9. details에 저장
        # ===================================================================
        details['waf_analysis_ref'] = ref_results
        details['waf_analysis_daily'] = daily_results
        details['waf_analysis_gap'] = gap_results
        details['df_cached_3months'] = df_cached_3months
        details['df_self_data'] = df_self_data
        details['avg_in_qty'] = avg_in_qty
        details['total_daily_qty'] = total_daily_qty

        return details


    def _export_to_excel(self, report, output_dir="./daily_reports_debug"):
        """Excel 보고서 생성 (Stage4: SheetBuilder 기반으로 전체 정리)"""
        try:
            PROJECT_ROOT = Path(__file__).parent.parent
            base_date = (datetime.now().date() - timedelta(days=1))
            date_folder_name = base_date.strftime("%Y%m%d")
            debug_dir = PROJECT_ROOT / output_dir / date_folder_name
            debug_dir.mkdir(parents=True, exist_ok=True)

            excel_path = debug_dir / f"Daily_Report_{date_folder_name}.xlsx"

            # 기존 파일 삭제(열려있을 경우 PermissionError 발생 가능)
            if excel_path.exists():
                try:
                    excel_path.unlink()
                except PermissionError:
                    raise PermissionError(f"엑셀 파일이 열려있습니다: {excel_path}")

            wb = Workbook()

            # -------------------------------------------------
            # Sheet 1: Prime 분석 (메인)
            # -------------------------------------------------
            ws = wb.active
            ws.title = "Prime 분석"
            builder = SheetBuilder(ws, start_row=1, start_col=1)

            # ===== 0) 타이틀 =====
            builder.title(f"Daily Report - {date_folder_name}", merge_from="A1", merge_to="L1")
            builder.blank(1)

            # -------------------------------------------------
            # 1) 3010 수율
            # -------------------------------------------------
            builder.title("[ 3010 WF RTY 수율 ]", merge_from=f"A{builder.row}", merge_to=f"L{builder.row}")
            builder.blank(1)

            data_3010 = report.get("DATA_3010_wafering_300") or {}
            chart_path_3010 = data_3010.get("chart_path")
            table_df_3010 = data_3010.get("table_df")

            # 차트 (있으면)
            if chart_path_3010:
                try:
                    builder.image_from_path(chart_path_3010, f"A{builder.row}", width=650, height=320)
                except Exception as e:
                    ws.cell(builder.row, 1, f"3010 차트 삽입 실패: {e}").font = Font(color="FF0000")
            # 표는 차트 오른쪽에 배치 (대략적)
            if table_df_3010 is not None and not getattr(table_df_3010, "empty", True):
                # 퍼센트 컬럼은 0~1로 변환 후 표시
                table_df_3010_fmt = table_df_3010.copy()
                pct_cols = ["월 목표", "월 실적", "일 목표", "일 실적", "Gap(월)", "Gap(일)"]
                for col in pct_cols:
                    if col in table_df_3010_fmt.columns:
                        table_df_3010_fmt[col] = pd.to_numeric(table_df_3010_fmt[col], errors="coerce") / 100.0
                number_formats = {c: "0.00%" for c in pct_cols if c in table_df_3010_fmt.columns}

                write_df_table(ws, table_df_3010_fmt, start_row=builder.row, start_col=9, number_formats=number_formats)

            # 다음 섹션으로 내려가기(차트 높이 감안)
            builder.row += 18
            builder.blank(1)

            # -------------------------------------------------
            # 2) Prime 불량률 (3210) 요약/차트
            # -------------------------------------------------
            builder.title("[ Prime 불량률 분석 (3210) ]", merge_from=f"A{builder.row}", merge_to=f"L{builder.row}")
            builder.blank(1)

            data_3210_details = report.get("DATA_3210_wafering_300_details") or {}
            prime_summary = data_3210_details.get("summary")
            prime_chart_path = data_3210_details.get("chart_path")
            top3_rej_groups = data_3210_details.get("top3_rej_groups") or []

            # 차트
            if prime_chart_path:
                try:
                    builder.image_from_path(prime_chart_path, f"A{builder.row}", width=650, height=340)
                except Exception as e:
                    ws.cell(builder.row, 1, f"Prime 차트 삽입 실패: {e}").font = Font(color="FF0000")

            # 표(오른쪽)
            if prime_summary is not None and not getattr(prime_summary, "empty", True):
                write_df_table(ws, prime_summary, start_row=builder.row, start_col=9)

            builder.row += 20
            builder.blank(1)

            # -------------------------------------------------
            # 3) 상위 3개 불량 세부분석 텍스트
            # -------------------------------------------------
            builder.title("[ 상위 3개 불량 세부분석 ]", merge_from=f"A{builder.row}", merge_to=f"L{builder.row}")
            builder.blank(1)

            detailed_lines = data_3210_details.get("detailed_analysis") or []
            if detailed_lines:
                for line in detailed_lines:
                    ws.cell(builder.row, 1, str(line)).alignment = LEFT
                    builder.row += 1
            else:
                ws.cell(builder.row, 1, "세부분석 결과 없음").alignment = LEFT
                builder.row += 1

            builder.blank(2)

            # -------------------------------------------------
            # 4) Prime 주요 열위 MID_GROUP (3개월 Ref vs 일실적)
            # -------------------------------------------------
            builder.title("[ Prime 주요 열위 불량 세부코드 (Ref 3개월 vs 일실적) ]", merge_from=f"A{builder.row}", merge_to=f"L{builder.row}")
            builder.blank(1)

            data_3210_3m = report.get("DATA_3210_wafering_300_3months") or {}
            top3_mid = data_3210_3m.get("top3_midgroup_analysis") if isinstance(data_3210_3m, dict) else None

            if top3_mid and isinstance(top3_mid, dict):
                # 분석 텍스트
                analysis_text = top3_mid.get("analysis")
                if analysis_text:
                    for line in str(analysis_text).split("\n"):
                        ws.cell(builder.row, 1, line).alignment = LEFT
                        builder.row += 1
                    builder.blank(1)

                # 플롯들 (그룹별)
                plot_paths = top3_mid.get("plot_paths") or {}
                if plot_paths:
                    for rej, pth in plot_paths.items():
                        builder.title(f"- {rej}", merge_from=f"A{builder.row}", merge_to=f"L{builder.row}", font=Font(size=12, bold=True))
                        builder.blank(1)
                        try:
                            builder.image_from_path(pth, f"A{builder.row}", width=650, height=320)
                        except Exception as e:
                            ws.cell(builder.row, 1, f"{rej} 플롯 삽입 실패: {e}").font = Font(color="FF0000")
                        builder.row += 18
                        builder.blank(1)

                # 표들 (그룹별)
                tables = top3_mid.get("tables") or {}
                if tables:
                    builder.title("[ 그룹별 Top3 표 ]", merge_from=f"A{builder.row}", merge_to=f"L{builder.row}", font=Font(size=12, bold=True))
                    builder.blank(1)
                    for rej, df_tbl in tables.items():
                        ws.cell(builder.row, 1, f"{rej}").font = Font(bold=True)
                        builder.row += 1
                        if df_tbl is not None and not getattr(df_tbl, "empty", True):
                            builder.table(df_tbl, start_col=1)
                        else:
                            ws.cell(builder.row, 1, "표 없음")
                            builder.row += 1
                        builder.blank(1)
            else:
                ws.cell(builder.row, 1, "3개월 Ref 비교 데이터 없음").alignment = LEFT
                builder.row += 1

            builder.blank(2)

            # -------------------------------------------------
            # 5) 제품 영향성 GAP (Ref vs Daily)
            # -------------------------------------------------
            builder.title("[ 제품 영향성 GAP (Ref 6개월 vs Daily) ]", merge_from=f"A{builder.row}", merge_to=f"L{builder.row}")
            builder.blank(1)

            product_gap = report.get("product_influence_gap")
            if product_gap is not None and not getattr(product_gap, "empty", True):
                builder.table(product_gap, start_col=1)
            else:
                ws.cell(builder.row, 1, "제품 영향성 GAP 데이터 없음").alignment = LEFT
                builder.row += 1

            builder.blank(2)

            # -------------------------------------------------
            # 6) LOT / WAF 상세 요약 (요약표/차트 경로가 있다면 자동 삽입)
            # -------------------------------------------------
            builder.title("[ LOT / WAF 상세 요약 ]", merge_from=f"A{builder.row}", merge_to=f"L{builder.row}")
            builder.blank(1)

            lot = report.get("DATA_LOT_3210_wafering_300_details") or {}
            waf = report.get("DATA_WAF_3210_wafering_300_details") or {}

            # LOT summary table
            lot_summary = lot.get("summary") if isinstance(lot, dict) else None
            if lot_summary is not None and not getattr(lot_summary, "empty", True):
                ws.cell(builder.row, 1, "LOT 요약").font = Font(bold=True)
                builder.row += 1
                builder.table(lot_summary, start_col=1)
                builder.blank(1)

            # LOT charts by group
            lot_charts = lot.get("rc_hg_gap_chart_path_by_group") if isinstance(lot, dict) else {}
            if lot_charts:
                ws.cell(builder.row, 1, "LOT 차트(그룹별)").font = Font(bold=True)
                builder.row += 1
                for group, pth in lot_charts.items():
                    ws.cell(builder.row, 1, f"- {group}").font = Font(bold=True)
                    builder.row += 1
                    try:
                        builder.image_from_path(pth, f"A{builder.row}", width=650, height=300)
                    except Exception as e:
                        ws.cell(builder.row, 1, f"LOT {group} 차트 삽입 실패: {e}").font = Font(color="FF0000")
                    builder.row += 17
                    builder.blank(1)

            lot_total_chart = lot.get("rc_hg_gap_chart_path_total") if isinstance(lot, dict) else None
            if lot_total_chart:
                ws.cell(builder.row, 1, "LOT 전체 차트").font = Font(bold=True)
                builder.row += 1
                try:
                    builder.image_from_path(lot_total_chart, f"A{builder.row}", width=650, height=300)
                except Exception as e:
                    ws.cell(builder.row, 1, f"LOT 전체 차트 삽입 실패: {e}").font = Font(color="FF0000")
                builder.row += 17
                builder.blank(1)

            # WAF summary if present
            waf_summary = waf.get("summary") if isinstance(waf, dict) else None
            if waf_summary is not None and not getattr(waf_summary, "empty", True):
                ws.cell(builder.row, 1, "WAF 요약").font = Font(bold=True)
                builder.row += 1
                builder.table(waf_summary, start_col=1)
                builder.blank(1)

            # (선택) WAF 차트 경로가 dict로 있다면 자동 삽입
            waf_charts = waf.get("charts") if isinstance(waf, dict) else {}
            if isinstance(waf_charts, dict) and waf_charts:
                ws.cell(builder.row, 1, "WAF 차트").font = Font(bold=True)
                builder.row += 1
                for name, pth in waf_charts.items():
                    ws.cell(builder.row, 1, f"- {name}").font = Font(bold=True)
                    builder.row += 1
                    try:
                        builder.image_from_path(pth, f"A{builder.row}", width=650, height=300)
                    except Exception as e:
                        ws.cell(builder.row, 1, f"WAF {name} 차트 삽입 실패: {e}").font = Font(color="FF0000")
                    builder.row += 17
                    builder.blank(1)

            # 보기 편하게 열 너비 기본값 설정
            default_widths = {
                "A": 18, "B": 16, "C": 16, "D": 16, "E": 16, "F": 16,
                "G": 16, "H": 16, "I": 16, "J": 16, "K": 16, "L": 16
            }
            for col, w in default_widths.items():
                ws.column_dimensions[col].width = w

            wb.save(excel_path)
            return excel_path

        except Exception as e:
            logger.error(f"Excel 보고서 생성 실패: {e}")
            raise

